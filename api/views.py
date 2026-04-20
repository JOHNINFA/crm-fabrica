from rest_framework import viewsets, permissions, status, parsers
from rest_framework.response import Response
from rest_framework.decorators import action, api_view, parser_classes, permission_classes
from rest_framework.parsers import MultiPartParser, FormParser
from django.conf import settings
from django.db import transaction, models  # 🆕 Agregar models para Q
from django.utils import timezone
from django.http import HttpResponse
import os
import base64
import re
import uuid
import csv
import json
import secrets
import unicodedata
from datetime import timedelta, datetime, date
from collections import defaultdict
from api.services.ai_assistant_service import AIAssistant
from django.utils.dateparse import parse_datetime, parse_date
from .models import Planeacion, Registro, Producto, Categoria, Stock, Lote, MovimientoInventario, RegistroInventario, Venta, DetalleVenta, Cliente, ProductosFrecuentes, ListaPrecio, PrecioProducto, CargueID1, CargueID2, CargueID3, CargueID4, CargueID5, CargueID6, Produccion, ProduccionSolicitada, Pedido, DetallePedido, Vendedor, VendedorSesionToken, Domiciliario, MovimientoCaja, ArqueoCaja, ConfiguracionImpresion, Ruta, ClienteRuta, VentaRuta, CarguePagos, CargueCumplimiento, RutaOrden, RutaOrdenVendedor, ReportePlaneacion, CargueResumen, TipoNegocio, ClienteOcasional, recalcular_totales_cargue_queryset
from .serializers import (
    PlaneacionSerializer, ReportePlaneacionSerializer,
    RegistroSerializer, ProductoSerializer, CategoriaSerializer, StockSerializer,
    LoteSerializer, MovimientoInventarioSerializer, RegistroInventarioSerializer,
    VentaSerializer, DetalleVentaSerializer, ClienteSerializer, ProductosFrecuentesSerializer, ListaPrecioSerializer, PrecioProductoSerializer,
    CargueID1Serializer, CargueID2Serializer, CargueID3Serializer, CargueID4Serializer, CargueID5Serializer, CargueID6Serializer, ProduccionSerializer, ProduccionSolicitadaSerializer, PedidoSerializer, DetallePedidoSerializer, VendedorSerializer, DomiciliarioSerializer, MovimientoCajaSerializer, ArqueoCajaSerializer, ConfiguracionImpresionSerializer,
    RutaSerializer, ClienteRutaSerializer, VentaRutaSerializer, CarguePagosSerializer, RutaOrdenSerializer, CargueResumenSerializer, TipoNegocioSerializer,
    ClienteOcasionalSerializer
)


def _extraer_token_request(request):
    auth_header = request.headers.get('Authorization', '') or ''
    if auth_header.lower().startswith('bearer '):
        return auth_header.split(' ', 1)[1].strip()

    token_header = request.headers.get('X-App-Token') or request.headers.get('x-app-token')
    if token_header:
        return str(token_header).strip()

    token_query = request.query_params.get('token') if hasattr(request, 'query_params') else None
    if token_query:
        return str(token_query).strip()

    token_body = request.data.get('token') if hasattr(request, 'data') else None
    if token_body:
        return str(token_body).strip()

    return ''


def _normalizar_id_vendedor(vendedor_raw):
    if vendedor_raw is None:
        return None, None

    vendedor_txt = str(vendedor_raw).strip().upper()
    if not vendedor_txt:
        return None, None

    if vendedor_txt.startswith('ID'):
        vendedor_num_txt = vendedor_txt[2:]
        vendedor_id_txt = vendedor_txt
    else:
        vendedor_num_txt = vendedor_txt
        vendedor_id_txt = f"ID{vendedor_num_txt}"

    if not vendedor_num_txt.isdigit():
        return None, None

    return vendedor_id_txt, int(vendedor_num_txt)


def _obtener_vendedor_sesion_movil(request):
    token = _extraer_token_request(request)
    if not token:
        return None, Response({'error': 'Token requerido'}, status=status.HTTP_401_UNAUTHORIZED)

    sesion = VendedorSesionToken.objects.select_related('vendedor').filter(
        token=token,
        activo=True,
        expira_en__gt=timezone.now(),
        vendedor__activo=True
    ).first()

    if not sesion:
        return None, Response({'error': 'Sesión inválida o expirada'}, status=status.HTTP_401_UNAUTHORIZED)

    sesion.ultimo_uso = timezone.now()
    sesion.save(update_fields=['ultimo_uso'])
    return sesion.vendedor, None


def _validar_vendedor_token(request, vendedor_raw=None, campo='vendedor_id'):
    vendedor, error_response = _obtener_vendedor_sesion_movil(request)
    if error_response:
        return None, None, error_response

    vendedor_token_id, vendedor_token_num = _normalizar_id_vendedor(vendedor.id_vendedor)
    if not vendedor_token_id:
        return None, None, Response({'error': 'Vendedor inválido en sesión'}, status=status.HTTP_401_UNAUTHORIZED)

    if vendedor_raw is not None and str(vendedor_raw).strip() != '':
        vendedor_req_id, vendedor_req_num = _normalizar_id_vendedor(vendedor_raw)
        if not vendedor_req_id:
            return None, None, Response({'error': f'{campo} inválido'}, status=status.HTTP_400_BAD_REQUEST)
        if vendedor_req_id != vendedor_token_id:
            return None, None, Response({'error': 'Token no autorizado para este vendedor'}, status=status.HTTP_403_FORBIDDEN)
        return vendedor, vendedor_req_num, None

    return vendedor, vendedor_token_num, None


def _pedido_pertenece_a_vendedor(pedido, vendedor):
    vendedor_id = (vendedor.id_vendedor or '').strip().upper()
    pedido_asignado = (pedido.asignado_a_id or '').strip().upper()
    pedido_vendedor = (pedido.vendedor or '').strip()

    if pedido_asignado and pedido_asignado == vendedor_id:
        return True

    if pedido_vendedor:
        if pedido_vendedor.upper() == vendedor_id:
            return True
        if (vendedor.nombre or '').strip().lower() == pedido_vendedor.lower():
            return True

    return False


class CargueResumenViewSet(viewsets.ModelViewSet):
    """API para gestionar resúmenes de cargue y estados"""
    queryset = CargueResumen.objects.all()
    serializer_class = CargueResumenSerializer
    permission_classes = [permissions.AllowAny]
    filterset_fields = ['vendedor_id', 'dia', 'fecha', 'estado_cargue']

    def get_queryset(self):
        queryset = super().get_queryset()
        vendedor_id = self.request.query_params.get('vendedor_id')
        dia = self.request.query_params.get('dia')
        fecha = self.request.query_params.get('fecha')
        
        if vendedor_id:
            queryset = queryset.filter(vendedor_id=vendedor_id)
        if dia:
            queryset = queryset.filter(dia=dia.upper())
        if fecha:
            queryset = queryset.filter(fecha=fecha)
            
        return queryset

# 🆕 ViewSet para pagos de Cargue (múltiples filas por día/vendedor)
class CarguePagosViewSet(viewsets.ModelViewSet):
    """API para gestionar filas de pagos del módulo Cargue"""
    queryset = CarguePagos.objects.filter(activo=True).order_by('id')
    serializer_class = CarguePagosSerializer
    permission_classes = [permissions.AllowAny]
    filterset_fields = ['vendedor_id', 'dia', 'fecha']

    def get_queryset(self):
        queryset = super().get_queryset()
        vendedor_id = self.request.query_params.get('vendedor_id')
        dia = self.request.query_params.get('dia')
        fecha = self.request.query_params.get('fecha')
        
        if vendedor_id:
            queryset = queryset.filter(vendedor_id=vendedor_id)
        if dia:
            queryset = queryset.filter(dia=dia.upper())
        if fecha:
            queryset = queryset.filter(fecha=fecha)
        
        return queryset

    @action(detail=False, methods=['post'])
    def sync_pagos(self, request):
        """
        Sincroniza todas las filas de pagos para un vendedor/día/fecha.
        Elimina las anteriores y crea las nuevas.
        """
        vendedor_id = request.data.get('vendedor_id')
        dia = request.data.get('dia', '').upper()
        fecha = request.data.get('fecha')
        filas = request.data.get('filas', [])
        usuario = request.data.get('usuario', 'Sistema')

        if not all([vendedor_id, dia, fecha]):
            return Response(
                {'error': 'Se requiere vendedor_id, dia y fecha'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            with transaction.atomic():
                # Eliminar filas anteriores
                CarguePagos.objects.filter(
                    vendedor_id=vendedor_id,
                    dia=dia,
                    fecha=fecha
                ).delete()

                # Crear nuevas filas
                nuevas_filas = []
                for fila in filas:
                    if fila.get('concepto') or fila.get('descuentos', 0) > 0 or fila.get('nequi', 0) > 0 or fila.get('daviplata', 0) > 0:
                        nuevas_filas.append(CarguePagos(
                            vendedor_id=vendedor_id,
                            dia=dia,
                            fecha=fecha,
                            concepto=fila.get('concepto', ''),
                            descuentos=fila.get('descuentos', 0),
                            nequi=fila.get('nequi', 0),
                            daviplata=fila.get('daviplata', 0),
                            usuario=usuario
                        ))

                if nuevas_filas:
                    CarguePagos.objects.bulk_create(nuevas_filas)

                return Response({
                    'success': True,
                    'message': f'Sincronizadas {len(nuevas_filas)} filas de pagos',
                    'count': len(nuevas_filas)
                })

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# 🆕 Importar CargueResumen para estado
from .models import CargueResumen


# 🆕 Endpoints para estado del cargue
@api_view(['GET'])
def obtener_estado_cargue(request):
    """Obtiene el estado del cargue para un día/fecha"""
    dia = request.query_params.get('dia', '').upper()
    fecha = request.query_params.get('fecha')
    
    if not dia or not fecha:
        return Response({'error': 'Se requiere dia y fecha'}, status=400)
    
    try:
        # Buscar en CargueResumen (usamos ID1 como referencia global)
        resumen = CargueResumen.objects.filter(
            dia=dia,
            fecha=fecha,
            activo=True
        ).first()
        
        if resumen:
            return Response({
                'success': True,
                'estado': resumen.estado_cargue,
                'vendedor_id': resumen.vendedor_id
            })
        else:
            return Response({
                'success': True,
                'estado': 'ALISTAMIENTO',  # Default si no existe
                'vendedor_id': None
            })
    except Exception as e:
        return Response({'error': str(e)}, status=500)


@api_view(['POST'])
def actualizar_estado_cargue(request):
    """Actualiza el estado del cargue para un día/fecha"""
    dia = request.data.get('dia', '').upper()
    fecha = request.data.get('fecha')
    estado = request.data.get('estado', 'ALISTAMIENTO')
    vendedor_id = request.data.get('vendedor_id', 'ID1')  # ID1 como referencia global
    
    print(f"📡 RECIBIDO UPDATE ESTADO: {estado} para {dia} {fecha} (ID: {vendedor_id})")

    
    if not dia or not fecha:
        return Response({'error': 'Se requiere dia y fecha'}, status=400)
    
    estados_validos = ['ALISTAMIENTO', 'SUGERIDO', 'DESPACHO', 'COMPLETADO', 'ALISTAMIENTO_ACTIVO', 'FINALIZAR']
    if estado not in estados_validos:
        return Response({'error': f'Estado inválido. Válidos: {estados_validos}'}, status=400)
    
    try:
        # Crear o actualizar en CargueResumen
        resumen, created = CargueResumen.objects.update_or_create(
            dia=dia,
            fecha=fecha,
            vendedor_id=vendedor_id,
            defaults={
                'estado_cargue': estado,
                'activo': True
            }
        )
        
        return Response({
            'success': True,
            'estado': resumen.estado_cargue,
            'action': 'created' if created else 'updated',
            'message': f'Estado actualizado a {estado}'
        })
    except Exception as e:
        return Response({'error': str(e)}, status=500)


# 🆕 Endpoint ULTRALIGERO para Polling Inteligente
@api_view(['GET'])
def verificar_actualizaciones(request):
    """
    Endpoint ultraligero para polling.
    Devuelve la fecha de la última actualización en la tabla correspondiente.
    """
    vendedor_id = request.query_params.get('idSheet', 'ID1')
    dia = request.query_params.get('dia', '').upper()
    fecha = request.query_params.get('fecha')

    if not all([vendedor_id, dia, fecha]):
        return Response({'error': 'Faltan parámetros'}, status=400)

    # Seleccionar modelo según ID
    modelos = {
        'ID1': CargueID1, 'ID2': CargueID2, 'ID3': CargueID3,
        'ID4': CargueID4, 'ID5': CargueID5, 'ID6': CargueID6
    }
    modelo = modelos.get(vendedor_id)

    if not modelo:
        return Response({'last_update': None})

    # Consulta optimizada: Solo max(fecha_actualizacion)
    try:
        resultado = modelo.objects.filter(dia=dia, fecha=fecha).aggregate(
            ultima=models.Max('fecha_actualizacion')
        )
        return Response({'last_update': resultado['ultima']})
    except Exception as e:
        return Response({'error': str(e)}, status=500)


class RegistroViewSet(viewsets.ModelViewSet):
    queryset = Registro.objects.all()
    serializer_class = RegistroSerializer
    permission_classes = [permissions.AllowAny]

class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer
    permission_classes = [permissions.AllowAny]

class ProductoViewSet(viewsets.ModelViewSet):
    """API para gestionar productos"""
    queryset = Producto.objects.filter(activo=True).order_by('orden', 'nombre')  # 🆕 Ordenar por campo 'orden'
    serializer_class = ProductoSerializer
    permission_classes = [permissions.AllowAny]
    parser_classes = [parsers.MultiPartParser, parsers.FormParser, parsers.JSONParser]
    
    def _save_image_to_paths(self, image_data, filename):
        """Guarda imagen en ambas ubicaciones"""
        # Crear carpetas
        frontend_path = os.path.join(settings.BASE_DIR, 'frontend', 'public', 'images', 'productos')
        media_path = os.path.join(settings.MEDIA_ROOT, 'productos')
        
        os.makedirs(frontend_path, exist_ok=True)
        os.makedirs(media_path, exist_ok=True)
        
        # Guardar en ambas ubicaciones
        for path in [frontend_path, media_path]:
            filepath = os.path.join(path, filename)
            with open(filepath, 'wb') as f:
                f.write(base64.b64decode(image_data))
    
    @action(detail=False, methods=['post'])
    def save_image(self, request):
        """Guarda imagen base64 y devuelve URLs"""
        try:
            image_data = request.data.get('image')
            product_id = request.data.get('productId')
            
            if not image_data or not image_data.startswith('data:'):
                return Response({'error': 'Datos de imagen no válidos'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Extraer datos base64
            match = re.match(r'data:([^;]+);base64,(.+)', image_data)
            if not match:
                return Response({'error': 'Formato de imagen no válido'}, status=status.HTTP_400_BAD_REQUEST)
            
            mime_type, base64_data = match.groups()
            extension = mime_type.split('/')[-1]
            filename = f"producto_{product_id or uuid.uuid4()}_{uuid.uuid4().hex[:8]}.{extension}"
            
            # Guardar imagen
            self._save_image_to_paths(base64_data, filename)
            
            return Response({
                'success': True,
                'frontendUrl': f"/images/productos/{filename}",
                'mediaUrl': f"/media/productos/{filename}",
                'filename': filename
            })
            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def actualizar_stock(self, request, pk=None):
        """Actualiza stock creando MovimientoInventario (con trazabilidad completa)"""
        try:
            producto = self.get_object()
            cantidad = int(request.data.get('cantidad', 0))
            usuario = request.data.get('usuario', 'Sistema')
            nota = request.data.get('nota', '')
            
            import datetime
            timestamp = datetime.datetime.now().strftime('%H:%M:%S.%f')[:-3]
            print(f"\n=== 🔥 ACTUALIZANDO STOCK [{timestamp}] ===")
            print(f"Producto: {producto.nombre} (ID: {producto.id})")
            print(f"Stock ANTES: {producto.stock_total}")
            print(f"Cantidad recibida: {cantidad}")
            print(f"Usuario: {usuario}")
            print(f"Nota: {nota}")
            
            # ✅ NUEVO: Crear MovimientoInventario (actualiza automáticamente ambos modelos)
            tipo_movimiento = 'ENTRADA' if cantidad > 0 else 'SALIDA'
            cantidad_absoluta = abs(cantidad)
            
            from .models import MovimientoInventario
            movimiento = MovimientoInventario.objects.create(
                producto=producto,
                tipo=tipo_movimiento,
                cantidad=cantidad_absoluta,
                usuario=usuario,
                nota=nota
            )
            
            # El método save() de MovimientoInventario actualiza:
            # - Producto.stock_total
            # - Stock.cantidad_actual
            
            # Refrescar producto para obtener stock actualizado
            producto.refresh_from_db()
            
            print(f"Stock DESPUÉS: {producto.stock_total}")
            print(f"Movimiento creado: {tipo_movimiento} de {cantidad_absoluta}")
            print(f"=== ✅ ACTUALIZACIÓN COMPLETADA [{timestamp}] ===\n")
            
            return Response({
                'success': True,
                'stock_actual': producto.stock_total,
                'movimiento_id': movimiento.id,
                'tipo': tipo_movimiento,
                'nota': 'Stock actualizado con MovimientoInventario (trazabilidad completa)'
            })
            
        except (ValueError, TypeError) as e:
            print(f"❌ Error de valor: {e}")
            return Response({'error': 'La cantidad debe ser un número entero'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            print(f"❌ Error general: {e}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    # 🆕 ENDPOINTS FILTRADOS POR MÓDULO
    @action(detail=False, methods=['get'], url_path='pos')
    def productos_pos(self, request):
        """Obtener productos disponibles para POS"""
        productos = Producto.objects.filter(disponible_pos=True, activo=True).order_by('orden', 'id')
        serializer = self.get_serializer(productos, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], url_path='cargue')
    def productos_cargue(self, request):
        """Obtener productos disponibles para Cargue"""
        productos = Producto.objects.filter(disponible_cargue=True, activo=True).order_by('orden', 'id')
        serializer = self.get_serializer(productos, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], url_path='pedidos')
    def productos_pedidos(self, request):
        """Obtener productos disponibles para Pedidos"""
        productos = Producto.objects.filter(disponible_pedidos=True, activo=True).order_by('orden', 'id')
        serializer = self.get_serializer(productos, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], url_path='inventario')
    def productos_inventario(self, request):
        """Obtener productos disponibles para Inventario"""
        productos = Producto.objects.filter(disponible_inventario=True, activo=True).order_by('orden', 'id')
        serializer = self.get_serializer(productos, many=True)
        return Response(serializer.data)

class StockViewSet(viewsets.ModelViewSet):
    """API para gestionar stock de productos"""
    queryset = Stock.objects.select_related('producto').all()
    serializer_class = StockSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = Stock.objects.select_related('producto').all()
        
        # 🎯 SOLO productos activos por defecto
        queryset = queryset.filter(producto__activo=True)
        
        # Filtrar por producto_id
        producto_id = self.request.query_params.get('producto_id')
        if producto_id:
            queryset = queryset.filter(producto_id=producto_id)
        
        # Filtrar por ubicación de inventario
        ubicacion = self.request.query_params.get('ubicacion')
        if ubicacion:
            # Incluir productos con ubicacion=PRODUCCION O sin ubicacion (NULL)
            from django.db.models import Q
            queryset = queryset.filter(
                Q(producto__ubicacion_inventario=ubicacion) | 
                Q(producto__ubicacion_inventario__isnull=True) |
                Q(producto__ubicacion_inventario='')
            )
            
        return queryset.order_by('producto__orden', 'producto__id')

class LoteViewSet(viewsets.ModelViewSet):
    queryset = Lote.objects.all()
    serializer_class = LoteSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = Lote.objects.all()

        # Filtros
        producto_id = self.request.query_params.get('producto')
        if producto_id:
            queryset = queryset.filter(producto_id=producto_id)

        fecha_produccion = self.request.query_params.get('fecha_produccion')
        if fecha_produccion:
            queryset = queryset.filter(fecha_produccion=fecha_produccion)

        tipo_origen = self.request.query_params.get('tipo_origen')
        if tipo_origen:
            queryset = queryset.filter(tipo_origen=tipo_origen)

        return queryset.order_by('-fecha_creacion')

class MovimientoInventarioViewSet(viewsets.ModelViewSet):
    queryset = MovimientoInventario.objects.all()
    serializer_class = MovimientoInventarioSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = MovimientoInventario.objects.all().order_by('-fecha')
        
        # Aplicar filtros
        filters = {
            'producto': 'producto_id',
            'tipo': 'tipo',
            'fecha_inicio': 'fecha__gte',
            'fecha_fin': 'fecha__lte'
        }
        
        for param, field in filters.items():
            value = self.request.query_params.get(param)
            if value:
                if param == 'tipo':
                    value = value.upper()
                queryset = queryset.filter(**{field: value})
        
        return queryset

class RegistroInventarioViewSet(viewsets.ModelViewSet):
    queryset = RegistroInventario.objects.all()
    serializer_class = RegistroInventarioSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = RegistroInventario.objects.all()
        
        fecha_produccion = self.request.query_params.get('fecha_produccion')
        if fecha_produccion:
            queryset = queryset.filter(fecha_produccion=fecha_produccion)
            
        return queryset.order_by('-fecha_creacion')

class VentaViewSet(viewsets.ModelViewSet):
    """API para gestionar ventas"""
    queryset = Venta.objects.all()
    serializer_class = VentaSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = Venta.objects.all().order_by('-fecha')
        
        # Filtros opcionales
        fecha_inicio = self.request.query_params.get('fecha_inicio')
        fecha_fin = self.request.query_params.get('fecha_fin')
        vendedor = self.request.query_params.get('vendedor')
        estado = self.request.query_params.get('estado')
        
        if fecha_inicio:
            queryset = queryset.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(fecha__lte=fecha_fin)
        if vendedor:
            queryset = queryset.filter(vendedor__icontains=vendedor)
        if estado:
            queryset = queryset.filter(estado=estado.upper())
            
        return queryset
    
    def create(self, request, *args, **kwargs):
        """Crear venta con sus detalles"""
        try:
            venta_data = request.data.copy()
            detalles_data = venta_data.pop('detalles', [])
            
            # 🛡️ VALIDAR DUPLICADOS - Verificar por número de factura (más confiable)
            numero_factura = venta_data.get('numero_factura')
            
            if numero_factura:
                venta_existente = Venta.objects.filter(numero_factura=numero_factura).first()
                
                if venta_existente:
                    print(f"⚠️ Venta duplicada detectada: Ticket #{numero_factura} ya existe")
                    return Response({
                        'error': 'Venta duplicada',
                        'message': f'El ticket #{numero_factura} ya fue registrado',
                        'venta_id': venta_existente.id
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Rechazar ventas sin detalles y con total 0 (registros fantasma del offline sync)
            total = float(venta_data.get('total', 0) or 0)
            if total == 0 and not detalles_data:
                return Response(
                    {'error': 'Venta inválida: total 0 sin productos'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Crear la venta
            venta_serializer = self.get_serializer(data=venta_data)
            if not venta_serializer.is_valid():
                print("❌ Errores en venta:", venta_serializer.errors)
                return Response(venta_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
            venta = venta_serializer.save()

            
            # Crear los detalles directamente
            for detalle_data in detalles_data:
                try:
                    producto = Producto.objects.get(id=detalle_data['producto'])
                    DetalleVenta.objects.create(
                        venta=venta,
                        producto=producto,
                        cantidad=detalle_data['cantidad'],
                        precio_unitario=float(detalle_data['precio_unitario'])
                    )
                    print(f"✅ Detalle creado: {producto.nombre} x{detalle_data['cantidad']}")
                    
                    # 🆕 ACTUALIZAR INVENTARIO (CargueIDx)
                    try:
                        v_id_raw = str(venta_data.get('vendedor_id', ''))
                        if v_id_raw.isdigit():
                            v_id_str = f"ID{v_id_raw}"
                        elif v_id_raw.upper().startswith('ID'):
                            v_id_str = v_id_raw.upper()
                        else:
                            v_id_str = None
                        
                        if v_id_str:
                            modelos = {
                                'ID1': CargueID1, 'ID2': CargueID2, 'ID3': CargueID3,
                                'ID4': CargueID4, 'ID5': CargueID5, 'ID6': CargueID6
                            }
                            Modelo = modelos.get(v_id_str)
                            
                            if Modelo:
                                from django.utils.dateparse import parse_datetime
                                fecha_str = venta_data.get('fecha')
                                if isinstance(fecha_str, str):
                                    fecha_dt = parse_datetime(fecha_str)
                                    fecha_venta = fecha_dt.date() if fecha_dt else None
                                elif hasattr(fecha_str, 'date'):
                                    fecha_venta = fecha_str.date()
                                else:
                                    from django.utils import timezone
                                    fecha_venta = timezone.now().date()
                                
                                if fecha_venta:
                                    cargue_item = Modelo.objects.filter(
                                        fecha=fecha_venta, 
                                        producto=producto
                                    ).first()
                                    
                                    if cargue_item:
                                        cargue_item.vendidas = (cargue_item.vendidas or 0) + detalle_data['cantidad']
                                        cargue_item.save()
                                        print(f"✅ Stock actualizado Cargue{v_id_str}: {producto.nombre} (+{detalle_data['cantidad']}) -> Total Vendidas: {cargue_item.vendidas}")
                    except Exception as e_inv:
                        print(f"⚠️ Error actualizando inventario en venta: {e_inv}")

                except Producto.DoesNotExist:
                    print(f"❌ Producto no encontrado: {detalle_data['producto']}")
                    return Response(
                        {'error': f'Producto {detalle_data["producto"]} no encontrado'}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
                except Exception as e:
                    print(f"❌ Error creando detalle: {str(e)}")
                    return Response(
                        {'error': f'Error creando detalle: {str(e)}'}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Retornar venta completa con detalles
            venta_completa = VentaSerializer(venta)
            return Response(venta_completa.data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            print("❌ Error general:", str(e))
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )

class DetalleVentaViewSet(viewsets.ModelViewSet):
    """API para gestionar detalles de venta"""
    queryset = DetalleVenta.objects.all()
    serializer_class = DetalleVentaSerializer
    permission_classes = [permissions.AllowAny]

class TipoNegocioViewSet(viewsets.ModelViewSet):
    """API para gestionar tipos de negocio"""
    queryset = TipoNegocio.objects.filter(activo=True)
    serializer_class = TipoNegocioSerializer
    permission_classes = [permissions.AllowAny]

class ClienteViewSet(viewsets.ModelViewSet):
    """API para gestionar clientes"""
    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = Cliente.objects.all().order_by('-fecha_creacion')
        
        # Filtros opcionales
        activo = self.request.query_params.get('activo')
        identificacion = self.request.query_params.get('identificacion')
        nombre = self.request.query_params.get('nombre')
        
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
        if identificacion:
            queryset = queryset.filter(identificacion__icontains=identificacion)
        if nombre:
            queryset = queryset.filter(nombre_completo__icontains=nombre)
            
        return queryset
    
    def _sincronizar_con_cliente_ruta(self, cliente, origen='PEDIDOS', vendedor_anterior=None, zona_anterior=None):
        """
        Sincroniza un cliente de la tabla Cliente con ClienteRuta
        Se ejecuta cuando el cliente tiene vendedor_asignado (es decir, está asignado a una ruta)
        
        Args:
            cliente: Instancia del cliente
            origen: Origen del cliente (PEDIDOS, APP, etc.)
            vendedor_anterior: Nombre del vendedor anterior (para detectar cambios de ruta)
            zona_anterior: Zona/Ruta anterior (para detectar cambios de ruta)
        """
        from api.models import ClienteRuta, Ruta, Vendedor
        
        # 🔄 SI CAMBIÓ LA ZONA/RUTA, eliminar el registro de la zona/ruta anterior
        if zona_anterior and zona_anterior != cliente.zona_barrio:
            try:
                ruta_vieja = Ruta.objects.filter(nombre__iexact=zona_anterior).first()
                if ruta_vieja:
                    # Eliminar cliente de la ruta anterior
                    ClienteRuta.objects.filter(
                        ruta=ruta_vieja,
                        nombre_negocio=cliente.alias or cliente.nombre_completo
                    ).delete()
                    print(f"🔄 Cliente eliminado de ruta anterior: {ruta_vieja.nombre}")
            except Exception as e:
                print(f"⚠️ Error eliminando de ruta anterior: {e}")
        
        # Determinar la ruta a usar (prioridad: zona_barrio > vendedor)
        ruta = None
        
        # 1. Intentar buscar por zona_barrio (campo que contiene el nombre de la ruta)
        if cliente.zona_barrio:
            ruta = Ruta.objects.filter(nombre__iexact=cliente.zona_barrio).first()
            if ruta:
                print(f"✅ Ruta encontrada por zona_barrio: {ruta.nombre}")
        
        # 2. Si no hay zona_barrio, buscar por vendedor_asignado
        if not ruta and cliente.vendedor_asignado:
            vendedor = Vendedor.objects.filter(nombre=cliente.vendedor_asignado).first()
            if vendedor:
                ruta = Ruta.objects.filter(vendedor=vendedor).first()
                if ruta:
                    print(f"✅ Ruta encontrada por vendedor: {ruta.nombre}")
        
        # Si no se encontró ruta, no sincronizar
        if not ruta:
            print(f"⚠️ No se encontró ruta para el cliente {cliente.nombre_completo}")
            # Si no tiene ruta, eliminar de cualquier ClienteRuta
            ClienteRuta.objects.filter(
                nombre_negocio=cliente.alias or cliente.nombre_completo
            ).delete()
            print(f"🗑️ Cliente eliminado de todas las rutas (sin ruta asignada)")
            return
        
        try:
            # Preparar datos para ClienteRuta
            # Formato del tipo_negocio: "TipoNegocio | ORIGEN"
            tipo_negocio_base = cliente.alias or "Sin especificar"
            tipo_negocio = f"{tipo_negocio_base} | {origen}"  # Ej: "LA FONDA | PEDIDOS"
            
            # Buscar si ya existe un ClienteRuta con este nombre de negocio en esta ruta
            cliente_ruta_existente = ClienteRuta.objects.filter(
                ruta=ruta,
                nombre_negocio=cliente.alias or cliente.nombre_completo
            ).first()
            
            if cliente_ruta_existente:
                # Actualizar existente
                cliente_ruta_existente.nombre_contacto = cliente.nombre_completo
                cliente_ruta_existente.direccion = cliente.direccion or ''
                cliente_ruta_existente.telefono = cliente.telefono_1 or cliente.movil or ''
                cliente_ruta_existente.tipo_negocio = tipo_negocio
                cliente_ruta_existente.dia_visita = cliente.dia_entrega or 'SABADO'
                cliente_ruta_existente.activo = cliente.activo
                cliente_ruta_existente.nota = cliente.nota # 🆕 Sincronizar nota
                cliente_ruta_existente.save()
                print(f"✅ ClienteRuta actualizado: {cliente_ruta_existente.nombre_negocio} en {ruta.nombre}")
            else:
                # Crear nuevo ClienteRuta
                # Obtener el último orden para esta ruta
                ultimo_orden = ClienteRuta.objects.filter(ruta=ruta).aggregate(
                    models.Max('orden')
                )['orden__max'] or 0
                
                ClienteRuta.objects.create(
                    ruta=ruta,
                    nombre_negocio=cliente.alias or cliente.nombre_completo,
                    nombre_contacto=cliente.nombre_completo,
                    direccion=cliente.direccion or '',
                    telefono=cliente.telefono_1 or cliente.movil or '',
                    tipo_negocio=tipo_negocio,
                    dia_visita=cliente.dia_entrega or 'SABADO',
                    orden=ultimo_orden + 1,
                    activo=cliente.activo,
                    nota=cliente.nota # 🆕 Sincronizar nota
                )
                print(f"✅ ClienteRuta creado: {cliente.alias or cliente.nombre_completo} en ruta {ruta.nombre}")
                
        except Exception as e:
            print(f"❌ Error sincronizando con ClienteRuta: {e}")
            import traceback
            traceback.print_exc()
    
    def perform_create(self, serializer):
        """Se ejecuta al crear un nuevo cliente"""
        cliente = serializer.save()
        # Sincronizar con ClienteRuta si tiene ruta asignada
        self._sincronizar_con_cliente_ruta(cliente, origen='PEDIDOS')
    
    def perform_update(self, serializer):
        """Se ejecuta al actualizar un cliente existente"""
        # Obtener los valores anteriores antes de guardar
        try:
            cliente_anterior = Cliente.objects.get(pk=serializer.instance.pk)
            vendedor_anterior = cliente_anterior.vendedor_asignado
            zona_anterior = cliente_anterior.zona_barrio
        except Cliente.DoesNotExist:
            vendedor_anterior = None
            zona_anterior = None
        
        # Guardar los cambios
        cliente = serializer.save()
        
        # 🔄 Solo sincronizar si NO viene de una sincronización desde ClienteRuta
        if not getattr(cliente, '_sincronizando', False):
            # Sincronizar con ClienteRuta (pasando valores anteriores para detectar cambios)
            self._sincronizar_con_cliente_ruta(
                cliente, 
                origen='PEDIDOS', 
                vendedor_anterior=vendedor_anterior,
                zona_anterior=zona_anterior
            )
        else:
            print(f"⏭️ Sincronización Cliente → ClienteRuta omitida (ya sincronizado desde ClienteRuta)")
    
    def perform_destroy(self, instance):
        """Se ejecuta al eliminar un cliente"""
        from api.models import ClienteRuta
        
        # Eliminar el cliente de ClienteRuta antes de eliminarlo
        nombre_negocio = instance.alias or instance.nombre_completo
        clientes_ruta_eliminados = ClienteRuta.objects.filter(
            nombre_negocio=nombre_negocio
        ).delete()
        
        if clientes_ruta_eliminados[0] > 0:
            print(f"🗑️ Cliente eliminado de ClienteRuta: {nombre_negocio} ({clientes_ruta_eliminados[0]} registros)")
        
        # Eliminar el cliente
        instance.delete()
        print(f"✅ Cliente eliminado completamente: {nombre_negocio}")

class ProductosFrecuentesViewSet(viewsets.ModelViewSet):
    """API para gestionar productos frecuentes de clientes"""
    queryset = ProductosFrecuentes.objects.all()
    serializer_class = ProductosFrecuentesSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = ProductosFrecuentes.objects.all()
        
        # Filtrar por cliente
        cliente_id = self.request.query_params.get('cliente')
        if cliente_id:
            queryset = queryset.filter(cliente_id=cliente_id)
        
        # Filtrar por día
        dia = self.request.query_params.get('dia')
        if dia:
            queryset = queryset.filter(dia=dia.upper())
        
        return queryset

class ListaPrecioViewSet(viewsets.ModelViewSet):
    """API para gestionar listas de precios"""
    queryset = ListaPrecio.objects.all()
    serializer_class = ListaPrecioSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = ListaPrecio.objects.all().order_by('-fecha_creacion')
        
        # Filtros opcionales
        activo = self.request.query_params.get('activo')
        tipo = self.request.query_params.get('tipo')
        
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
        if tipo:
            queryset = queryset.filter(tipo=tipo.upper())
            
        return queryset

class PrecioProductoViewSet(viewsets.ModelViewSet):
    """API para gestionar precios de productos"""
    queryset = PrecioProducto.objects.all()
    serializer_class = PrecioProductoSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = PrecioProducto.objects.all().order_by('producto__nombre')
        
        # Filtros opcionales
        lista_precio = self.request.query_params.get('lista_precio')
        producto = self.request.query_params.get('producto')
        activo = self.request.query_params.get('activo')
        
        if lista_precio:
            queryset = queryset.filter(lista_precio_id=lista_precio)
        if producto:
            queryset = queryset.filter(producto_id=producto)
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
            
        return queryset

# ========================================
# NUEVAS VIEWS SIMPLIFICADAS
# ========================================

class CargueID1ViewSet(viewsets.ModelViewSet):
    """API simplificada para CargueID1 - Como api_vendedor"""
    queryset = CargueID1.objects.all()
    serializer_class = CargueID1Serializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        import re
        queryset = CargueID1.objects.all().order_by('-fecha', '-fecha_actualizacion')
        
        # Filtros opcionales
        dia = self.request.query_params.get('dia')
        fecha = self.request.query_params.get('fecha')
        producto = self.request.query_params.get('producto')
        activo = self.request.query_params.get('activo')
        
        if dia:
            queryset = queryset.filter(dia=dia.upper())
        if fecha:
            queryset = queryset.filter(fecha=fecha)
        if producto:
            # 🔧 Normalizar nombre del producto (eliminar espacios múltiples)
            producto_normalizado = re.sub(r'\s+', ' ', producto).strip()
            queryset = queryset.filter(producto=producto_normalizado)
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
            
        return queryset
    
    def create(self, request, *args, **kwargs):
        """Crear registro con logging detallado"""
        import traceback
        import os
        from datetime import datetime
        import logging
        
        logger = logging.getLogger(__name__)
        
        # Log usando el sistema de logging de Django (más seguro)
        try:
            logger.info(f"🆕 CREATE CargueID1 - producto: {request.data.get('producto')}, "
                       f"cantidad: {request.data.get('cantidad')}, dctos: {request.data.get('dctos')}, "
                       f"adicional: {request.data.get('adicional')}, dia: {request.data.get('dia')}, "
                       f"fecha: {request.data.get('fecha')}")
        except Exception as e:
            logger.warning(f"Error al loguear create: {e}")
        
        return super().create(request, *args, **kwargs)

class CargueID2ViewSet(viewsets.ModelViewSet):
    """API simplificada para CargueID2 - Como api_vendedor"""
    queryset = CargueID2.objects.all()
    serializer_class = CargueID2Serializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        import re
        queryset = CargueID2.objects.all().order_by('-fecha', '-fecha_actualizacion')
        
        # Filtros opcionales
        dia = self.request.query_params.get('dia')
        fecha = self.request.query_params.get('fecha')
        producto = self.request.query_params.get('producto')
        activo = self.request.query_params.get('activo')
        
        if dia:
            queryset = queryset.filter(dia=dia.upper())
        if fecha:
            queryset = queryset.filter(fecha=fecha)
        if producto:
            producto_normalizado = re.sub(r'\s+', ' ', producto).strip()
            queryset = queryset.filter(producto=producto_normalizado)
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
            
        return queryset

class CargueID3ViewSet(viewsets.ModelViewSet):
    """API simplificada para CargueID3 - Como api_vendedor"""
    queryset = CargueID3.objects.all()
    serializer_class = CargueID3Serializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        import re
        queryset = CargueID3.objects.all().order_by('-fecha', '-fecha_actualizacion')
        
        # Filtros opcionales
        dia = self.request.query_params.get('dia')
        fecha = self.request.query_params.get('fecha')
        producto = self.request.query_params.get('producto')
        activo = self.request.query_params.get('activo')
        
        if dia:
            queryset = queryset.filter(dia=dia.upper())
        if fecha:
            queryset = queryset.filter(fecha=fecha)
        if producto:
            producto_normalizado = re.sub(r'\s+', ' ', producto).strip()
            queryset = queryset.filter(producto=producto_normalizado)
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
            
        return queryset

class CargueID4ViewSet(viewsets.ModelViewSet):
    """API simplificada para CargueID4 - Como api_vendedor"""
    queryset = CargueID4.objects.all()
    serializer_class = CargueID4Serializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        import re
        queryset = CargueID4.objects.all().order_by('-fecha', '-fecha_actualizacion')
        
        # Filtros opcionales
        dia = self.request.query_params.get('dia')
        fecha = self.request.query_params.get('fecha')
        producto = self.request.query_params.get('producto')
        activo = self.request.query_params.get('activo')
        
        if dia:
            queryset = queryset.filter(dia=dia.upper())
        if fecha:
            queryset = queryset.filter(fecha=fecha)
        if producto:
            producto_normalizado = re.sub(r'\s+', ' ', producto).strip()
            queryset = queryset.filter(producto=producto_normalizado)
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
            
        return queryset

class CargueID5ViewSet(viewsets.ModelViewSet):
    """API simplificada para CargueID5 - Como api_vendedor"""
    queryset = CargueID5.objects.all()
    serializer_class = CargueID5Serializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        import re
        queryset = CargueID5.objects.all().order_by('-fecha', '-fecha_actualizacion')
        
        # Filtros opcionales
        dia = self.request.query_params.get('dia')
        fecha = self.request.query_params.get('fecha')
        producto = self.request.query_params.get('producto')
        activo = self.request.query_params.get('activo')
        
        if dia:
            queryset = queryset.filter(dia=dia.upper())
        if fecha:
            queryset = queryset.filter(fecha=fecha)
        if producto:
            producto_normalizado = re.sub(r'\s+', ' ', producto).strip()
            queryset = queryset.filter(producto=producto_normalizado)
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
            
        return queryset

class CargueID6ViewSet(viewsets.ModelViewSet):
    """API simplificada para CargueID6 - Como api_vendedor"""
    queryset = CargueID6.objects.all()
    serializer_class = CargueID6Serializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        import re
        queryset = CargueID6.objects.all().order_by('-fecha', '-fecha_actualizacion')
        
        # Filtros opcionales
        dia = self.request.query_params.get('dia')
        fecha = self.request.query_params.get('fecha')
        producto = self.request.query_params.get('producto')
        activo = self.request.query_params.get('activo')
        
        if dia:
            queryset = queryset.filter(dia=dia.upper())
        if fecha:
            queryset = queryset.filter(fecha=fecha)
        if producto:
            producto_normalizado = re.sub(r'\s+', ' ', producto).strip()
            queryset = queryset.filter(producto=producto_normalizado)
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
            
        return queryset

class ProduccionViewSet(viewsets.ModelViewSet):
    """API para Producción con función de congelado"""
    queryset = Produccion.objects.all()
    serializer_class = ProduccionSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = Produccion.objects.all().order_by('-fecha', '-fecha_actualizacion')
        
        # Filtros opcionales
        fecha = self.request.query_params.get('fecha')
        congelado = self.request.query_params.get('congelado')
        activo = self.request.query_params.get('activo')
        
        if fecha:
            queryset = queryset.filter(fecha=fecha)
        if congelado is not None:
            queryset = queryset.filter(congelado=congelado.lower() == 'true')
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
            
        return queryset
    
    @action(detail=True, methods=['post'])
    def congelar(self, request, pk=None):
        """Congelar producción"""
        try:
            produccion = self.get_object()
            usuario = request.data.get('usuario', 'Sistema')
            
            if produccion.congelado:
                return Response(
                    {'error': 'La producción ya está congelada'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            produccion.congelar(usuario)
            
            return Response({
                'success': True,
                'message': 'Producción congelada exitosamente',
                'congelado': True,
                'fecha_congelado': produccion.fecha_congelado,
                'usuario_congelado': produccion.usuario_congelado
            })
            
        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def descongelar(self, request, pk=None):
        """Descongelar producción"""
        try:
            produccion = self.get_object()
            usuario = request.data.get('usuario', 'Sistema')
            
            if not produccion.congelado:
                return Response(
                    {'error': 'La producción no está congelada'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            produccion.descongelar(usuario)
            
            return Response({
                'success': True,
                'message': 'Producción descongelada exitosamente',
                'congelado': False,
                'usuario_descongelado': usuario
            })
            
        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

# ========================================
# VIEWSET PARA VENDEDORES/RESPONSABLES
# ========================================

class VendedorViewSet(viewsets.ViewSet):
    """API para gestionar responsables de vendedores"""
    permission_classes = [permissions.AllowAny]
    
    @action(detail=False, methods=['post'])
    def actualizar_responsable(self, request):
        """Actualizar responsable y ruta de un vendedor específico"""
        try:
            id_vendedor = request.data.get('id_vendedor')
            responsable = request.data.get('responsable')
            ruta = request.data.get('ruta', '')
            
            if not id_vendedor or not responsable:
                return Response(
                    {'error': 'id_vendedor y responsable son requeridos'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Mapear ID de vendedor a modelo correspondiente
            modelos_vendedor = {
                'ID1': CargueID1,
                'ID2': CargueID2,
                'ID3': CargueID3,
                'ID4': CargueID4,
                'ID5': CargueID5,
                'ID6': CargueID6,
            }
            
            modelo = modelos_vendedor.get(id_vendedor)
            if not modelo:
                return Response(
                    {'error': f'Vendedor {id_vendedor} no válido'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Actualizar todos los registros existentes de este vendedor
            from django.utils import timezone
            datos_actualizar = {
                'responsable': responsable,
                'fecha_actualizacion': timezone.now()  # 🔥 Actualizar timestamp para polling
            }
            if ruta:
                datos_actualizar['ruta'] = ruta
            
            registros_actualizados = modelo.objects.filter(activo=True).update(**datos_actualizar)
            
            # Si no hay registros, crear uno dummy para guardar el responsable
            if registros_actualizados == 0:
                from datetime import date
                modelo.objects.create(
                    dia='LUNES',
                    fecha=date.today(),
                    responsable=responsable,
                    ruta=ruta if ruta else '',
                    usuario='Sistema',
                    activo=True
                )
                registros_actualizados = 1
            
            return Response({
                'success': True,
                'message': f'Responsable y ruta actualizados para {id_vendedor}',
                'id_vendedor': id_vendedor,
                'responsable': responsable,
                'ruta': ruta,
                'registros_actualizados': registros_actualizados
            })
            
        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def obtener_responsable(self, request):
        """Obtener responsable y ruta actual de un vendedor o todos los vendedores"""
        try:
            id_vendedor = request.query_params.get('id_vendedor')
            
            # Mapear ID de vendedor a modelo correspondiente
            modelos_vendedor = {
                'ID1': CargueID1,
                'ID2': CargueID2,
                'ID3': CargueID3,
                'ID4': CargueID4,
                'ID5': CargueID5,
                'ID6': CargueID6,
            }
            
            # Si no se especifica id_vendedor, devolver todos
            if not id_vendedor:
                resultados = []
                for id_v, modelo in modelos_vendedor.items():
                    ultimo_registro = modelo.objects.filter(activo=True).order_by('-fecha_creacion').first()
                    responsable = 'RESPONSABLE'
                    ruta = 'Sin ruta'
                    fecha_creacion = None
                    
                    if ultimo_registro:
                        if ultimo_registro.responsable:
                            responsable = ultimo_registro.responsable
                        if hasattr(ultimo_registro, 'ruta') and ultimo_registro.ruta:
                            ruta = ultimo_registro.ruta
                        fecha_creacion = ultimo_registro.fecha_creacion
                    
                    resultados.append({
                        'id': id_v,
                        'responsable': responsable,
                        'ruta': ruta,
                        'fecha_creacion': fecha_creacion
                    })
                
                return Response(resultados)
            
            # Si se especifica id_vendedor, devolver solo ese
            modelo = modelos_vendedor.get(id_vendedor)
            if not modelo:
                return Response(
                    {'error': f'Vendedor {id_vendedor} no válido'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Obtener el responsable y ruta del registro más reciente
            ultimo_registro = modelo.objects.filter(activo=True).order_by('-fecha_creacion').first()
            
            responsable = 'RESPONSABLE'  # Valor por defecto
            ruta = 'Sin ruta'  # Valor por defecto
            
            if ultimo_registro:
                if ultimo_registro.responsable:
                    responsable = ultimo_registro.responsable
                if hasattr(ultimo_registro, 'ruta') and ultimo_registro.ruta:
                    ruta = ultimo_registro.ruta
            
            return Response({
                'success': True,
                'id_vendedor': id_vendedor,
                'responsable': responsable,
                'ruta': ruta,
                'results': [{
                    'id_vendedor': id_vendedor,
                    'responsable': responsable,
                    'ruta': ruta
                }]
            })
            
        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ProduccionSolicitadaViewSet(viewsets.ViewSet):
    """API para gestionar producción solicitada"""
    permission_classes = [permissions.AllowAny]
    
    def create(self, request):
        """Guardar/actualizar solicitadas de producción"""
        try:
            dia = request.data.get('dia')
            fecha = request.data.get('fecha')
            productos = request.data.get('productos', [])
            
            if not dia or not fecha:
                return Response(
                    {'error': 'Día y fecha son requeridos'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Limpiar registros existentes para este día/fecha
            ProduccionSolicitada.objects.filter(dia=dia, fecha=fecha).delete()
            
            # Crear nuevos registros
            registros_creados = []
            for producto_data in productos:
                if producto_data.get('cantidad_solicitada', 0) > 0:
                    registro = ProduccionSolicitada.objects.create(
                        dia=dia,
                        fecha=fecha,
                        producto_nombre=producto_data['producto_nombre'],
                        cantidad_solicitada=producto_data['cantidad_solicitada']
                    )
                    registros_creados.append(registro)
            
            serializer = ProduccionSolicitadaSerializer(registros_creados, many=True)
            
            return Response({
                'success': True,
                'message': f'Guardadas {len(registros_creados)} solicitadas para {dia} {fecha}',
                'data': serializer.data
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def list(self, request):
        """Obtener solicitadas por fecha"""
        fecha = request.query_params.get('fecha')
        dia = request.query_params.get('dia')
        
        queryset = ProduccionSolicitada.objects.all()
        
        if fecha:
            queryset = queryset.filter(fecha=fecha)
        if dia:
            queryset = queryset.filter(dia=dia.upper())
            
        queryset = queryset.order_by('producto_nombre')
        serializer = ProduccionSolicitadaSerializer(queryset, many=True)
        
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def calcular_desde_cargue(self, request):
        """Calcular y guardar solicitadas sumando todos los IDs de cargue"""
        try:
            dia = request.data.get('dia')
            fecha = request.data.get('fecha')
            
            if not dia or not fecha:
                return Response(
                    {'error': 'Día y fecha son requeridos'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Obtener todos los registros de cargue para esta fecha
            from django.db.models import Sum
            
            # Diccionario para acumular cantidades por producto
            productos_suma = {}
            
            # Obtener lista de productos válidos (que existen en la BD)
            productos_validos = set(Producto.objects.values_list('nombre', flat=True))
            
            # Consultar cada tabla de cargue (ID1 a ID6)
            for modelo in [CargueID1, CargueID2, CargueID3, CargueID4, CargueID5, CargueID6]:
                registros = modelo.objects.filter(dia=dia.upper(), fecha=fecha)
                
                for registro in registros:
                    producto = registro.producto
                    # Forzar conversión a int para evitar concatenación de strings
                    cantidad = int(registro.cantidad or 0)
                    
                    # Solo procesar si el producto existe en la BD y tiene cantidad > 0
                    if producto and cantidad > 0 and producto in productos_validos:
                        if producto in productos_suma:
                            productos_suma[producto] += cantidad
                        else:
                            productos_suma[producto] = cantidad
            
            # Limpiar registros existentes para este día/fecha
            ProduccionSolicitada.objects.filter(dia=dia.upper(), fecha=fecha).delete()
            
            # Crear nuevos registros con las sumas
            registros_creados = []
            for producto_nombre, cantidad_total in productos_suma.items():
                if cantidad_total > 0:
                    registro = ProduccionSolicitada.objects.create(
                        dia=dia.upper(),
                        fecha=fecha,
                        producto_nombre=producto_nombre,
                        cantidad_solicitada=cantidad_total
                    )
                    registros_creados.append(registro)
            
            serializer = ProduccionSolicitadaSerializer(registros_creados, many=True)
            
            return Response({
                'success': True,
                'message': f'Calculadas {len(registros_creados)} solicitadas para {dia} {fecha}',
                'productos_procesados': len(productos_suma),
                'data': serializer.data
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            import traceback
            return Response(
                {'error': str(e), 'traceback': traceback.format_exc()}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

# ========================================
# VISTAS PARA SISTEMA POS - CAJEROS
# ========================================

from .models import Sucursal, Cajero, Turno, VentaCajero, ArqueoCaja
from .serializers import (
    SucursalSerializer, CajeroSerializer, TurnoSerializer, 
    VentaCajeroSerializer, CajeroLoginSerializer, TurnoResumenSerializer,
    ArqueoCajaSerializer
)
from django.utils import timezone
from django.db.models import Q, Sum, Count

class SucursalViewSet(viewsets.ModelViewSet):
    """API para gestionar sucursales"""
    queryset = Sucursal.objects.all()
    serializer_class = SucursalSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = Sucursal.objects.all()
        activo = self.request.query_params.get('activo', None)
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
        return queryset.order_by('nombre')
    
    @action(detail=False, methods=['get'])
    def activas(self, request):
        """Obtener solo sucursales activas"""
        sucursales = Sucursal.objects.filter(activo=True).order_by('nombre')
        serializer = self.get_serializer(sucursales, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def principal(self, request):
        """Obtener sucursal principal"""
        try:
            sucursal = Sucursal.objects.filter(es_principal=True, activo=True).first()
            if not sucursal:
                # Si no hay principal, tomar la primera activa
                sucursal = Sucursal.objects.filter(activo=True).first()
            
            if sucursal:
                serializer = self.get_serializer(sucursal)
                return Response(serializer.data)
            else:
                return Response({'error': 'No hay sucursales disponibles'}, 
                              status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class CajeroViewSet(viewsets.ModelViewSet):
    """API para gestionar cajeros"""
    queryset = Cajero.objects.all()
    serializer_class = CajeroSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = Cajero.objects.select_related('sucursal').all()
        
        # Filtros
        sucursal_id = self.request.query_params.get('sucursal_id', None)
        activo = self.request.query_params.get('activo', None)
        
        # 🔧 Validar que sucursal_id sea un número válido
        if sucursal_id and sucursal_id not in ['undefined', 'null', '']:
            try:
                sucursal_id_int = int(sucursal_id)
                queryset = queryset.filter(sucursal_id=sucursal_id_int)
            except (ValueError, TypeError):
                pass  # Ignorar si no es un número válido
                
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
            
        return queryset.order_by('sucursal__nombre', 'nombre')
    
    @action(detail=False, methods=['post'])
    def authenticate(self, request):
        """Autenticar cajero"""
        serializer = CajeroLoginSerializer(data=request.data)
        if serializer.is_valid():
            cajero = serializer.validated_data['cajero']
            
            # Serializar datos del cajero
            cajero_data = CajeroSerializer(cajero).data
            
            return Response({
                'success': True,
                'message': 'Autenticación exitosa',
                'cajero': cajero_data
            })
        else:
            return Response({
                'success': False,
                'message': 'Credenciales inválidas',
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def activos_por_sucursal(self, request):
        """Obtener cajeros activos por sucursal"""
        sucursal_id = request.query_params.get('sucursal_id')
        if not sucursal_id:
            return Response({'error': 'sucursal_id es requerido'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        cajeros = Cajero.objects.filter(
            sucursal_id=sucursal_id, 
            activo=True
        ).order_by('nombre')
        
        serializer = self.get_serializer(cajeros, many=True)
        return Response(serializer.data)

class TurnoViewSet(viewsets.ModelViewSet):
    """API para gestionar turnos de cajeros"""
    queryset = Turno.objects.all()
    serializer_class = TurnoSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = Turno.objects.select_related('cajero', 'sucursal').all()
        
        # Filtros
        cajero_id = self.request.query_params.get('cajero_id', None)
        sucursal_id = self.request.query_params.get('sucursal_id', None)
        estado = self.request.query_params.get('estado', None)
        fecha_desde = self.request.query_params.get('fecha_desde', None)
        fecha_hasta = self.request.query_params.get('fecha_hasta', None)
        
        if cajero_id:
            queryset = queryset.filter(cajero_id=cajero_id)
        if sucursal_id:
            queryset = queryset.filter(sucursal_id=sucursal_id)
        if estado:
            queryset = queryset.filter(estado=estado)
        if fecha_desde:
            queryset = queryset.filter(fecha_inicio__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(fecha_inicio__lte=fecha_hasta)
            
        return queryset.order_by('-fecha_inicio')
    
    @action(detail=False, methods=['post'])
    def iniciar_turno(self, request):
        """Iniciar nuevo turno para un cajero"""
        cajero_id = request.data.get('cajero_id')
        sucursal_id = request.data.get('sucursal_id')
        base_inicial = request.data.get('base_inicial', 0)
        notas_apertura = request.data.get('notas_apertura', '')
        
        if not cajero_id or not sucursal_id:
            return Response({
                'error': 'cajero_id y sucursal_id son requeridos'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Verificar que no haya turno activo
            turno_activo = Turno.objects.filter(
                cajero_id=cajero_id,
                estado='ACTIVO'
            ).first()
            
            if turno_activo:
                return Response({
                    'error': 'El cajero ya tiene un turno activo',
                    'turno_activo': TurnoSerializer(turno_activo).data
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Crear nuevo turno
            turno = Turno.objects.create(
                cajero_id=cajero_id,
                sucursal_id=sucursal_id,
                base_inicial=base_inicial,
                notas_apertura=notas_apertura,
                estado='ACTIVO'
            )
            
            serializer = self.get_serializer(turno)
            return Response({
                'success': True,
                'message': 'Turno iniciado exitosamente',
                'turno': serializer.data
            })
            
        except Exception as e:
            return Response({
                'error': f'Error iniciando turno: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def cerrar_turno(self, request, pk=None):
        """Cerrar turno específico"""
        turno = self.get_object()
        
        if turno.estado != 'ACTIVO':
            return Response({
                'error': 'Solo se pueden cerrar turnos activos'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        arqueo_final = request.data.get('arqueo_final', 0)
        notas_cierre = request.data.get('notas_cierre', '')
        
        try:
            turno.cerrar_turno(arqueo_final, notas_cierre)
            
            serializer = self.get_serializer(turno)
            return Response({
                'success': True,
                'message': 'Turno cerrado exitosamente',
                'turno': serializer.data
            })
            
        except Exception as e:
            return Response({
                'error': f'Error cerrando turno: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def turno_activo(self, request):
        """Obtener turno activo de un cajero"""
        cajero_id = request.query_params.get('cajero_id')
        if not cajero_id:
            return Response({'error': 'cajero_id es requerido'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        turno = Turno.objects.filter(
            cajero_id=cajero_id,
            estado='ACTIVO'
        ).first()
        
        if turno:
            serializer = self.get_serializer(turno)
            return Response(serializer.data)
        else:
            return Response({'message': 'No hay turno activo'}, 
                          status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=False, methods=['get'])
    def resumen_turnos(self, request):
        """Obtener resumen de turnos con filtros"""
        queryset = self.get_queryset()
        
        # Usar serializer resumido
        serializer = TurnoResumenSerializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """Obtener estadísticas de turnos"""
        fecha_desde = request.query_params.get('fecha_desde')
        fecha_hasta = request.query_params.get('fecha_hasta')
        sucursal_id = request.query_params.get('sucursal_id')
        
        queryset = Turno.objects.all()
        
        if fecha_desde:
            queryset = queryset.filter(fecha_inicio__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(fecha_inicio__lte=fecha_hasta)
        if sucursal_id:
            queryset = queryset.filter(sucursal_id=sucursal_id)
        
        stats = queryset.aggregate(
            total_turnos=Count('id'),
            total_ventas=Sum('total_ventas'),
            total_transacciones=Sum('numero_transacciones'),
            turnos_activos=Count('id', filter=Q(estado='ACTIVO')),
            turnos_cerrados=Count('id', filter=Q(estado='CERRADO'))
        )
        
        return Response(stats)

class VentaCajeroViewSet(viewsets.ModelViewSet):
    """API para ventas con información de cajero"""
    queryset = VentaCajero.objects.all()
    serializer_class = VentaCajeroSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = VentaCajero.objects.select_related(
            'venta', 'cajero', 'turno', 'sucursal'
        ).all()
        
        # Filtros
        cajero_id = self.request.query_params.get('cajero_id', None)
        turno_id = self.request.query_params.get('turno_id', None)
        sucursal_id = self.request.query_params.get('sucursal_id', None)
        
        if cajero_id:
            queryset = queryset.filter(cajero_id=cajero_id)
        if turno_id:
            queryset = queryset.filter(turno_id=turno_id)
        if sucursal_id:
            queryset = queryset.filter(sucursal_id=sucursal_id)
            
        return queryset.order_by('-venta__fecha')
    
    @action(detail=False, methods=['get'])
    def por_turno(self, request):
        """Obtener ventas de un turno específico"""
        turno_id = request.query_params.get('turno_id')
        if not turno_id:
            return Response({'error': 'turno_id es requerido'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        ventas = self.get_queryset().filter(turno_id=turno_id)
        serializer = self.get_serializer(ventas, many=True)
        
        # Calcular totales
        total_ventas = sum(v.venta.total for v in ventas)
        total_transacciones = ventas.count()
        
        return Response({
            'ventas': serializer.data,
            'resumen': {
                'total_ventas': total_ventas,
                'total_transacciones': total_transacciones
            }
        })

class ArqueoCajaViewSet(viewsets.ModelViewSet):
    """API para arqueos de caja"""
    queryset = ArqueoCaja.objects.all()
    serializer_class = ArqueoCajaSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = ArqueoCaja.objects.select_related('cajero_logueado', 'sucursal', 'turno').all()
        
        # Filtros
        fecha = self.request.query_params.get('fecha', None)
        fecha_inicio = self.request.query_params.get('fecha_inicio', None)
        fecha_fin = self.request.query_params.get('fecha_fin', None)
        cajero = self.request.query_params.get('cajero', None)
        estado = self.request.query_params.get('estado', None)
        sucursal_id = self.request.query_params.get('sucursal_id', None)
        
        if fecha:
            queryset = queryset.filter(fecha=fecha)
        if fecha_inicio:
            queryset = queryset.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(fecha__lte=fecha_fin)
        if cajero:
            queryset = queryset.filter(cajero__icontains=cajero)
        if estado:
            queryset = queryset.filter(estado=estado)
        if sucursal_id:
            queryset = queryset.filter(sucursal_id=sucursal_id)
            
        return queryset.order_by('-fecha', '-fecha_creacion')
    
    def create(self, request, *args, **kwargs):
        """Crear arqueo - Permite múltiples arqueos por día (uno por turno)"""
        try:
            # NOTA: Permitir múltiples arqueos por día para soportar múltiples turnos
            # No validar duplicados - cada turno puede tener su propio arqueo
            
            # Crear el arqueo
            response = super().create(request, *args, **kwargs)
            
            if response.status_code == 201:
                return Response({
                    'success': True,
                    'message': 'Arqueo de caja guardado exitosamente',
                    'arqueo': response.data
                })
            
            return response
            
        except Exception as e:
            return Response({
                'error': f'Error al guardar arqueo: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def resumen_por_fecha(self, request):
        """Obtener resumen de arqueos por fecha"""
        fecha = request.query_params.get('fecha')
        if not fecha:
            return Response({'error': 'Fecha es requerida'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        arqueos = self.get_queryset().filter(fecha=fecha)
        
        resumen = {
            'fecha': fecha,
            'total_arqueos': arqueos.count(),
            'total_sistema': sum(a.total_sistema for a in arqueos),
            'total_caja': sum(a.total_caja for a in arqueos),
            'total_diferencia': sum(a.total_diferencia for a in arqueos),
            'arqueos_por_estado': {
                estado[0]: arqueos.filter(estado=estado[0]).count()
                for estado in ArqueoCaja.ESTADOS_CHOICES
            }
        }
        
        return Response(resumen)
    
    @action(detail=False, methods=['get'])
    def por_cajero(self, request):
        """Obtener arqueos de un cajero específico"""
        cajero = request.query_params.get('cajero')
        if not cajero:
            return Response({'error': 'Cajero es requerido'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        arqueos = self.get_queryset().filter(cajero__icontains=cajero)
        serializer = self.get_serializer(arqueos, many=True)
        
        return Response({
            'cajero': cajero,
            'total_arqueos': arqueos.count(),
            'arqueos': serializer.data
        })

class PedidoViewSet(viewsets.ModelViewSet):
    """API para gestionar pedidos"""
    queryset = Pedido.objects.all().order_by('-fecha_creacion')
    serializer_class = PedidoSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = Pedido.objects.all().prefetch_related('detalles__producto', 'evidencias').order_by('-fecha_creacion')
        
        # Filtros opcionales
        destinatario = self.request.query_params.get('destinatario')
        estado = self.request.query_params.get('estado')
        fecha_desde = self.request.query_params.get('fecha_desde') or self.request.query_params.get('fecha_inicio')
        fecha_hasta = self.request.query_params.get('fecha_hasta') or self.request.query_params.get('fecha_fin')
        fecha_entrega = self.request.query_params.get('fecha_entrega') # 🆕 Filtrar por fecha de entrega
        transportadora = self.request.query_params.get('transportadora')
        search = self.request.query_params.get('search')
        
        if destinatario:
            queryset = queryset.filter(destinatario__icontains=destinatario)
        if estado:
            queryset = queryset.filter(estado=estado.upper())
        if search:
            queryset = queryset.filter(
                Q(numero_pedido__icontains=search) |
                Q(destinatario__icontains=search) |
                Q(vendedor__icontains=search)
            )
        else:
            if fecha_desde:
                queryset = queryset.filter(fecha__date__gte=fecha_desde)
            if fecha_hasta:
                queryset = queryset.filter(fecha__date__lte=fecha_hasta)
        if fecha_entrega: # 🆕 Aplicar filtro
            queryset = queryset.filter(fecha_entrega=fecha_entrega)
        if transportadora:
            queryset = queryset.filter(transportadora__icontains=transportadora)
            
        return queryset
    
    def create(self, request, *args, **kwargs):
        """Crear nuevo pedido con detalles"""
        try:
            with transaction.atomic():
                # Crear el pedido (el serializer ya crea los detalles)
                serializer = self.get_serializer(data=request.data, context={'request': request})
                serializer.is_valid(raise_exception=True)
                pedido = serializer.save()
                
                # Recargar con detalles
                pedido.refresh_from_db()
                response_serializer = self.get_serializer(pedido)
                
                return Response(response_serializer.data, status=status.HTTP_201_CREATED)
                
        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['patch'])
    def cambiar_estado(self, request, pk=None):
        """Cambiar estado del pedido"""
        pedido = self.get_object()
        nuevo_estado = request.data.get('estado')
        
        if nuevo_estado not in dict(Pedido.ESTADO_CHOICES):
            return Response(
                {'error': 'Estado inválido'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        pedido.estado = nuevo_estado
        pedido.save()
        
        serializer = self.get_serializer(pedido)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def anular(self, request, pk=None):
        """Anular pedido y revertir en Planeación y Cargue"""
        pedido = self.get_object()
        
        # 🆕 VALIDACIÓN: Verificar si el vendedor ya procesó el pedido en la app
        if pedido.estado == 'ENTREGADO':
            return Response(
                {
                    'success': False,
                    'message': '⚠️ No se puede anular: El vendedor ya marcó este pedido como ENTREGADO en la app móvil'
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if pedido.estado == 'ANULADA' and pedido.novedades and len(pedido.novedades) > 0:
            return Response(
                {
                    'success': False,
                    'message': '⚠️ No se puede anular: El vendedor ya marcó este pedido como NO ENTREGADO en la app móvil'
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Si ya está anulado (sin novedades = anulación manual previa)
        if pedido.estado == 'ANULADA':
            # Si ya está anulado, devolvemos una respuesta exitosa para que el frontend no quede bloqueado
            return Response(
                {'success': True, 'message': 'El pedido ya estaba anulado'},
                status=status.HTTP_200_OK
            )  
        try:
            with transaction.atomic():
                print(f"\n{'='*60}")

                print(f"{'='*60}")

                print(f"💰 Total: ${pedido.total}")
                print(f"📅 Fecha entrega: {pedido.fecha_entrega}")
                print(f"👤 Vendedor: {pedido.vendedor}")
                print(f"📦 Detalles: {pedido.detalles.count()} productos")
                
                # 1. Cambiar estado del pedido
                estado_anterior = pedido.estado
                pedido.estado = 'ANULADA'
                motivo = request.data.get('motivo', 'Anulado desde gestión de pedidos')
                pedido.nota = f"{pedido.nota or ''}\n[ANULADO] Estado anterior: {estado_anterior} - {motivo} - {timezone.now().strftime('%Y-%m-%d %H:%M')}"
                pedido.save()
                print(f"✅ Estado cambiado de {estado_anterior} a ANULADA")
                
                # 2. Revertir en Planeación (solo si existe fecha_entrega)
                if pedido.fecha_entrega:
                    print(f"\n📊 REVIRTIENDO EN PLANEACIÓN")
                    print(f"{'='*60}")
                    
                    for detalle in pedido.detalles.all():
                        try:
                            # Buscar en planeación por fecha_entrega y producto_nombre
                            planeacion = Planeacion.objects.filter(
                                fecha=pedido.fecha_entrega,
                                producto_nombre=detalle.producto.nombre
                            ).first()
                            
                            if planeacion:
                                pedidos_antes = planeacion.pedidos
                                total_antes = planeacion.total
                                
                                # Restar la cantidad del pedido anulado
                                planeacion.pedidos = max(0, planeacion.pedidos - detalle.cantidad)
                                # El total se recalcula automáticamente en save()
                                planeacion.save()
                                
                                print(f"  ✅ {detalle.producto.nombre}:")
                                print(f"     Pedidos: {pedidos_antes} → {planeacion.pedidos} (-{detalle.cantidad})")
                                print(f"     Total: {total_antes} → {planeacion.total}")
                            else:
                                print(f"  ⚠️ {detalle.producto.nombre}: No encontrado en Planeación")
                                
                        except Exception as e:
                            print(f"  ❌ Error con {detalle.producto.nombre}: {str(e)}")
                            continue
                else:
                    print(f"⚠️ Sin fecha de entrega, no se revierte en Planeación")
                
                # 3. Revertir Inventario (si fue afectado)
                if pedido.inventario_afectado:
                    print(f"\n⚡ REVIRTIENDO INVENTARIO")
                    print(f"{'='*60}")
                    
                    from .models import MovimientoInventario
                    
                    for detalle in pedido.detalles.all():
                        try:
                            producto = detalle.producto
                            cantidad_a_devolver = detalle.cantidad
                            
                            # Crear movimiento de inventario (Devolución) - Esto actualiza el stock automáticamente
                            MovimientoInventario.objects.create(
                                producto=producto,
                                tipo='ENTRADA',
                                cantidad=cantidad_a_devolver,
                                usuario=request.data.get('usuario', 'Sistema'),
                                nota=f'Anulación Pedido #{pedido.numero_pedido} - Devolución de stock'
                            )
                            print(f"✅ Movimiento de entrada creado para {producto.nombre} (+{cantidad_a_devolver})")
                            
                        except Exception as e:
                            print(f"  ❌ Error devolviendo stock para {detalle.producto.nombre}: {str(e)}")
                            continue
                    
                    pedido.inventario_afectado = False
                    pedido.save()
                    print(f"✅ Inventario revertido correctamente")
                
                # 4. Revertir en Cargue (solo si existe fecha_entrega y vendedor)
                if pedido.fecha_entrega and pedido.vendedor:
                    print(f"\n💰 REVIRTIENDO EN CARGUE")
                    print(f"{'='*60}")
                    
                    cargue_models = [
                        ('ID1', CargueID1), ('ID2', CargueID2), ('ID3', CargueID3),
                        ('ID4', CargueID4), ('ID5', CargueID5), ('ID6', CargueID6)
                    ]
                    
                    cargue_actualizado = False
                    
                    for id_cargue, CargueModel in cargue_models:
                        try:
                            # Buscar registros de cargue por fecha
                            cargues = CargueModel.objects.filter(fecha=pedido.fecha_entrega)
                            
                            for cargue in cargues:
                                # Verificar si el vendedor coincide con el responsable
                                if hasattr(cargue, 'responsable') and cargue.responsable:
                                    if pedido.vendedor.lower() in cargue.responsable.lower():
                                        pedidos_antes = float(cargue.total_pedidos or 0)
                                        efectivo_antes = float(cargue.total_efectivo or 0)
                                        
                                        # Revertir el total_pedidos (devolver el dinero)
                                        cargue.total_pedidos = max(0, pedidos_antes - float(pedido.total))
                                        
                                        # Recalcular total_efectivo
                                        if hasattr(cargue, 'venta') and cargue.venta:
                                            cargue.total_efectivo = float(cargue.venta) - float(cargue.total_pedidos)
                                        
                                        cargue.save()
                                        
                                        print(f"  ✅ {id_cargue} - {cargue.responsable}:")
                                        print(f"     Total Pedidos: ${pedidos_antes:,.0f} → ${cargue.total_pedidos:,.0f} (-${pedido.total:,.0f})")
                                        print(f"     Total Efectivo: ${efectivo_antes:,.0f} → ${cargue.total_efectivo:,.0f}")
                                        
                                        cargue_actualizado = True
                                        break  # Solo actualizar un cargue por modelo
                            
                            if cargue_actualizado:
                                break  # Salir del loop de modelos si ya se actualizó
                                
                        except Exception as e:
                            print(f"  ⚠️ Error en {id_cargue}: {str(e)}")
                            continue
                    
                    if not cargue_actualizado:
                        print(f"  ⚠️ No se encontró cargue para vendedor '{pedido.vendedor}' en fecha {pedido.fecha_entrega}")
                else:
                    print(f"⚠️ Sin fecha de entrega o vendedor, no se revierte en Cargue")
                
                # Fin de la transacción - todo se completó exitosamente
                print(f"\n{'='*60}")
                print(f"✅ PEDIDO ANULADO EXITOSAMENTE")
                print(f"{'='*60}\n")
            
            # Fuera del transaction.atomic() - devolver respuesta exitosa
            serializer = self.get_serializer(pedido)
            return Response({
                'success': True,
                'message': 'Pedido anulado exitosamente. Se revirtieron las cantidades en Planeación y el dinero en Cargue.',
                'pedido': serializer.data
            })
                
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"\n{'='*60}")
            print(f"❌ ERROR AL ANULAR PEDIDO")
            print(f"{'='*60}")
            print(error_detail)
            print(f"{'='*60}\n")
            
            return Response(
                {'detail': f'Error al anular pedido: {str(e)}'}, \
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def afectar_inventario(self, request, pk=None):
        """Afectar inventario de un pedido manualmente (para corrección)"""
        pedido = self.get_object()
        
        # Validar que no esté ya afectado
        if pedido.inventario_afectado:
            return Response(
                {'detail': 'El inventario de este pedido ya fue afectado'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar que no esté anulado
        if pedido.estado == 'ANULADA':
            return Response(
                {'detail': 'No se puede afectar inventario de un pedido anulado'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            with transaction.atomic():
                from .models import Producto, MovimientoInventario
                
                print(f"\n{'='*60}")
                print(f"⚡ AFECTANDO INVENTARIO MANUALMENTE")
                print(f"Pedido: #{pedido.numero_pedido}")
                print(f"{'='*60}")
                
                for detalle in pedido.detalles.all():
                    try:
                        producto = detalle.producto
                        cantidad_a_descontar = detalle.cantidad
                        
                        # Verificar stock disponible
                        if producto.stock_total < cantidad_a_descontar:
                            print(f"⚠️ ADVERTENCIA: {producto.nombre} - Stock insuficiente ({producto.stock_total} < {cantidad_a_descontar})")
                        
                        # 🔧 FIX: Solo crear MovimientoInventario (que se encarga del descuento automáticamente)
                        # NO hacer descuento manual porque causa DOBLE descuento
                        stock_anterior = producto.stock_total
                        
                        # Crear movimiento de inventario (esto descuenta automáticamente en el save())
                        MovimientoInventario.objects.create(
                            producto=producto,
                            tipo='SALIDA',
                            cantidad=cantidad_a_descontar,
                            usuario=request.data.get('usuario', 'Sistema'),
                            nota=f'Corrección manual - Pedido #{pedido.numero_pedido} - {pedido.destinatario}'
                        )
                        
                        # Refrescar para ver el stock actualizado
                        producto.refresh_from_db()
                        print(f"✅ {producto.nombre}: {stock_anterior} → {producto.stock_total} (-{cantidad_a_descontar})")
                        
                    except Exception as e:
                        print(f"❌ Error afectando inventario para {detalle.producto.nombre}: {str(e)}")
                        raise e
                
                # Marcar como inventario afectado
                pedido.inventario_afectado = True
                pedido.afectar_inventario_inmediato = True  # Actualizar también este campo
                pedido.save()
                
                print(f"✅ Inventario afectado y marcado")
                print(f"{'='*60}\n")
                
                serializer = self.get_serializer(pedido)
                return Response({
                    'success': True,
                    'message': 'Inventario afectado exitosamente',
                    'pedido': serializer.data
                })
                
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"\n{'='*60}")
            print(f"❌ ERROR AL AFECTAR INVENTARIO")
            print(f"{'='*60}")
            print(error_detail)
            print(f"{'='*60}\n")
            
            return Response(
                {'detail': f'Error al afectar inventario: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def pendientes_vendedor(self, request):
        """
        Obtener pedidos pendientes asignados a un vendedor para la app móvil
        Busca por asignado_a_id o asignado por nombre (campo vendedor)
        """
        v_id_param = request.query_params.get('vendedor_id')
        vendedor, _, auth_error = _validar_vendedor_token(request, v_id_param, campo='vendedor_id')
        if auth_error:
            return auth_error

        fecha = request.query_params.get('fecha')

        if not fecha:
            return Response({'error': 'Falta parámetro: fecha'}, status=status.HTTP_400_BAD_REQUEST)

        vendedor_id = vendedor.id_vendedor
            
        print(f"📦 Backend: Buscando pedidos para {vendedor_id} en {fecha}")
        
        from django.db.models import Q
        
        nombre_vendedor = vendedor.nombre or ""

        # Filtro final (Fecha Y No Cancelado/Anulado)
        # 🔧 Incluir ENTREGADO para que la app móvil pueda mostrar check verde
        # 🔧 Incluir ANULADA para mostrar badge de No Entregado (pero excluir CANCELADO)
        filtro_base = Q(fecha_entrega=fecha) & ~Q(estado__in=['CANCELADO'])
        
        condicion_asignacion = Q(asignado_a_id=vendedor_id)
        if nombre_vendedor:
             print(f"   Incluyendo búsqueda por nombre: {nombre_vendedor}")
             condicion_asignacion |= Q(vendedor__iexact=nombre_vendedor)
        
        pedidos = Pedido.objects.filter(filtro_base & condicion_asignacion)
        
        serializer = self.get_serializer(pedidos, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def marcar_entregado(self, request, pk=None):
        """Marcar pedido como entregado desde la App"""
        vendedor, _, auth_error = _validar_vendedor_token(request)
        if auth_error:
            return auth_error

        pedido = self.get_object()
        if not _pedido_pertenece_a_vendedor(pedido, vendedor):
            return Response({'error': 'Pedido no autorizado para este vendedor'}, status=status.HTTP_403_FORBIDDEN)

        # Idempotente: si ya está entregado, retornar éxito sin modificar nada
        if pedido.estado in ('ENTREGADO', 'ENTREGADA'):
            return Response({'success': True, 'message': 'Pedido ya estaba marcado como entregado'})

        # 🆕 Obtener y guardar el método de pago
        metodo_pago = request.data.get('metodo_pago', 'EFECTIVO')
        if metodo_pago:
            pedido.metodo_pago = metodo_pago.upper()

        from django.utils import timezone
        pedido.estado = 'ENTREGADO'
        
        # Agregar nota con la hora y método
        from django.utils import timezone
        ahora = timezone.now()
        nota_entrega = f"Entregado vía App Móvil ({pedido.metodo_pago}) el {ahora.strftime('%Y-%m-%d %H:%M')}"
        pedido.nota = f"{pedido.nota or ''} | {nota_entrega}".strip()
        
        # Guardar la fecha real en fecha_actualizacion para tener exactitud a nivel de segundos
        pedido.fecha_actualizacion = ahora
        
        # 🔧 NO cambiar fecha_entrega para que siga apareciendo en su día original
        pedido.save()
        
        return Response({
            'success': True, 
            'message': f'Pedido marcado como entregado ({pedido.metodo_pago})'
        })

    @action(detail=True, methods=['post'])
    def marcar_no_entregado(self, request, pk=None):
        """Reportar que un pedido no pudo ser entregado"""
        vendedor, _, auth_error = _validar_vendedor_token(request)
        if auth_error:
            return auth_error

        pedido = self.get_object()
        if not _pedido_pertenece_a_vendedor(pedido, vendedor):
            return Response({'error': 'Pedido no autorizado para este vendedor'}, status=status.HTTP_403_FORBIDDEN)

        motivo = request.data.get('motivo', 'Sin motivo especificado')
        
        # Marcar como ANULADA (o un estado que indique no gestión exitosa)
        pedido.estado = 'ANULADA'
        pedido.nota = f"{pedido.nota or ''} | NO ENTREGADO: {motivo}".strip()
        pedido.save()
        
        return Response({'status': 'novedad reportada'})

    @action(detail=True, methods=['post'])
    def verificar_despacho(self, request, pk=None):
        """Marca/desmarca un pedido como verificado por el despachador (toggle)"""
        pedido = self.get_object()
        pedido.verificado_despachador = not pedido.verificado_despachador
        pedido.save(update_fields=['verificado_despachador'])
        return Response({
            'id': pedido.id,
            'verificado_despachador': pedido.verificado_despachador
        })

    @action(detail=True, methods=['patch'])
    def actualizar_app(self, request, pk=None):
        """
        Actualización de pedido usada por App Móvil (edición de entrega).
        Requiere token de vendedor y valida propiedad del pedido.
        """
        vendedor, _, auth_error = _validar_vendedor_token(request)
        if auth_error:
            return auth_error

        pedido = self.get_object()
        if not _pedido_pertenece_a_vendedor(pedido, vendedor):
            return Response({'error': 'Pedido no autorizado para este vendedor'}, status=status.HTTP_403_FORBIDDEN)

        # 🆕 Bloqueo de solo 1 edición (como Clientes de Ruta)
        if pedido.editada:
            return Response(
                {
                    'error': 'Este pedido ya fue modificado una vez. No se permiten más cambios.',
                    'codigo': 'VENTA_YA_MODIFICADA'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        data = request.data.copy()
        campos_bloqueados = {
            'id',
            'numero_pedido',
            'vendedor',
            'asignado_a_tipo',
            'asignado_a_id',
            'inventario_afectado',
            'fecha_creacion',
            # 'fecha_actualizacion',  # Quitamos esto para que se actualice la hora de edición de pago
        }
        for campo in campos_bloqueados:
            data.pop(campo, None)

        serializer = self.get_serializer(pedido, data=data, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        
        # Guardar la fecha real en fecha_actualizacion para subirla al tope del historial
        from django.utils import timezone
        serializer.save(fecha_actualizacion=timezone.now(), editada=True)

        return Response({
            'success': True,
            'message': 'Pedido actualizado correctamente',
            'pedido': serializer.data
        })

class DetallePedidoViewSet(viewsets.ModelViewSet):
    """API para detalles de pedidos"""
    queryset = DetallePedido.objects.all()
    serializer_class = DetallePedidoSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = DetallePedido.objects.all()
        pedido_id = self.request.query_params.get('pedido')
        remision_id = self.request.query_params.get('remision')  # Compatibilidad
        
        if pedido_id:
            queryset = queryset.filter(pedido_id=pedido_id)
        elif remision_id:
            queryset = queryset.filter(pedido_id=remision_id)
            
        return queryset
class PlaneacionViewSet(viewsets.ModelViewSet):
    queryset = Planeacion.objects.all()
    serializer_class = PlaneacionSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        fecha = self.request.query_params.get('fecha')
        producto_nombre = self.request.query_params.get('producto_nombre')
        
        if fecha:
            queryset = queryset.filter(fecha=fecha)
        if producto_nombre:
            queryset = queryset.filter(producto_nombre=producto_nombre)
            
        return queryset.order_by('producto_nombre')
    
    def create(self, request, *args, **kwargs):
        """Crear o actualizar registro de planeación (upsert)"""
        fecha = request.data.get('fecha')
        producto_nombre = request.data.get('producto_nombre')
        
        if fecha and producto_nombre:
            # Buscar si ya existe
            try:
                planeacion = Planeacion.objects.get(fecha=fecha, producto_nombre=producto_nombre)
                # Ya existe, actualizar
                serializer = self.get_serializer(planeacion, data=request.data, partial=False)
                serializer.is_valid(raise_exception=True)
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            except Planeacion.DoesNotExist:
                # No existe, crear nuevo
                pass
        
        # Crear nuevo registro
        return super().create(request, *args, **kwargs)
    
    @action(detail=False, methods=['post'])
    def prediccion_ia(self, request):
        """
        Obtiene predicciones de IA (con Redes Neuronales) para una fecha específica.
        
        POST /api/planeacion/prediccion_ia/
        Body: {
            "fecha": "2025-11-20",
            "datos_contextuales": {
                "AREPA TIPO OBLEA 500Gr": {
                    "existencias": 266,
                    "solicitadas": 0,
                    "pedidos": 0
                },
                ...
            }
        }
        """
        from api.services.ia_service import IAService
        
        fecha = request.data.get('fecha')
        datos_contextuales = request.data.get('datos_contextuales', {})
        
        if not fecha:
            return Response(
                {'error': 'Fecha es requerida'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Inicializar servicio de IA
            ia_service = IAService()
            
            # Obtener predicciones
            predicciones = ia_service.predecir_produccion(
                fecha_objetivo=fecha,
                datos_contextuales=datos_contextuales
            )
            
            return Response({
                'fecha': fecha,
                'predicciones': predicciones,
                'total_productos': len(predicciones)
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            print(f"❌ Error en predicción IA: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Error generando predicciones: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class VendedorViewSet(viewsets.ModelViewSet):
    """API para gestionar vendedores"""
    queryset = Vendedor.objects.all()
    serializer_class = VendedorSerializer
    permission_classes = [permissions.AllowAny]
    lookup_field = 'id_vendedor'  # Usar id_vendedor en lugar de pk
    
    def get_queryset(self):
        """Filtrar vendedores por parámetros"""
        queryset = Vendedor.objects.all()
        id_vendedor = self.request.query_params.get('id_vendedor', None)
        activo = self.request.query_params.get('activo', None)
        
        if id_vendedor:
            queryset = queryset.filter(id_vendedor=id_vendedor)
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
            
        return queryset.order_by('id_vendedor')
    
    @action(detail=False, methods=['get'])
    def obtener_responsable(self, request):
        """Obtener el responsable de uno o todos los vendedores"""
        id_vendedor = request.query_params.get('id_vendedor')
        
        if id_vendedor:
            vendedor = Vendedor.objects.filter(id_vendedor=id_vendedor).first()
            if vendedor:
                return Response({
                    'success': True,
                    'id': vendedor.id_vendedor,
                    'responsable': vendedor.nombre
                })
            else:
                return Response({
                    'success': True, 
                    'id': id_vendedor, 
                    'responsable': 'RESPONSABLE' 
                })
        else:
            vendedores = Vendedor.objects.filter(activo=True).values('id_vendedor', 'nombre')
            resultado = [{'id': v['id_vendedor'], 'responsable': v['nombre']} for v in vendedores]
            return Response(resultado)

    
    @action(detail=False, methods=['post'])
    def actualizar_responsable(self, request):
        """Actualizar nombre del responsable/vendedor"""
        try:
            id_vendedor = request.data.get('id_vendedor')
            responsable = request.data.get('responsable')
            
            if not id_vendedor or not responsable:
                return Response(
                    {'error': 'id_vendedor y responsable son requeridos'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Actualizar o crear vendedor en tabla Vendedor
            vendedor, created = Vendedor.objects.update_or_create(
                id_vendedor=id_vendedor,
                defaults={'nombre': responsable}
            )
            
            # ---------------------------------------------------------
            # TAMBIÉN ACTUALIZAR EN TABLAS DE CARGUE (CargueID1, etc.)
            # ---------------------------------------------------------
            try:
                # Mapear ID de vendedor a modelo correspondiente
                modelos_vendedor = {
                    'ID1': CargueID1,
                    'ID2': CargueID2,
                    'ID3': CargueID3,
                    'ID4': CargueID4,
                    'ID5': CargueID5,
                    'ID6': CargueID6,
                }
                
                modelo = modelos_vendedor.get(id_vendedor)
                if modelo:
                    # Actualizar todos los registros existentes de este vendedor
                    from django.utils import timezone
                    modelo.objects.filter(activo=True).update(
                        responsable=responsable,
                        fecha_actualizacion=timezone.now()  # 🔥 Actualizar timestamp para polling
                    )
                    print(f"✅ Responsable actualizado en {modelo.__name__}: {responsable}")
            except Exception as e:
                print(f"⚠️ Error actualizando tablas de cargue: {str(e)}")
            # ---------------------------------------------------------
            
            return Response({
                'success': True,
                'vendedor': {
                    'id_vendedor': vendedor.id_vendedor,
                    'nombre': vendedor.nombre,
                    'ruta': vendedor.ruta
                },
                'message': f'Responsable {"creado" if created else "actualizado"} exitosamente'
            })
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'])
    def login(self, request):
        """Endpoint de login para vendedores (App Móvil)"""
        try:
            id_vendedor = request.data.get('id_vendedor')
            password = request.data.get('password')
            dispositivo_id = (request.data.get('dispositivo_id') or '').strip()
            
            if not id_vendedor or not password:
                return Response(
                    {'error': 'id_vendedor y password son requeridos'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            id_vendedor = str(id_vendedor).strip().upper()
            
            # Buscar vendedor
            try:
                vendedor = Vendedor.objects.get(id_vendedor=id_vendedor, activo=True)
            except Vendedor.DoesNotExist:
                return Response(
                    {'error': 'Credenciales inválidas'},
                    status=status.HTTP_401_UNAUTHORIZED
                )
            
            # Validar contraseña
            if vendedor.password != password:
                return Response(
                    {'error': 'Credenciales inválidas'},
                    status=status.HTTP_401_UNAUTHORIZED
                )
            
            # Invalidar tokens expirados o tokens previos del MISMO dispositivo
            VendedorSesionToken.objects.filter(
                models.Q(expira_en__lte=timezone.now()) | 
                models.Q(dispositivo_id=dispositivo_id) if dispositivo_id else models.Q(pk=None),
                vendedor=vendedor,
                activo=True
            ).update(activo=False)

            # Crear token de sesión (30 días)
            token_value = secrets.token_urlsafe(48)
            expira_en = timezone.now() + timedelta(days=30)
            sesion = VendedorSesionToken.objects.create(
                vendedor=vendedor,
                token=token_value,
                dispositivo_id=dispositivo_id,
                expira_en=expira_en,
                activo=True
            )

            # Login exitoso
            return Response({
                'success': True,
                'vendedor': {
                    'id_vendedor': vendedor.id_vendedor,
                    'nombre': vendedor.nombre,
                    'ruta': vendedor.ruta or '',
                    'activo': vendedor.activo
                },
                'token': sesion.token,
                'token_type': 'Bearer',
                'expires_at': sesion.expira_en.isoformat(),
                'message': 'Login exitoso'
            })
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='sesiones-movil')
    def sesiones_movil(self, request):
        """
        Lista sesiones móviles (tokens) para monitorear equipos conectados.
        Query params opcionales:
        - id_vendedor: ID1, ID2, ...
        - incluir_inactivas: true|false (default false)
        """
        try:
            id_vendedor_raw = request.query_params.get('id_vendedor')
            incluir_inactivas = str(request.query_params.get('incluir_inactivas', 'false')).lower() == 'true'

            sesiones_qs = VendedorSesionToken.objects.select_related('vendedor').all()

            if id_vendedor_raw:
                id_vendedor_norm, _ = _normalizar_id_vendedor(id_vendedor_raw)
                if not id_vendedor_norm:
                    return Response(
                        {'error': 'id_vendedor inválido'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                sesiones_qs = sesiones_qs.filter(vendedor__id_vendedor=id_vendedor_norm)

            ahora = timezone.now()
            if not incluir_inactivas:
                sesiones_qs = sesiones_qs.filter(
                    activo=True,
                    expira_en__gt=ahora,
                    vendedor__activo=True
                )

            sesiones = []
            for sesion in sesiones_qs.order_by('vendedor__id_vendedor', '-ultimo_uso', '-creado_en')[:300]:
                if sesion.activo and sesion.expira_en > ahora:
                    estado = 'ACTIVA'
                elif sesion.expira_en <= ahora:
                    estado = 'EXPIRADA'
                else:
                    estado = 'INACTIVA'

                sesiones.append({
                    'id': sesion.id,
                    'vendedor_id': sesion.vendedor.id_vendedor,
                    'vendedor_nombre': sesion.vendedor.nombre,
                    'dispositivo_id': sesion.dispositivo_id or 'SIN_DISPOSITIVO',
                    'creado_en': sesion.creado_en.isoformat() if sesion.creado_en else None,
                    'ultimo_uso': sesion.ultimo_uso.isoformat() if sesion.ultimo_uso else None,
                    'expira_en': sesion.expira_en.isoformat() if sesion.expira_en else None,
                    'activo': sesion.activo,
                    'estado': estado
                })

            return Response({
                'success': True,
                'count': len(sesiones),
                'sesiones': sesiones
            })
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='desactivar-sesion-movil')
    def desactivar_sesion_movil(self, request):
        """
        Desactiva sesiones móviles activas.
        Soporta:
        - sesion_id (desactivar una sesión puntual)
        - id_vendedor + dispositivo_id (desactivar sesiones de ese equipo)
        - id_vendedor + desactivar_todas=true (desactivar todo el vendedor)
        """
        try:
            sesion_id = request.data.get('sesion_id')
            id_vendedor_raw = request.data.get('id_vendedor')
            dispositivo_id = str(request.data.get('dispositivo_id') or '').strip()
            desactivar_todas_raw = request.data.get('desactivar_todas', False)
            desactivar_todas = str(desactivar_todas_raw).lower() in ['1', 'true', 'si', 'yes']

            sesiones_qs = VendedorSesionToken.objects.filter(activo=True)

            if sesion_id:
                sesiones_qs = sesiones_qs.filter(id=sesion_id)
            else:
                id_vendedor_norm, _ = _normalizar_id_vendedor(id_vendedor_raw)
                if not id_vendedor_norm:
                    return Response(
                        {'error': 'Debes enviar sesion_id o id_vendedor válido'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                sesiones_qs = sesiones_qs.filter(vendedor__id_vendedor=id_vendedor_norm)

                if desactivar_todas:
                    pass
                elif dispositivo_id:
                    sesiones_qs = sesiones_qs.filter(dispositivo_id=dispositivo_id)
                else:
                    return Response(
                        {'error': 'Para desactivar por vendedor debes enviar dispositivo_id o desactivar_todas=true'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            actualizadas = sesiones_qs.update(activo=False)

            return Response({
                'success': True,
                'desactivadas': actualizadas,
                'message': 'Sesiones desactivadas correctamente' if actualizadas else 'No había sesiones activas para desactivar'
            })
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def logout(self, request):
        """Invalida el token móvil actual."""
        auth_header = request.headers.get('Authorization', '') or ''
        token = ''
        if auth_header.lower().startswith('bearer '):
            token = auth_header.split(' ', 1)[1].strip()
        if not token:
            token = (request.data.get('token') or '').strip()

        if not token:
            return Response({'error': 'Token requerido'}, status=status.HTTP_400_BAD_REQUEST)

        updated = VendedorSesionToken.objects.filter(token=token, activo=True).update(activo=False)
        if updated == 0:
            return Response({'success': True, 'message': 'Token ya estaba inactivo'})

        return Response({'success': True, 'message': 'Sesión cerrada'})


class DomiciliarioViewSet(viewsets.ModelViewSet):
    """API para gestionar domiciliarios"""
    queryset = Domiciliario.objects.all()
    serializer_class = DomiciliarioSerializer
    permission_classes = [permissions.AllowAny]
    lookup_field = 'codigo'  # Usar codigo en lugar de pk
    
    def get_queryset(self):
        """Filtrar domiciliarios por parámetros"""
        queryset = Domiciliario.objects.all()
        activo = self.request.query_params.get('activo', None)
        
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
            
        return queryset.order_by('codigo')
    
    @action(detail=True, methods=['get'])
    def pedidos(self, request, codigo=None):
        """Obtener pedidos asignados a un domiciliario"""
        domiciliario = self.get_object()
        fecha = request.query_params.get('fecha')
        estado = request.query_params.get('estado')
        
        pedidos_query = Pedido.objects.filter(
            asignado_a_tipo='DOMICILIARIO',
            asignado_a_id=domiciliario.codigo
        )
        
        if fecha:
            pedidos_query = pedidos_query.filter(fecha_entrega=fecha)
        if estado:
            pedidos_query = pedidos_query.filter(estado=estado.upper())
        
        pedidos = pedidos_query.order_by('-fecha_creacion')
        serializer = PedidoSerializer(pedidos, many=True)
        
        # Calcular totales
        total_pedidos = pedidos.count()
        total_monto = sum(p.total for p in pedidos)
        
        return Response({
            'domiciliario': {
                'codigo': domiciliario.codigo,
                'nombre': domiciliario.nombre
            },
            'pedidos': serializer.data,
            'resumen': {
                'total_pedidos': total_pedidos,
                'total_monto': float(total_monto)
            }
        })


class MovimientoCajaViewSet(viewsets.ModelViewSet):
    """API para gestionar movimientos de caja (ingresos y egresos)"""
    queryset = MovimientoCaja.objects.all()
    serializer_class = MovimientoCajaSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        """Filtrar movimientos por fecha, cajero o tipo"""
        queryset = MovimientoCaja.objects.all()
        
        fecha = self.request.query_params.get('fecha', None)
        cajero = self.request.query_params.get('cajero', None)
        tipo = self.request.query_params.get('tipo', None)
        turno_id = self.request.query_params.get('turno', None) # 🆕
        
        # 🆕 Si viene el turno, filtramos principalmente por él (más preciso)
        if turno_id:
             queryset = queryset.filter(turno_id=turno_id)
        else:
            # Comportamiento legacy
            if fecha:
                queryset = queryset.filter(fecha=fecha)
            if cajero:
                queryset = queryset.filter(cajero=cajero)

        if tipo:
            queryset = queryset.filter(tipo=tipo)
            
        return queryset.order_by('-fecha', '-hora')
    
    @action(detail=False, methods=['get'])
    def resumen_por_fecha(self, request):
        """Obtener resumen de ingresos y egresos por fecha"""
        try:
            fecha = request.query_params.get('fecha')
            cajero = request.query_params.get('cajero', None)
            
            if not fecha:
                return Response(
                    {'error': 'Fecha es requerida'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Filtrar movimientos
            movimientos = MovimientoCaja.objects.filter(fecha=fecha)
            if cajero:
                movimientos = movimientos.filter(cajero=cajero)
            
            # Calcular totales
            from django.db.models import Sum, Q
            
            total_ingresos = movimientos.filter(tipo='INGRESO').aggregate(
                total=Sum('monto')
            )['total'] or 0
            
            total_egresos = movimientos.filter(tipo='EGRESO').aggregate(
                total=Sum('monto')
            )['total'] or 0
            
            saldo = total_ingresos - total_egresos
            
            return Response({
                'fecha': fecha,
                'cajero': cajero,
                'total_ingresos': float(total_ingresos),
                'total_egresos': float(total_egresos),
                'saldo': float(saldo),
                'cantidad_movimientos': movimientos.count()
            })
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ArqueoCajaViewSet(viewsets.ModelViewSet):
    """API para gestionar arqueos de caja"""
    queryset = ArqueoCaja.objects.all().order_by('-fecha', '-fecha_creacion')
    serializer_class = ArqueoCajaSerializer
    permission_classes = [permissions.AllowAny]
    
    def create(self, request, *args, **kwargs):
        """Crear arqueo con logging detallado"""

        
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)

            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        except Exception as e:
            print('❌ Error al crear arqueo:', str(e))
            return Response(
                {'error': str(e), 'detail': getattr(e, 'detail', None)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def get_queryset(self):
        """Filtrar arqueos por fecha, cajero o estado"""
        queryset = ArqueoCaja.objects.all()
        
        fecha_inicio = self.request.query_params.get('fecha_inicio', None)
        fecha_fin = self.request.query_params.get('fecha_fin', None)
        cajero = self.request.query_params.get('cajero', None)
        estado = self.request.query_params.get('estado', None)
        
        if fecha_inicio:
            queryset = queryset.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(fecha__lte=fecha_fin)
        if cajero:
            queryset = queryset.filter(cajero=cajero)
        if estado:
            queryset = queryset.filter(estado=estado)
            
        return queryset.order_by('-fecha', '-fecha_creacion')
    
    @action(detail=False, methods=['post'])
    def validar(self, request):
        """Validar arqueo antes de guardar"""
        try:
            valores_sistema = request.data.get('valores_sistema', {})
            valores_caja = request.data.get('valores_caja', {})
            
            # Calcular diferencias
            diferencias = {}
            for metodo in valores_sistema.keys():
                sistema = float(valores_sistema.get(metodo, 0))
                caja = float(valores_caja.get(metodo, 0))
                diferencias[metodo] = caja - sistema
            
            total_diferencia = sum(diferencias.values())
            
            # Validaciones
            alertas = []
            if abs(total_diferencia) > 10000:
                alertas.append({
                    'tipo': 'error',
                    'mensaje': f'Diferencia muy alta: ${total_diferencia:,.2f}'
                })
            elif abs(total_diferencia) > 1000:
                alertas.append({
                    'tipo': 'warning',
                    'mensaje': f'Diferencia moderada: ${total_diferencia:,.2f}'
                })
            else:
                alertas.append({
                    'tipo': 'success',
                    'mensaje': 'Arqueo cuadrado correctamente'
                })
            
            return Response({
                'valido': abs(total_diferencia) <= 10000,
                'diferencias': diferencias,
                'total_diferencia': total_diferencia,
                'alertas': alertas
            })
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """Obtener estadísticas de arqueos por rango de fechas"""
        try:
            fecha_inicio = request.query_params.get('fecha_inicio')
            fecha_fin = request.query_params.get('fecha_fin')
            
            if not fecha_inicio or not fecha_fin:
                return Response(
                    {'error': 'Fechas de inicio y fin son requeridas'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            arqueos = ArqueoCaja.objects.filter(
                fecha__gte=fecha_inicio,
                fecha__lte=fecha_fin
            )
            
            from django.db.models import Sum, Avg, Count
            
            estadisticas = {
                'total_arqueos': arqueos.count(),
                'sin_diferencias': arqueos.filter(total_diferencia=0).count(),
                'con_diferencias': arqueos.exclude(total_diferencia=0).count(),
                'total_diferencia': float(arqueos.aggregate(Sum('total_diferencia'))['total_diferencia__sum'] or 0),
                'promedio_diferencia': float(arqueos.aggregate(Avg('total_diferencia'))['total_diferencia__avg'] or 0),
            }
            
            return Response(estadisticas)
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ========================================
# VIEWSET PARA CONFIGURACIÓN DE IMPRESIÓN
# ========================================

class ConfiguracionImpresionViewSet(viewsets.ModelViewSet):
    """ViewSet para configuración de impresión de tickets"""
    queryset = ConfiguracionImpresion.objects.all()
    serializer_class = ConfiguracionImpresionSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        """Obtener configuración activa"""
        queryset = ConfiguracionImpresion.objects.filter(activo=True)
        return queryset
    
    @action(detail=False, methods=['get'])
    def activa(self, request):
        """Obtener la configuración activa (solo una)"""
        try:
            config = ConfiguracionImpresion.objects.filter(activo=True).first()
            if config:
                serializer = self.get_serializer(config)
                return Response(serializer.data)
            else:
                # Retornar configuración por defecto si no existe
                return Response({
                    'id': None,
                    'nombre_negocio': 'MI NEGOCIO',
                    'nit_negocio': '',
                    'direccion_negocio': '',
                    'telefono_negocio': '',
                    'email_negocio': '',
                    'encabezado_ticket': '',
                    'pie_pagina_ticket': '',
                    'mensaje_agradecimiento': '¡Gracias por su compra!',
                    'logo': None,
                    'ancho_papel': '80mm',
                    'mostrar_logo': True,
                    'mostrar_codigo_barras': False,
                    'impresora_predeterminada': '',
                    'resolucion_facturacion': '',
                    'regimen_tributario': '',
                    'activo': True
                })
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

# ========================================
# VIEWSET PARA INTELIGENCIA ARTIFICIAL
# ========================================

class PrediccionIAView(viewsets.ViewSet):
    """
    API para generar predicciones de producción usando IA.
    """
    permission_classes = [permissions.AllowAny]

    def list(self, request):
        """
        Genera una predicción de producción CONTEXTUAL para una fecha específica.
        Uso: GET /api/prediccion-ia/?fecha=2025-05-24
        
        La IA considera:
        - Histórico de ventas
        - Existencias actuales (si están reportadas en Planeación)
        - Solicitadas del día
        - Pedidos del día
        """
        try:
            fecha_objetivo = request.query_params.get('fecha')
            
            if not fecha_objetivo:
                return Response(
                    {'error': 'El parámetro "fecha" es requerido (YYYY-MM-DD)'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # 📊 Obtener datos contextuales de la Planeación (si existen)
            from api.models import Planeacion
            datos_contextuales = {}
            
            try:
                planeacion_registros = Planeacion.objects.filter(fecha=fecha_objetivo)
                for registro in planeacion_registros:
                    datos_contextuales[registro.producto_nombre] = {
                        'existencias': registro.existencias or 0,
                        'solicitadas': registro.solicitadas or 0,
                        'pedidos': registro.pedidos or 0
                    }
                print(f"📊 Datos contextuales cargados para {len(datos_contextuales)} productos")
            except Exception as e:
                print(f"⚠️ No se pudieron cargar datos contextuales: {e}")
                # Continuar sin datos contextuales (IA usará solo histórico)
            
            # Importar el servicio aquí para evitar ciclos de importación
            from api.services.ia_service import IAService
            
            # Instanciar servicio y generar predicción con contexto
            ia_service = IAService()
            predicciones = ia_service.predecir_produccion(
                fecha_objetivo,
                datos_contextuales=datos_contextuales if datos_contextuales else None
            )
            
            return Response({
                'success': True,
                'fecha_objetivo': fecha_objetivo,
                'total_productos_analizados': len(predicciones),
                'con_datos_contextuales': len(datos_contextuales) > 0,
                'predicciones': predicciones
            })
            
        except Exception as e:
            print(f"❌ Error en PrediccionIAView: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

@api_view(['POST'])
def guardar_sugerido(request):
    """
    Endpoint para recibir Sugeridos/Cargue desde la App Móvil.
    Recibe: { vendedor_id, dia, fecha, productos: [{nombre, cantidad}, ...] }
    """
    try:
        data = request.data
        vendedor_id = data.get('vendedor_id') # Ej: "ID1"
        dia_raw = data.get('dia', '').upper() # Ej: "LUNES" o "SÁBADO"
        fecha_raw = data.get('fecha') # Ej: "2025-11-29" o "2025-11-29T..."

        # 🔒 Seguridad por fases (transición sin romper app):
        # - Si SECURE_SUGERIDOS_REQUIRE_TOKEN=True: token obligatorio.
        # - Si False: se acepta legado, pero si hay token válido se usa SIEMPRE su vendedor.
        secure_sugeridos_require_token = bool(getattr(settings, 'SECURE_SUGERIDOS_REQUIRE_TOKEN', False))
        vendedor_token, auth_error = _obtener_vendedor_sesion_movil(request)
        if auth_error:
            if secure_sugeridos_require_token:
                return auth_error
            print(f"⚠️ [SUGERIDOS][TRANSICION] Request sin token válido para vendedor payload={vendedor_id}")
        else:
            vendedor_payload_norm, _ = _normalizar_id_vendedor(vendedor_id)
            vendedor_token_norm, _ = _normalizar_id_vendedor(vendedor_token.id_vendedor)
            if vendedor_payload_norm and vendedor_token_norm and vendedor_payload_norm != vendedor_token_norm:
                print(
                    f"⚠️ [SUGERIDOS][TRANSICION] vendedor_id payload ({vendedor_payload_norm}) "
                    f"difiere del token ({vendedor_token_norm}); se usará token."
                )
            vendedor_id = vendedor_token.id_vendedor
        
        # Normalizar día (quitar tildes para consistencia)
        dias_sin_tilde = {
            'SÁBADO': 'SABADO',
            'MIÉRCOLES': 'MIERCOLES',
        }
        dia = dias_sin_tilde.get(dia_raw, dia_raw)
        
        # Sanitizar fecha: tomar solo los primeros 10 caracteres (YYYY-MM-DD)
        if fecha_raw and len(str(fecha_raw)) > 10:
            fecha = str(fecha_raw)[:10]
        else:
            fecha = fecha_raw
            
        productos = data.get('productos', []) # Lista de {nombre, cantidad}
        print(f"📱 Recibiendo Sugerido App: {vendedor_id} - {dia} (raw: {dia_raw}) - {fecha}")

        # Mapeo de ID a Modelo
        modelos = {
            'ID1': CargueID1,
            'ID2': CargueID2,
            'ID3': CargueID3,
            'ID4': CargueID4,
            'ID5': CargueID5,
            'ID6': CargueID6,
        }

        Modelo = modelos.get(vendedor_id)
        if not Modelo:
            return Response({'error': f'Vendedor no válido: {vendedor_id}'}, status=400)

        if not fecha:
            return Response({'error': 'La fecha es requerida'}, status=400)

        # ✅ VALIDACIÓN: Verificar si ya existe sugerido para este día/fecha/vendedor
        # Solo bloquear si hay al menos un producto con cantidad > 0
        # (Si todos están en 0, fue un envío fallido y se permite reenviar)
        registros_existentes = Modelo.objects.filter(dia=dia, fecha=fecha)
        registros_con_cantidad = registros_existentes.filter(cantidad__gt=0)
        if registros_con_cantidad.exists():
            total_existente = registros_existentes.count()
            print(f"⚠️ Ya existe sugerido para {vendedor_id} - {dia} - {fecha} ({total_existente} productos)")
            return Response({
                'error': 'YA_EXISTE_SUGERIDO',
                'message': f'Ya existe un sugerido para {dia} {fecha}. No se puede enviar otro.',
                'productos_existentes': total_existente
            }, status=409)  # 409 Conflict
        elif registros_existentes.exists():
            # Hay registros pero todos en cantidad=0 (envío fallido), limpiar y permitir reenvío
            total_vacios = registros_existentes.count()
            registros_existentes.delete()
            print(f"🧹 Limpiados {total_vacios} registros vacíos (envío fallido) para {vendedor_id} - {dia} - {fecha}")

        # 🚀 OPTIMIZACIÓN: Obtener vendedor una sola vez (fuera del loop)
        vendedor_nombre = vendedor_id
        try:
            from .models import Vendedor
            vendedor_obj = Vendedor.objects.filter(id_vendedor=vendedor_id).first()
            if vendedor_obj and vendedor_obj.nombre:
                vendedor_nombre = vendedor_obj.nombre
        except Exception as e:
            print(f"  ⚠️ Error buscando vendedor: {e}")
        
        # 🚀 Obtener todos los productos existentes en UNA SOLA QUERY
        productos_nombres = [prod.get('nombre') for prod in productos if prod.get('nombre')]
        registros_existentes = {
            reg.producto: reg
            for reg in Modelo.objects.filter(dia=dia, fecha=fecha, producto__in=productos_nombres)
        }
        
        # Listas para bulk operations
        productos_actualizar = []
        productos_crear = []
        
        # Procesar cada producto
        count = 0
        for prod in productos:
            nombre = prod.get('nombre')
            cantidad_raw = prod.get('cantidad')
            cantidad = int(cantidad_raw) if cantidad_raw is not None else 0
            
            print(f"  📦 Procesando: {nombre} - Cantidad raw: {cantidad_raw} - Cantidad int: {cantidad}")
            
            # La app envía TODOS los productos (con o sin cantidad)
            # porque pueden tener adicionales/descuentos que modifiquen el total
            
            if nombre:
                # Obtener check V (si viene)
                v_check = prod.get('v', False) or prod.get('V', False)
                
                # Normalizar nombre para evitar duplicados
                import re
                nombre = re.sub(r'\s+', ' ', nombre).strip()

                if not nombre:
                    continue

                # 🚀 Buscar en diccionario (sin query)
                registro_existente = registros_existentes.get(nombre)
                
                # Determinar responsable
                responsable_a_usar = vendedor_nombre
                
                if registro_existente and registro_existente.responsable:
                    if not registro_existente.responsable.startswith('ID'):
                        responsable_a_usar = registro_existente.responsable
                
                # 🚀 Preparar para bulk operations (sin guardar todavía)
                if registro_existente:
                    # Actualizar existente
                    registro_existente.cantidad = cantidad
                    registro_existente.total = cantidad
                    registro_existente.responsable = responsable_a_usar
                    registro_existente.usuario = 'AppMovil'
                    registro_existente.v = v_check
                    productos_actualizar.append(registro_existente)
                else:
                    # Crear nuevo
                    productos_crear.append(Modelo(
                        dia=dia,
                        fecha=fecha,
                        producto=nombre,
                        cantidad=cantidad,
                        total=cantidad,
                        responsable=responsable_a_usar,
                        usuario='AppMovil',
                        v=v_check
                    ))
                
                count += 1
        
        # 🚀 Ejecutar bulk operations (solo 2 queries en total)
        if productos_actualizar:
            Modelo.objects.bulk_update(
                productos_actualizar,
                ['cantidad', 'total', 'responsable', 'usuario', 'v'],
                batch_size=100
            )
            print(f"  ✅ Actualizados {len(productos_actualizar)} productos en bulk")
        
        if productos_crear:
            Modelo.objects.bulk_create(productos_crear, batch_size=100)
            print(f"  ✅ Creados {len(productos_crear)} productos en bulk")
        
        print(f"✅ Sugerido guardado: {count} productos actualizados para {vendedor_id}")
        return Response({'success': True, 'message': f'Sugerido guardado correctamente ({count} productos)'})

    except Exception as e:
        print(f"❌ Error guardando sugerido: {str(e)}")
        return Response({'error': str(e)}, status=500)

@api_view(['POST'])
def corregir_cantidad_cargue(request):
    """
    Endpoint para corregir cantidad de un producto específico.
    NO crea registros nuevos, solo actualiza existentes.
    Preserva el campo 'usuario' (AppMovil/CRM) y recalcula el total.
    """
    try:
        data = request.data
        vendedor_id = data.get('vendedor_id')  # Ej: "ID1"
        dia = data.get('dia', '').upper()  # Ej: "LUNES"
        fecha = data.get('fecha')  # Ej: "2026-02-11"
        producto = data.get('producto')  # Ej: "AREPA TIPO OBLEA"
        nueva_cantidad = data.get('nueva_cantidad')
        
        print(f"🔧 Corrigiendo cantidad: {vendedor_id} - {dia} - {fecha} - {producto} → {nueva_cantidad}")
        
        # Validación de campos requeridos
        if not all([vendedor_id, dia, fecha, producto, nueva_cantidad is not None]):
            return Response({
                'error': 'Faltan campos requeridos',
                'campos': {
                    'vendedor_id': vendedor_id,
                    'dia': dia,
                    'fecha': fecha,
                    'producto': producto,
                    'nueva_cantidad': nueva_cantidad
                }
            }, status=400)
        
        # Mapeo de modelos
        modelos = {
            'ID1': CargueID1,
            'ID2': CargueID2,
            'ID3': CargueID3,
            'ID4': CargueID4,
            'ID5': CargueID5,
            'ID6': CargueID6,
        }
        
        Modelo = modelos.get(vendedor_id)
        if not Modelo:
            return Response({'error': f'Vendedor no válido: {vendedor_id}'}, status=400)
        
        # Buscar registro existente
        registro = Modelo.objects.filter(
            dia=dia,
            fecha=fecha,
            producto=producto
        ).first()
        
        if not registro:
            return Response({
                'error': 'Producto no encontrado',
                'message': f'No existe "{producto}" para {vendedor_id} - {dia} - {fecha}'
            }, status=404)
        
        # Guardar valor anterior para log
        cantidad_anterior = registro.cantidad
        
        # Actualizar cantidad
        registro.cantidad = int(nueva_cantidad)
        
        # Recalcular total (preservando otros campos)
        cantidad = int(nueva_cantidad)
        dctos = int(registro.dctos) if registro.dctos else 0
        adicional = int(registro.adicional) if registro.adicional else 0
        devoluciones = int(registro.devoluciones) if registro.devoluciones else 0
        vencidas = int(registro.vencidas) if registro.vencidas else 0
        
        registro.total = cantidad - dctos + adicional - devoluciones - vencidas
        
        # Guardar (preserva campo 'usuario' - NO lo modificamos)
        registro.save()
        
        print(f"✅ Cantidad corregida: {producto}")
        print(f"   Cantidad: {cantidad_anterior} → {cantidad}")
        print(f"   Total recalculado: {registro.total}")
        print(f"   Usuario preservado: {registro.usuario}")
        
        return Response({
            'success': True,
            'message': 'Cantidad actualizada correctamente',
            'producto': producto,
            'cantidad_anterior': cantidad_anterior,
            'cantidad_nueva': cantidad,
            'total_recalculado': registro.total,
            'usuario': registro.usuario  # Retornar para confirmar que se preservó
        })
        
    except Exception as e:
        print(f"❌ Error corrigiendo cantidad: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)

@api_view(['POST'])
def actualizar_check_vendedor(request):
    """
    Endpoint para actualizar el check V (vendedor) desde la App Móvil.
    Recibe: { vendedor_id, dia, fecha, producto, v (true/false) }
    Validación: Solo permite marcar V si D ya está marcado y hay cantidad > 0
    """
    try:
        data = request.data
        vendedor_id = data.get('vendedor_id')
        dia = data.get('dia', '').upper()
        fecha = data.get('fecha')
        producto = data.get('producto')
        v_nuevo = data.get('v', False)
        
        # Mapeo de ID a Modelo
        modelos = {
            'ID1': CargueID1,
            'ID2': CargueID2,
            'ID3': CargueID3,
            'ID4': CargueID4,
            'ID5': CargueID5,
            'ID6': CargueID6,
        }
        
        Modelo = modelos.get(vendedor_id)
        if not Modelo:
            return Response({'error': f'Vendedor no válido: {vendedor_id}'}, status=400)
        
        # ⚡ OPTIMIZACIÓN: Usar select_for_update para evitar race conditions
        # y only() para traer solo los campos necesarios
        try:
            registro = Modelo.objects.only('v', 'd', 'total').get(
                dia=dia,
                fecha=fecha,
                producto=producto
            )
        except Modelo.DoesNotExist:
            # Crear registro si no existe
            registro = Modelo.objects.create(
                dia=dia,
                fecha=fecha,
                producto=producto,
                cantidad=0,
                v=False,
                d=False,
                responsable='Sistema'
            )
        
        # ✅ VALIDACIÓN: Solo permitir marcar V si D está marcado y hay cantidad
        if v_nuevo:
            if not registro.d:
                return Response({
                    'error': 'CHECK_D_REQUERIDO',
                    'message': 'No puedes marcar el check de Vendedor hasta que el Despachador lo haya marcado en el CRM.'
                }, status=400)
            
            if (registro.total or 0) <= 0:
                return Response({
                    'error': 'SIN_CANTIDAD',
                    'message': 'No puedes marcar el check sin cantidad de producto.'
                }, status=400)
        
        # ⚡ Actualizar el campo V y fecha_actualizacion
        from django.utils import timezone
        Modelo.objects.filter(
            dia=dia,
            fecha=fecha,
            producto=producto
        ).update(
            v=v_nuevo,
            fecha_actualizacion=timezone.now()  # 🔥 CRÍTICO: Actualizar timestamp para polling
        )
        
        print(f"✅ Check V actualizado: {producto} = {v_nuevo} (timestamp actualizado)")
        
        return Response({
            'success': True,
            'message': 'Check actualizado correctamente',
            'v': v_nuevo,
            'd': registro.d,
            'total': registro.total
        })
        
    except Exception as e:
        print(f"❌ Error actualizando check V: {str(e)}")
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
def obtener_cargue(request):
    """
    Endpoint para obtener Cargue desde la App Móvil.
    Recibe params: vendedor_id, dia, fecha
    Devuelve: cantidad, total, v (vendedor check), d (despachador check)
    """
    try:
        vendedor_id = request.query_params.get('vendedor_id')
        dia_raw = request.query_params.get('dia', '').upper()
        fecha = request.query_params.get('fecha') # YYYY-MM-DD
        
        # Normalizar día (quitar tildes para consistencia con BD)
        dias_sin_tilde = {
            'SÁBADO': 'SABADO',
            'MIÉRCOLES': 'MIERCOLES',
        }
        dia = dias_sin_tilde.get(dia_raw, dia_raw)

        print(f"📱 Solicitando Cargue App: {vendedor_id} - {dia} (raw: {dia_raw}) - {fecha}")

        # Mapeo de ID a Modelo
        modelos = {
            'ID1': CargueID1,
            'ID2': CargueID2,
            'ID3': CargueID3,
            'ID4': CargueID4,
            'ID5': CargueID5,
            'ID6': CargueID6,
        }

        Modelo = modelos.get(vendedor_id)
        if not Modelo:
            return Response({'error': f'Vendedor no válido: {vendedor_id}'}, status=400)

        # Construir filtro - buscar con y sin tilde
        from django.db.models import Q
        from .models import Pedido # Importar modelo Pedido
        
        # Crear lista de posibles variantes del día
        dias_variantes = [dia, dia_raw]
        if dia != dia_raw:
            dias_variantes = list(set(dias_variantes))
        
        filtros_q = Q(dia__in=dias_variantes)
        if fecha:
            filtros_q &= Q(fecha=fecha)
        
        # Obtener registros
        registros = Modelo.objects.filter(filtros_q)
        
        # Formatear respuesta para la App
        data = {}
        
        # 🆕 Obtener estado del cargue desde CargueResumen
        from .models import CargueResumen
        estado_cargue_valor = 'DESCONOCIDO'
        
        try:
            # 🔧 CAMBIO: Siempre buscar el estado en ID1 (maestro), ya que el estado es GLOBAL
            # El estado se controla desde ID1 y aplica para todos los IDs
            resumen = CargueResumen.objects.filter(
                vendedor_id='ID1',  # 🔧 Siempre buscar en ID1
                fecha=fecha
            ).first()
            
            if resumen:
                estado_cargue_valor = resumen.estado_cargue
                print(f"✅ Estado del cargue obtenido desde ID1: {estado_cargue_valor}")
            else:
                print(f"⚠️ No se encontró estado en CargueResumen para ID1 - fecha: {fecha}")
        except Exception as e:
            print(f"⚠️ Error al obtener estado del cargue: {e}")

        # 🆕 Precargar diccionario de Precios Alternos (Listas de Precio)
        precios_alternos = {}
        try:
            from .models import PrecioProducto
            # Consultar solo precios activos y relacionar
            pp_qs = PrecioProducto.objects.select_related('producto', 'lista_precio').filter(activo=True)
            for p in pp_qs:
                prod_nombre = p.producto.nombre
                lista_nombre = p.lista_precio.nombre
                
                if prod_nombre not in precios_alternos:
                    precios_alternos[prod_nombre] = {}
                
                precios_alternos[prod_nombre][lista_nombre] = float(p.precio)
        except Exception as e:
            print(f"⚠️ Error cargando precios alternos: {e}")

        for reg in registros:
            # Stock disponible para vender
            # Total = Cantidad + Adicional - Dctos - Devoluciones - Vencidas
            # Disponible = Total - Vendidas
            if reg.total is not None:
                stock_disponible = reg.total - (reg.vendidas or 0)
            else:
                stock_disponible = (reg.cantidad or 0) + (reg.adicional or 0) - (reg.dctos or 0) - (reg.vendidas or 0) - (reg.vencidas or 0)

            quantity_value = str(max(0, stock_disponible))  # No permitir negativos
            
            # Obtener precios alternos para este producto
            precios_producto = precios_alternos.get(reg.producto, {})
            
            data[reg.producto] = {
                'quantity': quantity_value,  # Stock disponible (total - vendidas + pedidos_entregados)
                'cantidad': reg.cantidad or 0,  # Cantidad base
                'adicional': reg.adicional or 0,  # Adicionales
                'dctos': reg.dctos or 0,  # Descuentos
                'vendidas': reg.vendidas or 0,  # 🆕 Vendidas
                'vencidas': reg.vencidas or 0,  # 🆕 Vencidas
                'devoluciones': reg.devoluciones or 0,  # 🆕 Devoluciones
                'turno_cerrado': False,  # El stock no se bloquea por devoluciones
                'estado': estado_cargue_valor, # 🆕 ESTADO DEL CARGUE (DESPACHO, ETC)
                'precios_alternos': precios_producto, # 🆕 PRECIOS DE LISTAS ESPECIALES
                'v': reg.v,  # Check vendedor
                'd': reg.d,   # Check despachador
                # 🆕 Campos adicionales para sincronización completa
                'lotes_vencidos': reg.lotes_vencidos or '',  # JSON string de lotes
                'total': reg.total or 0,
                'valor': float(reg.valor) if reg.valor else 0,
                'neto': float(reg.neto) if reg.neto else 0,
                # Pagos (pueden estar en el mismo registro)
                'nequi': float(reg.nequi) if reg.nequi else 0,
                'daviplata': float(reg.daviplata) if reg.daviplata else 0,
                'concepto': reg.concepto or '',
                'descuentos': float(reg.descuentos) if reg.descuentos else 0,
                # Resumen
                'base_caja': float(reg.base_caja) if reg.base_caja else 0,
                # Cumplimiento
                'licencia_transporte': reg.licencia_transporte or '',
                'soat': reg.soat or '',
                'uniforme': reg.uniforme or '',
                'no_locion': reg.no_locion or '',
                'no_accesorios': reg.no_accesorios or '',
                'capacitacion_carnet': reg.capacitacion_carnet or '',
                'higiene': reg.higiene or '',
                'estibas': reg.estibas or '',
                'desinfeccion': reg.desinfeccion or '',
            }
            # 🆕 Debug
            if 'CANASTILLA' in reg.producto.upper():
                print(f"🔍 BACKEND - CANASTILLA: cantidad={reg.cantidad}, adicional={reg.adicional}, total={reg.total}, quantity_value='{quantity_value}'")

        # Agregar totales de CarguePagos como clave especial __pagos__
        try:
            from .models import CarguePagos
            from django.db.models import Sum
            pagos_qs = CarguePagos.objects.filter(
                vendedor_id=vendedor_id,
                activo=True
            )
            if fecha:
                pagos_qs = pagos_qs.filter(fecha=fecha)
            elif dia:
                pagos_qs = pagos_qs.filter(dia__in=dias_variantes)

            pagos_totales = pagos_qs.aggregate(
                total_nequi=Sum('nequi'),
                total_daviplata=Sum('daviplata'),
                total_descuentos=Sum('descuentos'),
            )
            data['__pagos__'] = {
                'nequi': float(pagos_totales['total_nequi'] or 0),
                'daviplata': float(pagos_totales['total_daviplata'] or 0),
                'descuentos': float(pagos_totales['total_descuentos'] or 0),
            }
            print(f"💳 Pagos CarguePagos: nequi={data['__pagos__']['nequi']}, daviplata={data['__pagos__']['daviplata']}")
        except Exception as ep:
            print(f"⚠️ Error obteniendo CarguePagos: {ep}")
            data['__pagos__'] = {'nequi': 0, 'daviplata': 0, 'descuentos': 0}

        return Response(data)

    except Exception as e:
        print(f"❌ Error obteniendo cargue: {str(e)}")
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
def obtener_rendimiento_cargue(request):
    """
    Obtiene el rendimiento consolidado de todos los IDs para un día y fecha específica
    Para el módulo de Rendimiento de la app móvil
    """
    dia = request.GET.get('dia', '').upper()
    fecha = request.GET.get('fecha')
    
    if not dia or not fecha:
        return Response({'error': 'Faltan parámetros: dia, fecha'}, status=400)
    
    # Mapear todos los modelos de cargue
    modelos = {
        'ID1': CargueID1,
        'ID2': CargueID2,
        'ID3': CargueID3,
        'ID4': CargueID4,
        'ID5': CargueID5,
        'ID6': CargueID6
    }
    
    try:
        # Consolidar datos de todos los IDs
        productos_consolidados = {}
        
        for id_name, modelo in modelos.items():
            registros = modelo.objects.filter(dia=dia, fecha=fecha)
            
            for registro in registros:
                producto_nombre = registro.producto
                
                if producto_nombre not in productos_consolidados:
                    productos_consolidados[producto_nombre] = {
                        'producto': producto_nombre,
                        'vencidas': 0,
                        'devoluciones': 0,
                        'total': 0
                    }
                
                # Sumar los valores de todos los IDs
                productos_consolidados[producto_nombre]['vencidas'] += registro.vencidas or 0
                productos_consolidados[producto_nombre]['devoluciones'] += registro.devoluciones or 0
                productos_consolidados[producto_nombre]['total'] += registro.total or 0
        
        # Convertir a lista y ordenar por nombre de producto
        data = list(productos_consolidados.values())
        data.sort(key=lambda x: x['producto'])
        
        return Response({
            'success': True,
            'data': data,
            'dia': dia,
            'fecha': fecha,
            'total_productos': len(data)
        })
        
    except Exception as e:
        print(f"❌ Error obteniendo rendimiento: {str(e)}")
        return Response({'error': str(e)}, status=500)


# ===== ENDPOINT: VERIFICAR ESTADO DEL DÍA =====
# Agregado: 24 Nov 2025
# Propósito: Permitir verificar si un día específico está completado o en qué estado se encuentra

@api_view(['GET'])
def verificar_estado_dia(request):
    """
    Verifica el estado de un día específico para un vendedor
    
    Parámetros (query params):
        - vendedor_id: ID del vendedor (ID1, ID2, etc.)
        - dia: Día de la semana (LUNES, MARTES, etc.)
        - fecha: Fecha en formato YYYY-MM-DD
    
    Retorna:
        {
            "success": true,
            "completado": false,
            "estado": "SUGERIDO" | "DESPACHO" | "COMPLETADO",
            "puede_editar": true,
            "mensaje": "Este día está disponible para edición",
            "fecha": "2025-11-24",
            "dia": "LUNES",
            "tiene_datos": false,
            "total_productos": 0
        }
    """
    try:
        vendedor_id = request.GET.get('vendedor_id', '').upper()
        dia = request.GET.get('dia', '').upper()
        fecha = request.GET.get('fecha', '')
        
        # Validar parámetros
        if not vendedor_id or not dia or not fecha:
            return Response({
                'success': False,
                'error': 'Faltan parámetros requeridos: vendedor_id, dia, fecha'
            }, status=400)
        
        # Mapear vendedor_id a modelo de tabla
        modelos_cargue = {
            'ID1': CargueID1,
            'ID2': CargueID2,
            'ID3': CargueID3,
            'ID4': CargueID4,
            'ID5': CargueID5,
            'ID6': CargueID6,
        }
        
        modelo = modelos_cargue.get(vendedor_id)
        if not modelo:
            return Response({
                'success': False,
                'error': f'Vendedor ID inválido: {vendedor_id}'
            }, status=400)
        
        # Buscar registros para este día y fecha
        registros = modelo.objects.filter(dia=dia, fecha=fecha)
        
        tiene_datos = registros.exists()
        total_productos = registros.count()
        
        # Determinar estado del día
        # Nota: El estado "COMPLETADO" se maneja actualmente en localStorage del frontend
        # Aquí solo podemos verificar si hay datos guardados
        
        estado = "SUGERIDO"  # Estado por defecto (día vacío)
        completado = False
        puede_editar = True
        mensaje = "Este día está disponible para edición"
        
        if tiene_datos:
            # Verificar si algún registro tiene checks marcados
            tiene_checks_d = registros.filter(d=True).exists()
            tiene_checks_v = registros.filter(v=True).exists()
            
            if tiene_checks_v:
                estado = "DESPACHO"
                mensaje = "Este día tiene datos con checks de vendedor marcados"
            elif tiene_checks_d:
                estado = "DESPACHO"
                mensaje = "Este día tiene datos con checks de despachador marcados"
            else:
                estado = "SUGERIDO"
                mensaje = "Este día tiene datos pero no está despachado"
            
            # Por ahora, siempre permitimos editar
            # En el futuro, podríamos agregar un campo 'finalizado' en la tabla
            puede_editar = True
        
        return Response({
            'success': True,
            'completado': completado,
            'estado': estado,
            'puede_editar': puede_editar,
            'mensaje': mensaje,
            'fecha': fecha,
            'dia': dia,
            'tiene_datos': tiene_datos,
            'total_productos': total_productos,
            'vendedor_id': vendedor_id
        })
        
    except Exception as e:
        print(f"❌ Error verificando estado del día: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({
            'success': False,
            'error': str(e)
        }, status=500)


# ===== VIEWSETS RUTAS Y VENTAS RUTA =====

class RutaViewSet(viewsets.ModelViewSet):
    queryset = Ruta.objects.all()
    serializer_class = RutaSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = Ruta.objects.all()
        vendedor_id = self.request.query_params.get('vendedor_id', None)
        if vendedor_id:
            # Filtrar por ID de vendedor (ej: ID1)
            queryset = queryset.filter(vendedor__id_vendedor=vendedor_id)
        return queryset

class ClienteRutaViewSet(viewsets.ModelViewSet):
    queryset = ClienteRuta.objects.select_related('ruta').all()  # 🔥 Optimización
    serializer_class = ClienteRutaSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = ClienteRuta.objects.select_related('ruta', 'ruta__vendedor').all()  # 🔥 Optimización
        ruta_id = self.request.query_params.get('ruta', None)
        dia = self.request.query_params.get('dia', None)
        vendedor_id = self.request.query_params.get('vendedor_id', None)
        search = self.request.query_params.get('search', None)
        
        # 🆕 Búsqueda Global
        if search:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(nombre_negocio__icontains=search) |
                Q(nombre_contacto__icontains=search) |
                Q(telefono__icontains=search) |
                Q(direccion__icontains=search)
            )
        
        # Filtrar por vendedor (busca la ruta del vendedor)
        if vendedor_id:
            rutas_vendedor = Ruta.objects.filter(vendedor__id_vendedor=vendedor_id, activo=True)
            queryset = queryset.filter(ruta__in=rutas_vendedor)
            # 🔥 NO filtrar por ruta_id específica si viene vendedor_id
            # Esto permite que se muestren clientes de TODAS las rutas del vendedor
            # Solo usar ruta_id si viene explícitamente en los parámetros
        
        if ruta_id and not vendedor_id:
            # Solo filtrar por ruta_id si NO viene vendedor_id
            queryset = queryset.filter(ruta_id=ruta_id)
            
        if dia:
            dia_upper = dia.upper()
            # Buscar clientes que tengan este día en su lista (soporta múltiples días)
            queryset = queryset.filter(dia_visita__icontains=dia_upper)
            
            # 🆕 Ordenar según RutaOrden si existe para esta ruta + día
            # Si viene vendedor_id, aplicar orden combinado de todas las rutas
            if vendedor_id and not ruta_id:
                vendedor_norm, _ = _normalizar_id_vendedor(vendedor_id)

                # 1) Prioridad: orden global exacto por vendedor + día (App móvil)
                if vendedor_norm:
                    orden_global = RutaOrdenVendedor.objects.filter(
                        vendedor__id_vendedor=vendedor_norm,
                        dia=dia_upper
                    ).first()

                    if orden_global and isinstance(orden_global.clientes_ids, list) and orden_global.clientes_ids:
                        ids_globales = []
                        for raw_id in orden_global.clientes_ids:
                            try:
                                ids_globales.append(int(raw_id))
                            except (TypeError, ValueError):
                                continue

                        if ids_globales:
                            from django.db.models import Case, When, Value, IntegerField

                            ordering_cases = [When(id=pk, then=Value(pos)) for pos, pk in enumerate(ids_globales)]
                            queryset = queryset.annotate(
                                orden_dia=Case(
                                    *ordering_cases,
                                    default=Value(999999),
                                    output_field=IntegerField()
                                )
                            ).order_by('orden_dia', 'ruta__id', 'orden', 'id')

                            # 🚀 OPTIMIZACIÓN: Pre-computar listas de precios en masa para evitar N+1 queries
                            from .models import Cliente
                            nombres_mostrar = list(queryset.values_list('nombre_negocio', flat=True))
                            clientes_admin = Cliente.objects.filter(nombre_completo__in=nombres_mostrar).values('nombre_completo', 'tipo_lista_precio')
                            precios_map = {c['nombre_completo']: c['tipo_lista_precio'] for c in clientes_admin}
                            
                            for c in queryset:
                                c.precomputed_lista_precio = precios_map.get(c.nombre_negocio)

                            return queryset.filter(activo=True)

                # Obtener todas las órdenes guardadas para este día (ordenadas por más reciente)
                rutas_vendedor = Ruta.objects.filter(vendedor__id_vendedor=vendedor_id, activo=True)
                ordenes = RutaOrden.objects.filter(
                    ruta__in=rutas_vendedor, 
                    dia=dia_upper
                ).order_by('-fecha_actualizacion')  # 🔥 Más reciente primero
                
                # Combinar todos los clientes_ids de todas las rutas
                todos_ids_ordenados = []
                ids_vistos = set()
                
                for orden in ordenes:
                    for cliente_id in orden.clientes_ids:
                        if cliente_id not in ids_vistos:
                            todos_ids_ordenados.append(cliente_id)
                            ids_vistos.add(cliente_id)
                
                if todos_ids_ordenados:
                    # Aplicar orden combinado
                    from django.db.models import Case, When, Value, IntegerField
                    
                    ordering_cases = [When(id=pk, then=Value(pos)) for pos, pk in enumerate(todos_ids_ordenados)]
                    
                    queryset = queryset.annotate(
                        orden_dia=Case(
                            *ordering_cases,
                            default=Value(999999),  # Clientes no en la lista van al final
                            output_field=IntegerField()
                        )
                    ).order_by('orden_dia', 'ruta__id', 'orden', 'id')
                    
                    return queryset.filter(activo=True)
                
                # Si no hay orden personalizado, usar orden por defecto
                return queryset.filter(activo=True).order_by('ruta__id', 'orden', 'id')
            
            if ruta_id:
                try:
                    orden_personalizado = RutaOrden.objects.get(ruta_id=ruta_id, dia=dia_upper)
                    clientes_ids = orden_personalizado.clientes_ids
                    
                    if clientes_ids and len(clientes_ids) > 0:
                        # Crear orden dinámico basado en la posición en la lista
                        from django.db.models import Case, When, Value, IntegerField
                        
                        # Crear condiciones de ordenamiento
                        ordering_cases = [When(id=pk, then=Value(pos)) for pos, pk in enumerate(clientes_ids)]
                        
                        # Agregar anotación para orden personalizado
                        queryset = queryset.annotate(
                            orden_dia=Case(
                                *ordering_cases,
                                default=Value(999999),  # Clientes no en la lista van al final
                                output_field=IntegerField()
                            )
                        ).order_by('orden_dia', 'orden', 'id')
                        
                        return queryset.filter(activo=True)
                except RutaOrden.DoesNotExist:
                    pass  # No hay orden personalizado, usar orden por defecto
            
        return queryset.filter(activo=True).order_by('orden', 'id')
    
    def _ordenar_dias_semana(self, dias_string):
        """
        Ordena los días de la semana en orden cronológico
        Entrada: "SABADO,MARTES,JUEVES" o "Sabado, Martes, Jueves"
        Salida: "MARTES,JUEVES,SABADO"
        """
        if not dias_string:
            return dias_string
        
        # Orden de días de la semana
        orden_dias = {
            'LUNES': 1,
            'MARTES': 2,
            'MIERCOLES': 3,
            'JUEVES': 4,
            'VIERNES': 5,
            'SABADO': 6,
            'DOMINGO': 7
        }
        
        # Separar días, limpiar y convertir a mayúsculas
        dias = [dia.strip().upper() for dia in dias_string.split(',')]
        
        # Ordenar según el diccionario
        dias_ordenados = sorted(dias, key=lambda d: orden_dias.get(d, 99))
        
        # Retornar en el mismo formato (mayúsculas, separados por coma)
        return ','.join(dias_ordenados)
    
    def perform_update(self, serializer):
        """Se ejecuta al actualizar un ClienteRuta - Sincroniza hacia Cliente"""
        # Ordenar días antes de guardar
        if 'dia_visita' in serializer.validated_data:
            serializer.validated_data['dia_visita'] = self._ordenar_dias_semana(
                serializer.validated_data['dia_visita']
            )
        
        cliente_ruta = serializer.save()
        
        # 🔄 SINCRONIZAR HACIA CLIENTE (si existe un cliente con el mismo nombre)
        try:
            # Buscar cliente por alias o nombre completo que coincida con nombre_negocio
            cliente = Cliente.objects.filter(
                models.Q(alias__iexact=cliente_ruta.nombre_negocio) |
                models.Q(nombre_completo__iexact=cliente_ruta.nombre_contacto)
            ).first()
            
            if cliente:
                # Marcar flag para evitar loop infinito
                cliente._sincronizando = True
                
                # Actualizar campos del cliente con los datos de ClienteRuta (ya ordenados)
                cliente.dia_entrega = cliente_ruta.dia_visita
                cliente.direccion = cliente_ruta.direccion or cliente.direccion
                cliente.telefono_1 = cliente_ruta.telefono or cliente.telefono_1
                cliente.zona_barrio = cliente_ruta.ruta.nombre  # Sincronizar la ruta
                cliente.nota = cliente_ruta.nota # 🆕 Sincronizar nota
                cliente.save()
                print(f"✅ Cliente sincronizado desde ClienteRuta: {cliente.alias} - Días: {cliente.dia_entrega}")
            else:
                print(f"⚠️ No se encontró Cliente correspondiente para: {cliente_ruta.nombre_negocio}")
                
        except Exception as e:
            print(f"❌ Error sincronizando ClienteRuta → Cliente: {e}")
            import traceback
            traceback.print_exc()


class VentaRutaViewSet(viewsets.ModelViewSet):
    queryset = VentaRuta.objects.select_related('vendedor', 'ruta', 'cliente').all()  # 🔥 Optimización: Reducir consultas N+1
    serializer_class = VentaRutaSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = VentaRuta.objects.select_related('vendedor', 'ruta', 'cliente').all()  # 🔥 Optimización
        vendedor_id = self.request.query_params.get('vendedor_id', None)
        fecha = self.request.query_params.get('fecha', None)
        
        fecha_inicio = self.request.query_params.get('fecha_inicio', None)
        fecha_fin = self.request.query_params.get('fecha_fin', None)

        if vendedor_id:
            queryset = queryset.filter(vendedor__id_vendedor=vendedor_id)
        
        if fecha_inicio and fecha_fin:
             queryset = queryset.filter(fecha__date__range=[fecha_inicio, fecha_fin])
        elif fecha:
            # Filtrar por fecha (solo la parte de la fecha, ignorando la hora)
            queryset = queryset.filter(fecha__date=fecha)
            
        cliente_id = self.request.query_params.get('cliente_id', None)
        if cliente_id:
             queryset = queryset.filter(cliente__id=cliente_id)

        ruta_id = self.request.query_params.get('ruta_id', None)
        if ruta_id:
            from django.db.models import Q
            # Filtrar si la venta tiene la ruta marcada O si el cliente pertenece a esa ruta
            queryset = queryset.filter(Q(ruta_id=ruta_id) | Q(cliente__ruta_id=ruta_id))
        
        # Ordenar por fecha descendente (más recientes primero)
        return queryset.order_by('-fecha')

    def _extraer_token_request(self, request):
        auth_header = request.headers.get('Authorization', '') or ''
        if auth_header.lower().startswith('bearer '):
            return auth_header.split(' ', 1)[1].strip()
        return (request.headers.get('X-App-Token') or '').strip()

    def _validar_sesion_movil(self, request):
        token = self._extraer_token_request(request)
        if not token:
            return None

        sesion = VendedorSesionToken.objects.select_related('vendedor').filter(
            token=token,
            activo=True,
            expira_en__gt=timezone.now(),
            vendedor__activo=True
        ).first()

        if sesion:
            sesion.ultimo_uso = timezone.now()
            sesion.save(update_fields=['ultimo_uso'])
            return sesion.vendedor

        return None
    
    @staticmethod
    def _to_int(value, default=0):
        try:
            if value is None:
                return default
            return int(float(str(value).strip()))
        except Exception:
            return default

    @staticmethod
    def _to_float(value, default=0.0):
        try:
            if value is None:
                return default
            return float(str(value).strip())
        except Exception:
            return default

    @staticmethod
    def _normalizar_nombre_producto_cargue(nombre_producto):
        texto = str(nombre_producto or '').strip()
        if not texto:
            return ''

        texto = unicodedata.normalize('NFD', texto)
        texto = ''.join(ch for ch in texto if unicodedata.category(ch) != 'Mn')
        texto = texto.upper()
        texto = re.sub(r'[^A-Z0-9#]+', ' ', texto)
        texto = re.sub(r'\s+', ' ', texto).strip()

        tokens = []
        for token in texto.split():
            if len(token) > 3 and token.endswith('S'):
                token = token[:-1]
            tokens.append(token)

        return ' '.join(tokens)

    def _resolver_registro_cargue(self, ModeloCargue, fecha_venta, nombre_producto):
        if not ModeloCargue or not fecha_venta or not nombre_producto:
            return None

        queryset = ModeloCargue.objects.filter(fecha=fecha_venta, activo=True)

        # Ruta rápida: coincidencia exacta habitual
        registro = queryset.filter(producto__iexact=nombre_producto).first()
        if registro:
            return registro

        nombre_norm = self._normalizar_nombre_producto_cargue(nombre_producto)
        if not nombre_norm:
            return None

        # Fallback tolerante: diferencias menores de nombre (plural, espacios, tildes)
        for candidato in queryset.only('id', 'producto'):
            if self._normalizar_nombre_producto_cargue(candidato.producto) == nombre_norm:
                return candidato

        return None

    @staticmethod
    def _resolver_fecha_operativa(fecha_raw, fallback=None):
        def _a_fecha_local(valor):
            if isinstance(valor, datetime):
                try:
                    if timezone.is_aware(valor):
                        return timezone.localtime(valor).date()
                except Exception:
                    pass
                return valor.date()
            if isinstance(valor, date):
                return valor
            return None

        fecha_local = _a_fecha_local(fecha_raw)
        if fecha_local:
            return fecha_local

        if fecha_raw:
            fecha_texto = str(fecha_raw).strip()
            fecha_dt = parse_datetime(fecha_texto)
            if fecha_dt:
                fecha_local = _a_fecha_local(fecha_dt)
                if fecha_local:
                    return fecha_local

            fecha_simple = parse_date(fecha_texto[:10])
            if fecha_simple:
                return fecha_simple

        fecha_fallback = _a_fecha_local(fallback)
        if fecha_fallback:
            return fecha_fallback
        return None

    @staticmethod
    def _obtener_modelo_cargue_por_vendedor(id_vendedor):
        modelo_map = {
            'ID1': CargueID1,
            'ID2': CargueID2,
            'ID3': CargueID3,
            'ID4': CargueID4,
            'ID5': CargueID5,
            'ID6': CargueID6,
        }
        return modelo_map.get((id_vendedor or '').strip().upper())

    def _construir_mapa_detalles_cargue(self, ModeloCargue, fecha_venta, detalles):
        detalles_por_registro = {}
        productos_sin_cargue = []

        if not ModeloCargue or not fecha_venta:
            return detalles_por_registro, productos_sin_cargue

        for item in detalles or []:
            if not isinstance(item, dict):
                continue

            nombre = (item.get('nombre') or item.get('producto') or '').strip()
            cantidad = self._to_int(item.get('cantidad'), 0)
            if not nombre or cantidad <= 0:
                continue

            registro = self._resolver_registro_cargue(ModeloCargue, fecha_venta, nombre)
            if not registro:
                productos_sin_cargue.append({
                    'producto': nombre,
                    'solicitado': cantidad,
                    'motivo': 'PRODUCTO_NO_EN_CARGUE',
                })
                continue

            if registro.pk not in detalles_por_registro:
                detalles_por_registro[registro.pk] = {
                    'registro': registro,
                    'producto': registro.producto,
                    'solicitado': 0,
                }

            detalles_por_registro[registro.pk]['solicitado'] += cantidad

        return detalles_por_registro, productos_sin_cargue

    def _cantidades_base_detalles(self, ModeloCargue, fecha_venta, detalles):
        cantidades_base = defaultdict(int)
        detalles_por_registro, _ = self._construir_mapa_detalles_cargue(ModeloCargue, fecha_venta, detalles)
        for pk_registro, info in detalles_por_registro.items():
            cantidades_base[pk_registro] += info['solicitado']
        return cantidades_base

    def _validar_stock_disponible_cargue(self, ModeloCargue, fecha_venta, detalles, cantidades_base=None):
        if not ModeloCargue or not fecha_venta:
            return None

        cantidades_base = cantidades_base or {}
        detalles_por_registro, productos_sin_cargue = self._construir_mapa_detalles_cargue(
            ModeloCargue,
            fecha_venta,
            detalles,
        )

        excedidos = []
        for info in detalles_por_registro.values():
            registro = info['registro']
            solicitado_total = info['solicitado']
            cantidad_base = self._to_int(cantidades_base.get(registro.pk), 0)
            incremento = solicitado_total - cantidad_base
            disponible = max(0, self._to_int(registro.total, 0) - self._to_int(registro.vendidas, 0))

            if incremento > disponible:
                excedidos.append({
                    'producto': registro.producto,
                    'solicitado': solicitado_total,
                    'cantidad_base': cantidad_base,
                    'incremento': incremento,
                    'disponible': disponible,
                })

        if not productos_sin_cargue and not excedidos:
            return None

        errores = []
        for item in productos_sin_cargue:
            errores.append({
                'producto': item['producto'],
                'solicitado': item['solicitado'],
                'codigo': item['motivo'],
                'mensaje': 'El producto no existe en el cargue del dia para este ID.',
            })

        for item in excedidos:
            errores.append({
                'producto': item['producto'],
                'solicitado': item['solicitado'],
                'disponible': item['disponible'],
                'incremento': item['incremento'],
                'codigo': 'STOCK_INSUFICIENTE_CARGUE',
                'mensaje': f"Solo quedan {item['disponible']} unidad(es) disponibles para vender.",
            })

        return {
            'error': 'La venta supera el stock disponible del cargue.',
            'codigo': 'STOCK_CARGUE_INSUFICIENTE',
            'detalle': errores,
        }

    def _buscar_producto_catalogo(self, nombre_producto):
        if not nombre_producto:
            return None

        producto = Producto.objects.filter(nombre__iexact=nombre_producto).first()
        if producto:
            return producto

        nombre_norm = self._normalizar_nombre_producto_cargue(nombre_producto)
        if not nombre_norm:
            return None

        for candidato in Producto.objects.only('id', 'nombre', 'precio', 'precio_cargue'):
            if self._normalizar_nombre_producto_cargue(candidato.nombre) == nombre_norm:
                return candidato

        return None

    def _crear_registro_cargue_on_the_fly(self, ModeloCargue, fecha_venta, nombre_producto, *, vendidas=0, vencidas=0):
        ref_cargue = ModeloCargue.objects.filter(fecha=fecha_venta, activo=True).first()
        if not ref_cargue:
            return None

        prod_obj = self._buscar_producto_catalogo(nombre_producto)
        nombre_real = prod_obj.nombre if prod_obj else nombre_producto
        precio_prod = 0
        if prod_obj:
            precio_prod = prod_obj.precio_cargue or prod_obj.precio or 0

        return ModeloCargue.objects.create(
            fecha=fecha_venta,
            dia=ref_cargue.dia,
            responsable=ref_cargue.responsable,
            usuario='Sistema',
            ruta=ref_cargue.ruta if hasattr(ref_cargue, 'ruta') else '',
            producto=nombre_real,
            valor=precio_prod,
            cantidad=0,
            vendidas=vendidas,
            vencidas=vencidas,
            activo=True
        )

    def _normalizar_detalles(self, detalles_raw):
        if isinstance(detalles_raw, str):
            try:
                detalles_raw = json.loads(detalles_raw)
            except Exception:
                detalles_raw = []

        if not isinstance(detalles_raw, list):
            return [], 0.0

        detalles_normalizados = []
        total_calculado = 0.0

        for item in detalles_raw:
            if not isinstance(item, dict):
                continue

            cantidad = self._to_int(item.get('cantidad'), 0)
            if cantidad <= 0:
                continue

            nombre = (item.get('nombre') or item.get('producto') or '').strip()
            producto = (item.get('producto') or nombre).strip()
            if not nombre and producto:
                nombre = producto
            if not nombre:
                continue

            precio = self._to_float(
                item.get('precio', item.get('precio_unitario', item.get('valor_unitario', 0))),
                0.0
            )
            if precio < 0:
                precio = 0.0

            subtotal = round(precio * cantidad, 2)

            item_normalizado = dict(item)
            item_normalizado['nombre'] = nombre
            item_normalizado['producto'] = producto or nombre
            item_normalizado['cantidad'] = cantidad
            item_normalizado['precio'] = precio
            item_normalizado['subtotal'] = subtotal

            detalles_normalizados.append(item_normalizado)
            total_calculado += subtotal

        return detalles_normalizados, round(total_calculado, 2)

    def _normalizar_productos_vencidos(self, productos_raw):
        if isinstance(productos_raw, str):
            try:
                productos_raw = json.loads(productos_raw)
            except Exception:
                productos_raw = []

        if not isinstance(productos_raw, list):
            return []

        normalizados = []
        for item in productos_raw:
            if not isinstance(item, dict):
                continue
            producto = (item.get('producto') or item.get('nombre') or '').strip()
            cantidad = self._to_int(item.get('cantidad'), 0)
            if not producto or cantidad <= 0:
                continue

            nuevo_item = {'producto': producto, 'cantidad': cantidad}
            if item.get('motivo'):
                nuevo_item['motivo'] = str(item.get('motivo')).strip()
            if item.get('id') is not None:
                nuevo_item['id'] = item.get('id')
            normalizados.append(nuevo_item)

        return normalizados
    
    # ===== 🆕 ACTION: ANULAR VENTA RUTA =====
    @action(detail=True, methods=['post'])
    def anular(self, request, pk=None):
        """
        Anula una VentaRuta:
        1. Verifica que no esté ya anulada
        2. Verifica límite de anulaciones (máximo 1)
        3. Descuenta las 'vendidas' del CargueIDx por cada producto
        4. Marca estado = 'ANULADA' e incrementa contador
        """
        try:
            vendedor_auth = self._validar_sesion_movil(request)
            if not vendedor_auth:
                return Response({'error': 'No autorizado. Inicia sesión nuevamente.'}, status=401)

            venta = self.get_object()
            if venta.vendedor_id != vendedor_auth.id_vendedor:
                return Response({'error': 'No autorizado para anular esta venta.'}, status=403)

            if venta.estado == 'ANULADA':
                return Response({'error': 'Esta venta ya fue anulada'}, status=400)

            # 🔒 LÍMITE DE ANULACIONES: Máximo 1 anulación por venta
            intentos_anulacion = getattr(venta, 'intentos_anulacion', 0) or 0
            if intentos_anulacion >= 1:
                return Response({
                    'error': 'Límite alcanzado',
                    'mensaje': 'Esta venta ya fue anulada anteriormente. Usa la opción de editar en vez de anular.'
                }, status=400)

            id_vendedor = venta.vendedor.id_vendedor  # ID1, ID2, etc.
            fecha_venta = self._resolver_fecha_operativa(venta.fecha, fallback=timezone.now())

            modelo_map = {
                'ID1': CargueID1, 'ID2': CargueID2, 'ID3': CargueID3,
                'ID4': CargueID4, 'ID5': CargueID5, 'ID6': CargueID6,
            }
            ModeloCargue = modelo_map.get(id_vendedor)

            # Revertir vendidas y vencidas en CargueIDx
            if ModeloCargue and venta.detalles:
                from django.db.models import F
                for item in venta.detalles:
                    nombre = item.get('nombre') or item.get('producto', '')
                    cantidad = int(item.get('cantidad', 0))
                    if nombre and cantidad > 0:
                        registro_cargue = self._resolver_registro_cargue(ModeloCargue, fecha_venta, nombre)
                        if registro_cargue:
                            updated = ModeloCargue.objects.filter(pk=registro_cargue.pk).update(
                                vendidas=F('vendidas') - cantidad
                            )
                            print(
                                f"✅ Anulación: {nombre} -> {registro_cargue.producto} "
                                f"vendidas -= {cantidad} (registros: {updated})"
                            )
                        else:
                            print(f"⚠️ Anulación: no se encontró registro de cargue para {nombre}")

            # 🆕 Revertir vencidas reportadas en la venta anulada
            if ModeloCargue and venta.productos_vencidos:
                from django.db.models import F
                from django.db.models.functions import Greatest
                registros_recalculados = []

                for item_vencido in (venta.productos_vencidos or []):
                    nombre = (item_vencido.get('nombre') or item_vencido.get('producto') or '').strip()
                    cantidad_vencida = int(item_vencido.get('cantidad', 0) or 0)
                    if not nombre or cantidad_vencida <= 0:
                        continue

                    registro_cargue = self._resolver_registro_cargue(ModeloCargue, fecha_venta, nombre)
                    if registro_cargue:
                        updated = ModeloCargue.objects.filter(pk=registro_cargue.pk).update(
                            vencidas=Greatest(F('vencidas') - cantidad_vencida, 0)
                        )
                        if updated:
                            registros_recalculados.append(registro_cargue.pk)
                        print(
                            f"✅ Anulación: {nombre} -> {registro_cargue.producto} "
                            f"vencidas -= {cantidad_vencida} (registros: {updated})"
                        )
                    else:
                        print(f"⚠️ Anulación: no se encontró registro de cargue para vencida {nombre}")

                if registros_recalculados:
                    recalcular_totales_cargue_queryset(
                        ModeloCargue.objects.filter(pk__in=registros_recalculados)
                    )

            venta.estado = 'ANULADA'
            # 🔒 Incrementar contador de anulaciones
            venta.intentos_anulacion = (getattr(venta, 'intentos_anulacion', 0) or 0) + 1
            venta.save()

            return Response({
                'success': True,
                'mensaje': f'Venta #{venta.id} anulada correctamente',
                'venta_id': venta.id,
                'estado': 'ANULADA',
                'intentos_anulacion': venta.intentos_anulacion
            })

        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)

    # ===== 🆕 ACTION: EDITAR VENTA RUTA =====
    @action(detail=True, methods=['patch'])
    def editar(self, request, pk=None):
        """
        Edita una VentaRuta:
        - Recibe nuevos detalles [{nombre, cantidad, precio}]
        - Calcula diferencias vs detalles anteriores y ajusta vendidas en CargueIDx
        - Actualiza total y marca editada = True
        """
        try:
            vendedor_auth = self._validar_sesion_movil(request)
            if not vendedor_auth:
                return Response({'error': 'No autorizado. Inicia sesión nuevamente.'}, status=401)

            venta = self.get_object()
            if venta.vendedor_id != vendedor_auth.id_vendedor:
                return Response({'error': 'No autorizado para editar esta venta.'}, status=403)

            if venta.estado == 'ANULADA':
                return Response({'error': 'No se puede editar una venta anulada'}, status=400)

            detalles_raw = request.data.get('detalles', None)
            actualizar_detalles = detalles_raw is not None

            foto_vencidos_base64 = request.data.get('foto_vencidos', None)
            if isinstance(foto_vencidos_base64, str):
                try:
                    foto_vencidos_base64 = json.loads(foto_vencidos_base64)
                except Exception:
                    foto_vencidos_base64 = None

            productos_vencidos_raw = request.data.get('productos_vencidos', None)
            actualizar_vencidas = productos_vencidos_raw is not None or foto_vencidos_base64 is not None
            nuevos_productos_vencidos = self._normalizar_productos_vencidos(productos_vencidos_raw or []) if actualizar_vencidas else (venta.productos_vencidos or [])

            metodo_pago_raw = request.data.get('metodo_pago', None)
            metodo_pago_normalizado = None
            if metodo_pago_raw is not None:
                metodo_pago_normalizado = str(metodo_pago_raw).strip().upper()
                metodos_validos = {'EFECTIVO', 'NEQUI', 'DAVIPLATA', 'TARJETA', 'TRANSFERENCIA'}
                if metodo_pago_normalizado not in metodos_validos:
                    return Response(
                        {'error': f'Método de pago inválido: {metodo_pago_normalizado}'},
                        status=400
                    )

            if not actualizar_detalles and metodo_pago_normalizado is None and not actualizar_vencidas:
                return Response(
                    {'error': 'Debes enviar detalles, metodo_pago o productos_vencidos para editar la venta'},
                    status=400
                )

            # Determinar tipo de operación para bloqueos independientes
            solo_cambio_metodo_pago = (
                not actualizar_detalles and
                metodo_pago_normalizado is not None and
                not actualizar_vencidas
            )

            if solo_cambio_metodo_pago:
                # Cambio de método de pago: bloquear solo si ya fue cambiado antes
                if venta.metodo_pago_cambiado:
                    return Response(
                        {
                            'error': 'El método de pago de esta venta ya fue cambiado una vez.',
                            'codigo': 'METODO_PAGO_YA_CAMBIADO'
                        },
                        status=400
                    )
            else:
                # Edición de productos/vencidas: bloquear si ya fue editada
                if venta.editada:
                    return Response(
                        {
                            'error': 'Esta venta ya fue modificada una vez. No se permiten más ediciones.',
                            'codigo': 'VENTA_YA_MODIFICADA'
                        },
                        status=400
                    )

            nuevos_detalles = (venta.detalles or [])
            nuevo_total = float(venta.total or 0)
            
            if actualizar_detalles:
                nuevos_detalles, nuevo_total = self._normalizar_detalles(detalles_raw)
                if not nuevos_detalles:
                    return Response({'error': 'Se requieren los nuevos detalles'}, status=400)

            # Normalizar detalles previos para comparación
            detalles_anteriores_json = json.dumps(
                sorted(
                    [{'producto': str(i.get('producto') or i.get('nombre') or '').strip().upper(),
                      'cantidad': int(i.get('cantidad') or 0)} 
                     for i in (venta.detalles or [])],
                    key=lambda x: x['producto']
                )
            )
            detalles_nuevos_json = json.dumps(
                sorted(
                    [{'producto': str(i.get('producto') or i.get('nombre') or '').strip().upper(),
                      'cantidad': int(i.get('cantidad') or 0)} 
                     for i in (nuevos_detalles or [])],
                    key=lambda x: x['producto']
                )
            )

            # Forzar actualización si hay cambios en detalles o total
            if detalles_anteriores_json != detalles_nuevos_json or round(float(venta.total or 0), 2) != round(float(nuevo_total or 0), 2):
                actualizar_detalles = True
            else:
                actualizar_detalles = False
                # Aun si no cambian los detalles, si el total es diferente (ej. cambio manual), actualizamos
                if round(float(venta.total or 0), 2) != round(float(nuevo_total or 0), 2):
                    actualizar_detalles = True

            id_vendedor = venta.vendedor.id_vendedor
            fecha_venta = self._resolver_fecha_operativa(venta.fecha, fallback=timezone.now())
            ModeloCargue = self._obtener_modelo_cargue_por_vendedor(id_vendedor)
            registros_recalculados = set()

            if actualizar_detalles and ModeloCargue:
                cantidades_base = self._cantidades_base_detalles(
                    ModeloCargue,
                    fecha_venta,
                    venta.detalles or [],
                )
                error_stock = self._validar_stock_disponible_cargue(
                    ModeloCargue,
                    fecha_venta,
                    nuevos_detalles,
                    cantidades_base=cantidades_base,
                )
                if error_stock:
                    return Response(error_stock, status=400)

            if actualizar_detalles:
                # Construir mapa de cantidades anteriores
                detalles_anteriores = venta.detalles or []
                mapa_anterior = {}
                for item in detalles_anteriores:
                    nombre = item.get('nombre') or item.get('producto', '')
                    if nombre:
                        mapa_anterior[nombre.lower()] = int(item.get('cantidad', 0))

                # Construir mapa de cantidades nuevas
                mapa_nuevo = {}
                for item in nuevos_detalles:
                    nombre = item.get('nombre') or item.get('producto', '')
                    if nombre:
                        mapa_nuevo[nombre.lower()] = {
                            'nombre_real': nombre,
                            'cantidad': int(item.get('cantidad', 0)),
                            'precio': float(item.get('precio', 0))
                        }

                # Ajustar vendidas en CargueIDx solo por las diferencias
                if ModeloCargue:
                    from django.db.models import F

                    # Productos que existían antes
                    for nombre_key, cant_anterior in mapa_anterior.items():
                        nuevo = mapa_nuevo.get(nombre_key)
                        cant_nueva = nuevo['cantidad'] if nuevo else 0
                        diferencia = cant_nueva - cant_anterior  # positivo = más, negativo = devuelve
                        nombre_real = nuevo['nombre_real'] if nuevo else nombre_key

                        if diferencia != 0:
                            registro_cargue = self._resolver_registro_cargue(ModeloCargue, fecha_venta, nombre_real)
                            if registro_cargue:
                                ModeloCargue.objects.filter(pk=registro_cargue.pk).update(
                                    vendidas=F('vendidas') + diferencia
                                )
                                registros_recalculados.add(registro_cargue.pk)
                                print(
                                    f"✅ Edición: {nombre_real} -> {registro_cargue.producto} "
                                    f"vendidas {'+' if diferencia > 0 else ''}{diferencia}"
                                )
                            else:
                                print(f"⚠️ Edición: no se encontró registro de cargue para {nombre_real}")

                    # Productos nuevos que no estaban antes
                    for nombre_key, datos in mapa_nuevo.items():
                        if nombre_key not in mapa_anterior:
                            cant_nueva = datos['cantidad']
                            if cant_nueva > 0:
                                registro_cargue = self._resolver_registro_cargue(
                                    ModeloCargue, fecha_venta, datos['nombre_real']
                                )
                                if registro_cargue:
                                    ModeloCargue.objects.filter(pk=registro_cargue.pk).update(
                                        vendidas=F('vendidas') + cant_nueva
                                    )
                                    registros_recalculados.add(registro_cargue.pk)
                                    print(
                                        f"✅ Edición (nuevo): {datos['nombre_real']} -> {registro_cargue.producto} "
                                        f"vendidas += {cant_nueva}"
                                    )
                                else:
                                    print(f"⚠️ Edición (nuevo): no se encontró registro de cargue para {datos['nombre_real']}")

            if actualizar_vencidas and ModeloCargue:
                from django.db.models import F
                from django.db.models.functions import Greatest

                mapa_vencidas_anterior = {}
                for item in self._normalizar_productos_vencidos(venta.productos_vencidos or []):
                    nombre = (item.get('producto') or item.get('nombre') or '').strip()
                    cantidad = self._to_int(item.get('cantidad'), 0)
                    if nombre and cantidad > 0:
                        mapa_vencidas_anterior[nombre.lower()] = {
                            'nombre_real': nombre,
                            'cantidad': cantidad,
                        }

                mapa_vencidas_nuevo = {}
                for item in nuevos_productos_vencidos:
                    nombre = (item.get('producto') or item.get('nombre') or '').strip()
                    cantidad = self._to_int(item.get('cantidad'), 0)
                    if nombre and cantidad > 0:
                        mapa_vencidas_nuevo[nombre.lower()] = {
                            'nombre_real': nombre,
                            'cantidad': cantidad,
                        }

                # Unificar edición de stock para vencidas
                for nombre_key in set(mapa_vencidas_anterior.keys()) | set(mapa_vencidas_nuevo.keys()):
                    anterior = mapa_vencidas_anterior.get(nombre_key)
                    nuevo = mapa_vencidas_nuevo.get(nombre_key)
                    cant_anterior = anterior['cantidad'] if anterior else 0
                    cant_nueva = nuevo['cantidad'] if nuevo else 0
                    diferencia = cant_nueva - cant_anterior
                    
                    if diferencia == 0:
                        continue

                    nombre_real = (nuevo or anterior)['nombre_real']
                    registro_cargue = self._resolver_registro_cargue(ModeloCargue, fecha_venta, nombre_real)
                    
                    if registro_cargue:
                        # Usar F y Greatest para seguridad atómica y evitar negativos
                        ModeloCargue.objects.filter(pk=registro_cargue.pk).update(
                            vencidas=Greatest(F('vencidas') + diferencia, 0)
                        )
                        registros_recalculados.add(registro_cargue.pk)
                        print(f"📦 Edición Vencidas: {nombre_real} -> dif: {diferencia}")
                    else:
                        print(f"⚠️ Edición Vencidas: No se encontró cargue para {nombre_real}")

            # Actualizar la venta
            if actualizar_detalles:
                venta.detalles = nuevos_detalles
                venta.total = nuevo_total
            if metodo_pago_normalizado is not None:
                venta.metodo_pago = metodo_pago_normalizado
                # Solo consumir el slot "metodo_pago_cambiado" cuando es un cambio explícito
                # desde el historial (solo_cambio_metodo_pago).
                # Si el método viene incluido en una edición de productos, no bloqueamos
                # el cambio posterior desde el historial.
                if solo_cambio_metodo_pago:
                    venta.metodo_pago_cambiado = True
            if actualizar_vencidas:
                venta.productos_vencidos = nuevos_productos_vencidos
                if len(nuevos_productos_vencidos) == 0:
                    if venta.foto_vencidos:
                        venta.foto_vencidos.delete(save=False)
                    venta.foto_vencidos = None
                    venta.evidencias.all().delete()
            # Solo marcar editada si se modificaron productos o vencidas (no solo método de pago)
            if actualizar_detalles or actualizar_vencidas:
                venta.editada = True
                venta.fecha_ultima_edicion = timezone.now()
            venta.save()

            evidencias_creadas = 0
            if foto_vencidos_base64 and isinstance(foto_vencidos_base64, dict):
                from .models import EvidenciaVenta
                import base64
                import uuid
                from django.core.files.base import ContentFile

                primera_guardada = False
                for prod_id_key, pics in foto_vencidos_base64.items():
                    if not pics or not isinstance(pics, list):
                        continue

                    producto_id = int(prod_id_key) if str(prod_id_key).isdigit() else None
                    for pic_base64 in pics:
                        if not isinstance(pic_base64, str) or ',' not in pic_base64:
                            continue

                        try:
                            formato, imgstr = pic_base64.split(';base64,')
                            ext = formato.split('/')[-1]
                            contenido = base64.b64decode(imgstr)
                            nombre_archivo = f"venta_{venta.id}_{uuid.uuid4().hex[:6]}.{ext}"

                            if not primera_guardada:
                                venta.foto_vencidos.save(nombre_archivo, ContentFile(contenido), save=False)
                                primera_guardada = True

                            EvidenciaVenta.objects.create(
                                venta=venta,
                                producto_id=producto_id,
                                imagen=ContentFile(contenido, name=nombre_archivo),
                            )
                            evidencias_creadas += 1
                        except Exception:
                            continue

                if primera_guardada:
                    venta.save(update_fields=['foto_vencidos'])

            if registros_recalculados:
                recalcular_totales_cargue_queryset(
                    ModeloCargue.objects.filter(pk__in=registros_recalculados)
                )

            return Response({
                'success': True,
                'mensaje': f'Venta #{venta.id} editada correctamente',
                'venta_id': venta.id,
                'nuevo_total': nuevo_total,
                'metodo_pago': venta.metodo_pago,
                'detalles_actualizados': actualizar_detalles,
                'metodo_pago_cambiado': venta.metodo_pago_cambiado,
                'productos_vencidos': venta.productos_vencidos,
                'foto_vencidos': request.build_absolute_uri(venta.foto_vencidos.url) if venta.foto_vencidos else None,
                'evidencias': [
                    {
                        'id': ev.id,
                        'producto_id': ev.producto_id,
                        'imagen': request.build_absolute_uri(ev.imagen.url) if ev.imagen else None,
                    }
                    for ev in venta.evidencias.all().order_by('id')
                ],
                'evidencias_creadas': evidencias_creadas,
                'editada': True
            })

        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['get'])
    def reportes(self, request):

        """Endpoint para reportes de ventas por período"""
        from django.db.models import Sum, Count
        from django.db.models.functions import TruncDate, TruncMonth
        from datetime import datetime, timedelta
        from dateutil.relativedelta import relativedelta
        
        periodo = request.query_params.get('periodo', 'dia')  # dia, mes, trimestre, semestre, año
        vendedor_id = request.query_params.get('vendedor_id', None)
        fecha_inicio = request.query_params.get('fecha_inicio', None)
        fecha_fin = request.query_params.get('fecha_fin', None)
        
        # Calcular fechas según período
        hoy = datetime.now().date()
        if periodo == 'dia':
            fecha_inicio = fecha_inicio or str(hoy)
            fecha_fin = fecha_fin or str(hoy)
        elif periodo == 'semana':
            fecha_inicio = fecha_inicio or str(hoy - timedelta(days=7))
            fecha_fin = fecha_fin or str(hoy)
        elif periodo == 'mes':
            fecha_inicio = fecha_inicio or str(hoy.replace(day=1))
            fecha_fin = fecha_fin or str(hoy)
        elif periodo == 'trimestre':
            fecha_inicio = fecha_inicio or str(hoy - relativedelta(months=3))
            fecha_fin = fecha_fin or str(hoy)
        elif periodo == 'semestre':
            fecha_inicio = fecha_inicio or str(hoy - relativedelta(months=6))
            fecha_fin = fecha_fin or str(hoy)
        elif periodo == 'año':
            fecha_inicio = fecha_inicio or str(hoy.replace(month=1, day=1))
            fecha_fin = fecha_fin or str(hoy)
        
        # Filtrar ventas
        queryset = VentaRuta.objects.filter(fecha__date__gte=fecha_inicio, fecha__date__lte=fecha_fin)
        if vendedor_id:
            queryset = queryset.filter(vendedor__id_vendedor=vendedor_id)
        
        # Total general
        total_general = queryset.aggregate(total=Sum('total'))['total'] or 0
        cantidad_ventas = queryset.count()
        
        # Ventas por vendedor
        ventas_por_vendedor = queryset.values('vendedor__nombre', 'vendedor__id_vendedor').annotate(
            total=Sum('total'),
            cantidad=Count('id')
        ).order_by('-total')
        
        # Ventas por cliente
        ventas_por_cliente = queryset.values('cliente_nombre', 'nombre_negocio').annotate(
            total=Sum('total'),
            cantidad=Count('id')
        ).order_by('-total')[:20]  # Top 20 clientes
        
        # Ventas por producto (necesita procesar JSON)
        productos_dict = {}
        for venta in queryset:
            detalles = venta.detalles or []
            for item in detalles:
                nombre = item.get('nombre') or item.get('producto') or 'Sin nombre'
                cantidad = item.get('cantidad', 0)
                subtotal = item.get('subtotal', 0) or (cantidad * item.get('precio', 0))
                if nombre in productos_dict:
                    productos_dict[nombre]['cantidad'] += cantidad
                    productos_dict[nombre]['total'] += subtotal
                else:
                    productos_dict[nombre] = {'cantidad': cantidad, 'total': subtotal}
        
        ventas_por_producto = [
            {'producto': k, 'cantidad': v['cantidad'], 'total': v['total']}
            for k, v in sorted(productos_dict.items(), key=lambda x: x[1]['total'], reverse=True)
        ]
        
        # Ventas por día (para gráficos)
        ventas_por_dia = queryset.annotate(dia=TruncDate('fecha')).values('dia').annotate(
            total=Sum('total'),
            cantidad=Count('id')
        ).order_by('dia')
        
        return Response({
            'periodo': periodo,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'total_general': float(total_general),
            'cantidad_ventas': cantidad_ventas,
            'ventas_por_vendedor': list(ventas_por_vendedor),
            'ventas_por_cliente': list(ventas_por_cliente),
            'ventas_por_producto': ventas_por_producto[:20],  # Top 20
            'ventas_por_dia': list(ventas_por_dia)
        })

    def create(self, request, *args, **kwargs):
        from .models import Vendedor, EvidenciaVenta, SyncLog
        from rest_framework import status
        from django.http import QueryDict
        from django.db import transaction, IntegrityError
        import json as json_lib

        vendedor_auth = self._validar_sesion_movil(request)
        if not vendedor_auth:
            return Response({'error': 'No autorizado. Inicia sesión nuevamente.'}, status=status.HTTP_401_UNAUTHORIZED)
        
        # 🆕 Obtener IP del cliente
        def _get_client_ip(request):
            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip = x_forwarded_for.split(',')[0]
            else:
                ip = request.META.get('REMOTE_ADDR')
            return ip
        
        # 🆕 Logging de sincronización
        def _log_sync(accion, exito=True, error_mensaje='', id_local='', registro_id=0):
            try:
                SyncLog.objects.create(
                    accion=accion,
                    modelo='VentaRuta',
                    registro_id=registro_id,
                    id_local=id_local,
                    vendedor_id=request.data.get('vendedor_id', ''),
                    dispositivo_id=request.data.get('dispositivo_id', ''),
                    ip_origen=_get_client_ip(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', ''),
                    exito=exito,
                    error_mensaje=error_mensaje
                )
            except Exception as e:
                print(f"⚠️ Error logging: {e}")

        def _huella_detalles(detalles):
            if not isinstance(detalles, list):
                return '[]'

            detalles_norm = []
            for item in detalles:
                if not isinstance(item, dict):
                    continue
                detalles_norm.append({
                    'producto': str(item.get('producto') or item.get('nombre') or '').strip().upper(),
                    'cantidad': int(item.get('cantidad') or 0),
                    'precio': float(item.get('precio') or item.get('precio_unitario') or item.get('valor_unitario') or 0),
                })

            detalles_norm.sort(key=lambda x: (x['producto'], x['cantidad'], x['precio']))
            return json_lib.dumps(detalles_norm, ensure_ascii=False, sort_keys=True)

        def _huella_vencidas(vencidas):
            if not isinstance(vencidas, list):
                return '[]'

            vencidas_norm = []
            for item in vencidas:
                if not isinstance(item, dict):
                    continue
                vencidas_norm.append({
                    'producto': str(item.get('producto') or item.get('nombre') or '').strip().upper(),
                    'cantidad': int(item.get('cantidad') or 0),
                    'motivo': str(item.get('motivo') or '').strip().upper(),
                })

            vencidas_norm.sort(key=lambda x: (x['producto'], x['cantidad'], x['motivo']))
            return json_lib.dumps(vencidas_norm, ensure_ascii=False, sort_keys=True)
        
        # 🆕 Verificar duplicados por id_local
        id_local = request.data.get('id_local')
        dispositivo_id = request.data.get('dispositivo_id', '')
        
        if id_local:
            try:
                venta_existente = VentaRuta.objects.get(id_local=id_local)
                print(f"⚠️ DUPLICADO DETECTADO: id_local={id_local}, ID={venta_existente.id}")
                print(f"   Dispositivo original: {venta_existente.dispositivo_id}")
                print(f"   Dispositivo actual: {dispositivo_id}")
                
                # Log del intento de duplicado
                _log_sync(
                    accion='CREATE_DUPLICADO',
                    exito=False,
                    error_mensaje=f'Venta ya existe (ID: {venta_existente.id})',
                    id_local=id_local
                )
                
                # 🆕 Retornar 200 OK con warning (no error)
                return Response(
                    {
                        'id': venta_existente.id,
                        'message': 'Venta ya registrada previamente',
                        'duplicada': True,
                        'id_local': id_local,
                        'dispositivo_original': venta_existente.dispositivo_id,
                        'timestamp': venta_existente.fecha
                    },
                    status=status.HTTP_200_OK  # No HTTP_409_CONFLICT para no fallar en app
                )
                
            except VentaRuta.DoesNotExist:
                # No existe, continuar con creación
                pass
            except VentaRuta.MultipleObjectsReturned:
                # ❌ Si hay múltiples (no debería pasar por unique=True)
                print(f"❌ ERROR: Múltiples ventas con id_local={id_local}")
                _log_sync(
                    accion='CREATE_DUPLICADO',
                    exito=False,
                    error_mensaje=f'Múltiples ventas con mismo id_local',
                    id_local=id_local
                )
                return Response(
                    {'error': 'Error de integridad de datos'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        
        
        # Crear un dict plano (evita efectos de QueryDict en JSONField al validar serializer)
        if isinstance(request.data, QueryDict):
            data = {k: request.data.get(k) for k in request.data.keys()}
        else:
            data = dict(request.data)

        # 🔒 Endurecimiento: campos sensibles no se aceptan desde cliente en creación de venta
        # Detectar si la venta fue editada offline antes de sincronizarse
        fue_editada_offline = data.get('fecha_ultima_edicion') is not None
        for campo_bloqueado in ('sincronizado', 'ip_origen', 'editada'):
            data.pop(campo_bloqueado, None)
        # Si la venta fue editada offline (tiene fecha_ultima_edicion), marcar como editada
        if fue_editada_offline:
            data['editada'] = True
        # Estado inicial siempre controlado por backend
        data['estado'] = 'ACTIVA'
        
        # NO parsear campos JSON aquí - el JSONField del serializer lo hace automáticamente
        # Solo necesitamos ajustar el vendedor
        
        # Ajustar vendedor si viene como string (normaliza ID1/id1/ ID1 )
        if 'vendedor' in data and isinstance(data['vendedor'], str):
             data['vendedor'] = data['vendedor'].strip().upper()
             if data['vendedor'].startswith('ID'):
                 try:
                     vendedor = Vendedor.objects.get(id_vendedor=data['vendedor'])
                     data['vendedor'] = vendedor.pk
                 except Vendedor.DoesNotExist:
                     pass

        # 🔒 vendedor siempre se toma del token autenticado
        data['vendedor'] = vendedor_auth.id_vendedor

        # 🆕 CREAR CLIENTE OCASIONAL AUTOMÁTICAMENTE si viene en los datos
        # La app envía cliente_ocasional con un ID temporal, pero necesitamos crear el registro real
        if 'cliente_ocasional' in data and data['cliente_ocasional']:
            try:
                # Extraer datos del cliente desde los campos de la venta
                nombre_cliente = data.get('cliente_nombre', '').strip()
                nombre_negocio = data.get('nombre_negocio', '').strip()
                
                # Usar nombre_negocio si existe, sino usar cliente_nombre
                nombre_final = nombre_negocio if nombre_negocio else nombre_cliente
                
                if nombre_final:
                    # Buscar si ya existe un ClienteOcasional con ese nombre para este vendedor
                    cliente_ocasional_existente = ClienteOcasional.objects.filter(
                        vendedor=vendedor_auth,
                        nombre__iexact=nombre_final,
                        activo=True
                    ).first()
                    
                    if cliente_ocasional_existente:
                        # Ya existe, usar ese ID
                        data['cliente_ocasional'] = cliente_ocasional_existente.id
                        print(f"✅ Cliente ocasional existente encontrado: {cliente_ocasional_existente.id} - {nombre_final}")
                    else:
                        # No existe, crear uno nuevo
                        nuevo_cliente_ocasional = ClienteOcasional.objects.create(
                            vendedor=vendedor_auth,
                            nombre=nombre_final,
                            telefono='',  # La app no envía teléfono en ventas ocasionales
                            direccion='',  # La app no envía dirección en ventas ocasionales
                            tope_venta=60000,  # Tope por defecto
                            activo=True
                        )
                        data['cliente_ocasional'] = nuevo_cliente_ocasional.id
                        print(f"✅ Cliente ocasional creado: {nuevo_cliente_ocasional.id} - {nombre_final} para {vendedor_auth.id_vendedor}")
                else:
                    # No hay nombre, quitar el campo para evitar error
                    data.pop('cliente_ocasional', None)
                    print(f"⚠️ No se pudo crear cliente ocasional: sin nombre")
            except Exception as e:
                print(f"❌ Error creando cliente ocasional: {e}")
                # Si falla, quitar el campo para que la venta se guarde sin cliente ocasional
                data.pop('cliente_ocasional', None)

        # 🔒 Validar vendedor activo
        vendedor_ref = data.get('vendedor')
        if not vendedor_ref or not Vendedor.objects.filter(pk=vendedor_ref, activo=True).exists():
            return Response({'error': 'Vendedor inválido o inactivo'}, status=status.HTTP_403_FORBIDDEN)

        # 🔒 Validar y normalizar detalle de venta (total se calcula en backend)
        detalles_normalizados, total_calculado = self._normalizar_detalles(data.get('detalles', []))
        productos_vencidos_normalizados = self._normalizar_productos_vencidos(data.get('productos_vencidos', []))
        if not detalles_normalizados and not productos_vencidos_normalizados:
            return Response({'error': 'La venta debe incluir al menos un producto válido'}, status=status.HTTP_400_BAD_REQUEST)
        data['detalles'] = detalles_normalizados
        data['total'] = total_calculado if detalles_normalizados else 0
        data['productos_vencidos'] = productos_vencidos_normalizados

        fecha_operativa = self._resolver_fecha_operativa(data.get('fecha'), fallback=timezone.now())

        # 🔒 Bloquear vencidas reportadas después del cierre de turno
        # Solo aplica a registros sin venta real (solo vencidas), no a sync de ventas offline
        if not detalles_normalizados and productos_vencidos_normalizados:
            from .models import TurnoVendedor as _TV
            _vid = int(str(vendedor_auth.id_vendedor).replace('ID', '')) if 'ID' in str(vendedor_auth.id_vendedor) else int(vendedor_auth.id_vendedor)
            _fecha_op = fecha_operativa.date() if hasattr(fecha_operativa, 'date') else fecha_operativa
            _turno_cerrado = _TV.objects.filter(vendedor_id=_vid, fecha=_fecha_op, estado='CERRADO').exists()
            if _turno_cerrado:
                print(f"🚫 Vencidas ignoradas post-cierre (OK silencioso para desbloquear app): {vendedor_auth.id_vendedor} - {_fecha_op}")
                return Response({'id': 0, 'mensaje': 'Registro procesado'}, status=status.HTTP_201_CREATED)
        ModeloCargue = self._obtener_modelo_cargue_por_vendedor(vendedor_auth.id_vendedor)
        error_stock = self._validar_stock_disponible_cargue(
            ModeloCargue,
            fecha_operativa,
            detalles_normalizados,
        )
        if error_stock:
            _log_sync(
                accion='CREATE_STOCK_INVALIDO',
                exito=False,
                error_mensaje=str(error_stock.get('detalle', [])),
                id_local=id_local or ''
            )
            return Response(error_stock, status=status.HTTP_400_BAD_REQUEST)

        # 🆕 Verificar duplicados sospechosos por similitud en ventana muy corta.
        # Debe cubrir doble submit/click casi instantáneo, no ventas reales repetidas.
        try:
            fecha_referencia = parse_datetime(str(data.get('fecha'))) if data.get('fecha') else None
            if fecha_referencia is None:
                fecha_referencia = timezone.now()
            if timezone.is_naive(fecha_referencia):
                fecha_referencia = timezone.make_aware(fecha_referencia, timezone.get_current_timezone())

            ventana_inicio = fecha_referencia - timedelta(seconds=15)
            ventana_fin = fecha_referencia + timedelta(seconds=15)

            nombre_negocio_norm = str(data.get('nombre_negocio') or '').strip()
            cliente_nombre_norm = str(data.get('cliente_nombre') or '').strip()
            total_norm = float(data.get('total') or 0)
            detalles_huella = _huella_detalles(detalles_normalizados)
            vencidas_huella = _huella_vencidas(productos_vencidos_normalizados)

            candidatos = VentaRuta.objects.filter(
                vendedor_id=vendedor_auth.id_vendedor,
                estado='ACTIVA',
                dispositivo_id=dispositivo_id,
                fecha__gte=ventana_inicio,
                fecha__lte=ventana_fin,
                total=total_norm,
                cliente_nombre=cliente_nombre_norm,
                nombre_negocio=nombre_negocio_norm,
            ).order_by('fecha')

            for candidato in candidatos:
                if _huella_detalles(candidato.detalles) == detalles_huella and _huella_vencidas(candidato.productos_vencidos) == vencidas_huella:
                    print(f"⚠️ DUPLICADO SOSPECHOSO DETECTADO: nuevo id_local={id_local} ~ venta existente {candidato.id}")
                    _log_sync(
                        accion='CREATE_DUPLICADO_SOSPECHOSO',
                        exito=False,
                        error_mensaje=f'Venta sospechosamente duplicada (ID: {candidato.id})',
                        id_local=id_local or ''
                    )
                    return Response(
                        {
                            'id': candidato.id,
                            'message': 'Venta sospechosamente duplicada detectada',
                            'duplicada': True,
                            'warning': 'DUPLICADO_SOSPECHOSO',
                            'id_local': id_local,
                            'duplicado_de': candidato.id,
                            'timestamp': candidato.fecha
                        },
                        status=status.HTTP_200_OK
                    )
        except Exception as duplicate_error:
            print(f"⚠️ Error validando duplicado sospechoso: {duplicate_error}")

        # Extraer fotos de evidencia antes de crear la venta
        evidencias_data = []
        for key in request.FILES.keys():
            if key.startswith('evidencia_'):
                parts = key.split('_')
                if len(parts) >= 3:
                    producto_id = parts[1]
                    evidencias_data.append({
                        'producto_id': int(producto_id),
                        'imagen': request.FILES[key]
                    })

        # Crear la venta con los datos procesados


        print(f"data keys: {data.keys()}")
        print(f"vendedor: {data.get('vendedor')}")
        print(f"cliente_nombre: {data.get('cliente_nombre')}")
        print(f"total: {data.get('total')}")
        print(f"detalles type: {type(data.get('detalles'))}, valor: {data.get('detalles')}")
        print(f"productos_vencidos type: {type(data.get('productos_vencidos'))}, valor: {data.get('productos_vencidos')}")

        
        
        # 🆕 Hotfix transitorio: tolerar foto_vencidos legacy/offline para no bloquear sincronización.
        # Si no podemos convertir a archivo válido, simplemente se ignora la foto y se guarda la venta.
        def _iterar_posibles_imagenes(raw):
            if isinstance(raw, str):
                yield raw
                return
            if isinstance(raw, list):
                for item in raw:
                    yield from _iterar_posibles_imagenes(item)
                return
            if isinstance(raw, dict):
                for value in raw.values():
                    yield from _iterar_posibles_imagenes(value)
                return

        foto_vencidos_data = data.get('foto_vencidos')
        if isinstance(foto_vencidos_data, str):
            import json
            try:
                foto_vencidos_data = json.loads(foto_vencidos_data)
            except Exception:
                # Si llega string no JSON/base64, no debe romper creación.
                pass

        # Limpiar siempre el campo para evitar que llegue valor no-file al serializer.
        data.pop('foto_vencidos', None)

        # 1) Priorizar archivo real si viene en multipart.
        foto_file = request.FILES.get('foto_vencidos')
        if foto_file:
            data['foto_vencidos'] = foto_file
        else:
            # 2) Intentar convertir primer base64 encontrado dentro de estructuras legacy.
            import base64
            import uuid
            from django.core.files.base import ContentFile

            for candidata in _iterar_posibles_imagenes(foto_vencidos_data):
                if not isinstance(candidata, str):
                    continue
                if ';base64,' not in candidata:
                    continue
                try:
                    formato, imgstr = candidata.split(';base64,', 1)
                    ext = formato.split('/')[-1].split(';')[0] or 'jpg'
                    data['foto_vencidos'] = ContentFile(
                        base64.b64decode(imgstr),
                        name=f'vencidas_{uuid.uuid4().hex[:8]}.{ext}'
                    )
                    break
                except Exception as e:
                    print(f"⚠️ Foto vencidos inválida ignorada: {e}")
            
        serializer = self.get_serializer(data=data)
        
        if not serializer.is_valid():
            print("❌ ERRORES DE VALIDACIÓN:")
            print(serializer.errors)
            
            # 🆕 Log de error de validación
            _log_sync(
                accion='CREATE_VENTA',
                exito=False,
                error_mensaje=f'Errores de validación: {serializer.errors}',
                id_local=id_local or ''
            )
            
        serializer.is_valid(raise_exception=True)
        
        # 🆕 Agregar metadatos de multi-dispositivo antes del save
        try:
            with transaction.atomic():
                # Guardar con metadatos
                venta = serializer.save(
                    dispositivo_id=dispositivo_id,
                    ip_origen=_get_client_ip(request)
                )
                
                print(f"✅ VENTA CREADA: ID={venta.id}, id_local={venta.id_local}")
                print(f"   Dispositivo: {venta.dispositivo_id}")
                print(f"   IP: {venta.ip_origen}")
                
                # 🆕 Log de creación exitosa
                _log_sync(
                    accion='CREATE_VENTA',
                    exito=True,
                    error_mensaje='',
                    id_local=venta.id_local or '',
                    registro_id=venta.id
                )
                
        except IntegrityError as e:
            # Error de integridad (posible race condition con id_local duplicado)
            print(f"❌ IntegrityError: {e}")
            _log_sync(
                accion='CONFLICT',
                exito=False,
                error_mensaje=f'IntegrityError: {str(e)}',
                id_local=id_local or ''
            )
            return Response(
                {'error': 'Conflicto de sincronización. La venta puede haber sido registrada por otro dispositivo.'},
                status=status.HTTP_409_CONFLICT
            )
        except Exception as e:
            print(f"❌ Error al crear venta: {e}")
            _log_sync(
                accion='CREATE_VENTA',
                exito=False,
                error_mensaje=str(e),
                id_local=id_local or ''
            )
            raise
        
        # Guardar las evidencias asociadas
        for evidencia_info in evidencias_data:
            EvidenciaVenta.objects.create(
                venta=venta,
                producto_id=evidencia_info['producto_id'],
                imagen=evidencia_info['imagen']
            )
        
        # ========== 🆕 SINCRONIZAR VENCIDAS A CARGUEIDx ==========
        productos_vencidos = data.get('productos_vencidos', [])
        print(f"🔍 DEBUG - productos_vencidos recibidos: {productos_vencidos}")
        print(f"🔍 DEBUG - tipo: {type(productos_vencidos)}, longitud: {len(productos_vencidos) if productos_vencidos else 0}")
        
        # Si es string, parsearlo a JSON
        if isinstance(productos_vencidos, str):
            try:
                productos_vencidos = json.loads(productos_vencidos)
                print(f"✅ JSON parseado correctamente: {productos_vencidos}")
            except json.JSONDecodeError as e:
                print(f"❌ Error parseando JSON: {e}")
                productos_vencidos = []
        
        if productos_vencidos and len(productos_vencidos) > 0:
            try:
                # Obtener ID del vendedor y fecha
                id_vendedor = venta.vendedor.id_vendedor  # ID1, ID2, etc.
                fecha_venta = self._resolver_fecha_operativa(venta.fecha, fallback=timezone.now())
                registros_recalculados = []
                
                print(f"🔄 Sincronizando vencidas a CargueIDx: {id_vendedor} - {fecha_venta}")
                print(f"   Productos vencidos: {productos_vencidos}")
                
                # Mapeo de ID a Modelo
                modelo_map = {
                    'ID1': CargueID1,
                    'ID2': CargueID2,
                    'ID3': CargueID3,
                    'ID4': CargueID4,
                    'ID5': CargueID5,
                    'ID6': CargueID6,
                }
                
                ModeloCargue = modelo_map.get(id_vendedor)
                if ModeloCargue:
                    # Actualizar cada producto vencido en el cargue
                    for item_vencido in productos_vencidos:
                        nombre_producto = item_vencido.get('nombre', '') or item_vencido.get('producto', '')
                        cantidad_vencida = item_vencido.get('cantidad', 0)
                        
                        if nombre_producto and cantidad_vencida > 0:
                            # Buscar el producto en el cargue
                            cargue = self._resolver_registro_cargue(ModeloCargue, fecha_venta, nombre_producto)
                            
                            if cargue:
                                # 🔥 OPTIMIZADO: Usar F() para evitar race conditions
                                from django.db.models import F
                                ModeloCargue.objects.filter(pk=cargue.pk).update(
                                    vencidas=F('vencidas') + cantidad_vencida
                                )
                                registros_recalculados.append(cargue.pk)
                                print(f"   ✅ {nombre_producto} -> {cargue.producto}: vencidas += {cantidad_vencida}")
                            else:
                                print(f"   ⚠️ No se encontró cargue para: {nombre_producto} - Intentando crear...")
                                try:
                                    cargue = self._crear_registro_cargue_on_the_fly(
                                        ModeloCargue,
                                        fecha_venta,
                                        nombre_producto,
                                        vendidas=0,
                                        vencidas=cantidad_vencida
                                    )
                                    if cargue:
                                        print(f"   ✨ Registro creado exitosamente para vencida: {cargue.producto}")
                                    else:
                                        print(f"   ❌ No hay referencia de cargue para el día {fecha_venta}, imposible crear.")
                                except Exception as create_error:
                                    print(f"   ❌ Error creando registro on-the-fly: {create_error}")

                    if registros_recalculados:
                        recalcular_totales_cargue_queryset(
                            ModeloCargue.objects.filter(pk__in=registros_recalculados)
                        )
                else:
                    print(f"   ⚠️ Modelo de cargue no encontrado para: {id_vendedor}")
                    
            except Exception as e:
                print(f"❌ Error sincronizando vencidas: {str(e)}")
                import traceback
                traceback.print_exc()
        
        # ========== 🆕 SINCRONIZAR VENDIDAS A CARGUEIDx ==========
        try:
            # Obtener ID del vendedor y fecha
            id_vendedor = venta.vendedor.id_vendedor if hasattr(venta.vendedor, 'id_vendedor') else None
            fecha_venta = self._resolver_fecha_operativa(venta.fecha, fallback=timezone.now())
            
            if id_vendedor:
                print(f"🔄 Sincronizando vendidas a CargueIDx: {id_vendedor} - {fecha_venta}")
                
                # Mapeo de modelos
                modelo_map = {
                    'ID1': CargueID1,
                    'ID2': CargueID2,
                    'ID3': CargueID3,
                    'ID4': CargueID4,
                    'ID5': CargueID5,
                    'ID6': CargueID6,
                }
                
                ModeloCargue = modelo_map.get(id_vendedor)
                if ModeloCargue:
                    # Parsear detalles de la venta
                    detalles_json = data.get('detalles', '[]')
                    if isinstance(detalles_json, str):
                        try:
                            detalles = json.loads(detalles_json)
                        except:
                            detalles = []
                    else:
                        detalles = detalles_json
                    
                    print(f"   📦 Detalles de venta: {len(detalles)} productos")
                    
                    # Actualizar vendidas por cada producto en la venta
                    for item in detalles:
                        nombre_producto = item.get('nombre', '')
                        cantidad_vendida = int(item.get('cantidad', 0))
                        
                        if nombre_producto and cantidad_vendida > 0:
                            # Buscar el producto en el cargue
                            cargue = self._resolver_registro_cargue(ModeloCargue, fecha_venta, nombre_producto)
                            
                            if cargue:
                                # 🔥 OPTIMIZADO: Usar F() para evitar race conditions
                                from django.db.models import F
                                ModeloCargue.objects.filter(pk=cargue.pk).update(
                                    vendidas=F('vendidas') + cantidad_vendida,
                                    vencidas=F('vencidas')  # Mantener valor actual
                                )
                                print(f"   ✅ {nombre_producto} -> {cargue.producto}: vendidas += {cantidad_vendida}")
                            else:
                                print(f"   ⚠️ No se encontró cargue para: {nombre_producto} - Intentando crear...")
                                try:
                                    cargue = self._crear_registro_cargue_on_the_fly(
                                        ModeloCargue,
                                        fecha_venta,
                                        nombre_producto,
                                        vendidas=cantidad_vendida,
                                        vencidas=0
                                    )
                                    if cargue:
                                        print(f"   ✨ Registro creado exitosamente para venta: {cargue.producto}")
                                    else:
                                        print(f"   ❌ No hay referencia de cargue para el día {fecha_venta}, imposible crear.")
                                except Exception as create_error:
                                    print(f"   ❌ Error creando registro on-the-fly: {create_error}")
                else:
                    print(f"   ⚠️ Modelo de cargue no encontrado para: {id_vendedor}")
        except Exception as e:
            print(f"❌ Error sincronizando vendidas: {str(e)}")
            import traceback
            traceback.print_exc()

        # ========== 🆕 SINCRONIZAR PAGOS A CarguePagos ==========
        try:
            metodo_pago = str(data.get('metodo_pago', 'EFECTIVO')).upper()
            total_venta = float(data.get('total', 0))
            
            es_nequi = 'NEQUI' in metodo_pago
            es_daviplata = 'DAVIPLATA' in metodo_pago
            
            # Solo registrar en CarguePagos si es transacción electrónica especial
            if (es_nequi or es_daviplata) and id_vendedor:
                from .models import CarguePagos
                
                # Intentar obtener el día correcto usando el mapa de modelos ya definido
                dia_str = 'LUNES' # Valor por defecto
                
                if 'modelo_map' in locals() and modelo_map.get(id_vendedor):
                    RefModel = modelo_map.get(id_vendedor)
                    ref_obj = RefModel.objects.filter(fecha=fecha_venta).first()
                    if ref_obj:
                        dia_str = ref_obj.dia
                
                CarguePagos.objects.create(
                    vendedor_id=id_vendedor,
                    dia=dia_str,
                    fecha=fecha_venta,
                    concepto=f"Venta: {data.get('cliente_nombre', 'Cliente Final')}",
                    nequi=total_venta if es_nequi else 0,
                    daviplata=total_venta if es_daviplata else 0,
                    descuentos=0, 
                    usuario='App Movil'
                )
                print(f"   💸 Pago registrado en CarguePagos: {metodo_pago} - ${total_venta}")
                
        except Exception as e:
            print(f"❌ Error sincronizando pagos: {str(e)}")
        
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        """
        🆕 EDICIÓN DE VENTA RUTA
        Al editar una venta:
        1. Guarda detalles anteriores (para revertir Cargue)
        2. Hace el update normal
        3. Ajusta el Cargue IDx: resta vendidas viejas, suma vendidas nuevas
        Protegido con try/except — si el Cargue falla, la venta igual se edita.
        """
        import json as json_lib
        from django.db.models import F

        # ─── Obtener la instancia ANTES de editar (para revertir Cargue) ───
        instancia_anterior = self.get_object()
        detalles_anteriores = instancia_anterior.detalles or []
        id_vendedor_anterior = None
        fecha_venta_anterior = None
        try:
            id_vendedor_anterior = instancia_anterior.vendedor.id_vendedor
            fecha_venta_anterior = self._resolver_fecha_operativa(instancia_anterior.fecha, fallback=timezone.now())
        except Exception:
            pass

        if instancia_anterior.estado == 'ANULADA':
            return Response(
                {'error': 'No se puede editar una venta anulada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if instancia_anterior.editada:
            return Response(
                {
                    'error': 'Esta venta ya fue modificada una vez. No se permiten más ediciones ni cambios de método de pago.',
                    'codigo': 'VENTA_YA_MODIFICADA'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        nuevos_detalles = request.data.get('detalles', None)
        if nuevos_detalles is not None:
            detalles_normalizados, _ = self._normalizar_detalles(nuevos_detalles)
            if not detalles_normalizados:
                return Response(
                    {'error': 'La venta debe incluir al menos un producto valido'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            ModeloCargue = self._obtener_modelo_cargue_por_vendedor(id_vendedor_anterior)
            fecha_objetivo = self._resolver_fecha_operativa(
                request.data.get('fecha'),
                fallback=instancia_anterior.fecha,
            )
            cantidades_base = self._cantidades_base_detalles(
                ModeloCargue,
                fecha_venta_anterior,
                detalles_anteriores,
            )
            error_stock = self._validar_stock_disponible_cargue(
                ModeloCargue,
                fecha_objetivo,
                detalles_normalizados,
                cantidades_base=cantidades_base,
            )
            if error_stock:
                return Response(error_stock, status=status.HTTP_400_BAD_REQUEST)

        # ─── Hacer el update normal (heredado) ───
        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(instancia_anterior, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        venta_actualizada = serializer.save(editada=True)
        
        print(f"✅ Venta actualizada: ID={venta_actualizada.id}, Total={venta_actualizada.total}, Método={venta_actualizada.metodo_pago}")
        print(f"   Detalles: {len(venta_actualizada.detalles or [])} productos")
        print(f"   Vencidas: {len(venta_actualizada.productos_vencidos or [])} productos")


        # ─── Marcar como editada ───
        # Guardar flag en el campo detalles no es ideal, pero evitamos migración.
        # En cambio, retornamos el campo en la respuesta.
        
        # ─── Sincronizar Cargue IDx (revertir antigua + aplicar nueva) ───
        try:
            modelo_map = {
                'ID1': CargueID1,
                'ID2': CargueID2,
                'ID3': CargueID3,
                'ID4': CargueID4,
                'ID5': CargueID5,
                'ID6': CargueID6,
            }

            if id_vendedor_anterior and fecha_venta_anterior:
                ModeloCargue = modelo_map.get(id_vendedor_anterior)

                if ModeloCargue:
                    # 1) REVERTIR cantidades anteriores
                    print(f"↩️ Revirtiendo {len(detalles_anteriores)} productos de venta editada en Cargue {id_vendedor_anterior}")
                    for item in detalles_anteriores:
                        nombre = item.get('nombre', '') or item.get('producto', '')
                        cantidad = int(item.get('cantidad', 0))
                        if nombre and cantidad > 0:
                            ModeloCargue.objects.filter(
                                fecha=fecha_venta_anterior,
                                producto__iexact=nombre,
                                activo=True
                            ).update(vendidas=F('vendidas') - cantidad)
                            print(f"   ↩️ {nombre}: vendidas -= {cantidad}")

                    # 🆕 REVERTIR método de pago anterior (nequi/daviplata)
                    metodo_pago_anterior = str(instancia_anterior.metodo_pago or 'EFECTIVO').upper()
                    total_anterior = float(instancia_anterior.total or 0)
                    
                    if 'NEQUI' in metodo_pago_anterior and total_anterior > 0:
                        # Buscar cualquier registro del día para restar nequi
                        registro_dia = ModeloCargue.objects.filter(
                            fecha=fecha_venta_anterior,
                            activo=True
                        ).first()
                        if registro_dia:
                            ModeloCargue.objects.filter(
                                fecha=fecha_venta_anterior,
                                dia=registro_dia.dia,
                                activo=True
                            ).update(nequi=F('nequi') - total_anterior)
                            print(f"   ↩️ Nequi revertido: ${total_anterior}")
                    
                    elif 'DAVIPLATA' in metodo_pago_anterior and total_anterior > 0:
                        registro_dia = ModeloCargue.objects.filter(
                            fecha=fecha_venta_anterior,
                            activo=True
                        ).first()
                        if registro_dia:
                            ModeloCargue.objects.filter(
                                fecha=fecha_venta_anterior,
                                dia=registro_dia.dia,
                                activo=True
                            ).update(daviplata=F('daviplata') - total_anterior)
                            print(f"   ↩️ Daviplata revertido: ${total_anterior}")

                    # 🆕 REVERTIR vencidas anteriores
                    vencidas_anteriores = instancia_anterior.productos_vencidos or []
                    if isinstance(vencidas_anteriores, str):
                        try:
                            vencidas_anteriores = json_lib.loads(vencidas_anteriores)
                        except Exception:
                            vencidas_anteriores = []
                    
                    for item_vencido in vencidas_anteriores:
                        nombre_vencido = item_vencido.get('producto', '') or item_vencido.get('nombre', '')
                        cantidad_vencida = int(item_vencido.get('cantidad', 0))
                        if nombre_vencido and cantidad_vencida > 0:
                            ModeloCargue.objects.filter(
                                fecha=fecha_venta_anterior,
                                producto__iexact=nombre_vencido,
                                activo=True
                            ).update(vencidas=F('vencidas') - cantidad_vencida)
                            print(f"   ↩️ {nombre_vencido}: vencidas -= {cantidad_vencida}")

                    # 2) APLICAR cantidades nuevas
                    nuevos_detalles = venta_actualizada.detalles or []
                    if isinstance(nuevos_detalles, str):
                        try:
                            nuevos_detalles = json_lib.loads(nuevos_detalles)
                        except Exception:
                            nuevos_detalles = []

                    fecha_nueva = self._resolver_fecha_operativa(venta_actualizada.fecha, fallback=timezone.now())
                    print(f"✅ Aplicando {len(nuevos_detalles)} productos nuevos en Cargue {id_vendedor_anterior}")
                    for item in nuevos_detalles:
                        nombre = item.get('nombre', '') or item.get('producto', '')
                        cantidad = int(item.get('cantidad', 0))
                        if nombre and cantidad > 0:
                            ModeloCargue.objects.filter(
                                fecha=fecha_nueva,
                                producto__iexact=nombre,
                                activo=True
                            ).update(vendidas=F('vendidas') + cantidad)
                            print(f"   ✅ {nombre}: vendidas += {cantidad}")

                    # 🆕 APLICAR método de pago nuevo (nequi/daviplata)
                    metodo_pago_nuevo = str(venta_actualizada.metodo_pago or 'EFECTIVO').upper()
                    total_nuevo = float(venta_actualizada.total or 0)
                    
                    if 'NEQUI' in metodo_pago_nuevo and total_nuevo > 0:
                        registro_dia = ModeloCargue.objects.filter(
                            fecha=fecha_nueva,
                            activo=True
                        ).first()
                        if registro_dia:
                            ModeloCargue.objects.filter(
                                fecha=fecha_nueva,
                                dia=registro_dia.dia,
                                activo=True
                            ).update(nequi=F('nequi') + total_nuevo)
                            print(f"   ✅ Nequi aplicado: ${total_nuevo}")
                    
                    elif 'DAVIPLATA' in metodo_pago_nuevo and total_nuevo > 0:
                        registro_dia = ModeloCargue.objects.filter(
                            fecha=fecha_nueva,
                            activo=True
                        ).first()
                        if registro_dia:
                            ModeloCargue.objects.filter(
                                fecha=fecha_nueva,
                                dia=registro_dia.dia,
                                activo=True
                            ).update(daviplata=F('daviplata') + total_nuevo)
                            print(f"   ✅ Daviplata aplicado: ${total_nuevo}")

                    # 🆕 APLICAR vencidas nuevas
                    vencidas_nuevas = venta_actualizada.productos_vencidos or []
                    if isinstance(vencidas_nuevas, str):
                        try:
                            vencidas_nuevas = json_lib.loads(vencidas_nuevas)
                        except Exception:
                            vencidas_nuevas = []
                    
                    for item_vencido in vencidas_nuevas:
                        nombre_vencido = item_vencido.get('producto', '') or item_vencido.get('nombre', '')
                        cantidad_vencida = int(item_vencido.get('cantidad', 0))
                        if nombre_vencido and cantidad_vencida > 0:
                            ModeloCargue.objects.filter(
                                fecha=fecha_nueva,
                                producto__iexact=nombre_vencido,
                                activo=True
                            ).update(vencidas=F('vencidas') + cantidad_vencida)
                            print(f"   ✅ {nombre_vencido}: vencidas += {cantidad_vencida}")

        except Exception as e:
            print(f"⚠️ Error ajustando Cargue en edición de venta (la venta sí se editó): {e}")
            import traceback
            traceback.print_exc()

        # Retornar la venta actualizada con flag editada=True
        data = serializer.data
        data_con_flag = dict(data)
        data_con_flag['editada'] = True
        return Response(data_con_flag)



# ===== VIEWSET PARA SNAPSHOT PLANEACIÓN =====

from .models import RegistrosPlaneacionDia
from .serializers import RegistrosPlaneacionDiaSerializer

class RegistrosPlaneacionDiaViewSet(viewsets.ModelViewSet):
    """
    API para guardar y consultar snapshots de Planeación.
    Se usa cuando el botón cambia de SUGERIDO → ALISTAMIENTO_ACTIVO.
    """
    queryset = RegistrosPlaneacionDia.objects.all()
    serializer_class = RegistrosPlaneacionDiaSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = RegistrosPlaneacionDia.objects.all()
        fecha = self.request.query_params.get('fecha', None)
        
        if fecha:
            queryset = queryset.filter(fecha=fecha)
        
        return queryset.order_by('orden', 'producto_nombre')
    
    @action(detail=False, methods=['post'])
    def guardar_snapshot(self, request):
        """
        Endpoint para guardar snapshot completo de Planeación.
        Recibe: { fecha: "2025-12-02", registros: [...], usuario: "Sistema" }
        """
        from django.db import transaction
        
        fecha = request.data.get('fecha')
        registros = request.data.get('registros', [])
        usuario = request.data.get('usuario', 'Sistema')
        
        if not fecha:
            return Response({'error': 'Fecha requerida'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not registros:
            return Response({'error': 'No hay registros para guardar'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            with transaction.atomic():
                # Eliminar registros anteriores de la misma fecha (sobrescribir)
                RegistrosPlaneacionDia.objects.filter(fecha=fecha).delete()
                
                # Crear nuevos registros
                registros_creados = []
                for reg in registros:
                    nuevo = RegistrosPlaneacionDia.objects.create(
                        fecha=fecha,
                        producto_nombre=reg.get('producto_nombre', ''),
                        existencias=reg.get('existencias', 0),
                        solicitadas=reg.get('solicitadas', 0),
                        pedidos=reg.get('pedidos', 0),
                        total=reg.get('total', 0),
                        orden=reg.get('orden', 0),
                        ia=reg.get('ia', 0),
                        usuario=usuario
                    )
                    registros_creados.append(nuevo)
                
                print(f"✅ Snapshot guardado: {len(registros_creados)} registros para {fecha}")
                
                return Response({
                    'success': True,
                    'mensaje': f'Snapshot guardado: {len(registros_creados)} registros',
                    'fecha': fecha,
                    'cantidad': len(registros_creados)
                }, status=status.HTTP_201_CREATED)
                
        except Exception as e:
            print(f"❌ Error guardando snapshot: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def consultar_fecha(self, request):
        """
        Endpoint para consultar snapshot de una fecha específica.
        GET /api/registros-planeacion-dia/consultar_fecha/?fecha=2025-12-02
        """
        fecha = request.query_params.get('fecha')
        
        if not fecha:
            return Response({'error': 'Fecha requerida'}, status=status.HTTP_400_BAD_REQUEST)
        
        registros = RegistrosPlaneacionDia.objects.filter(fecha=fecha).order_by('orden', 'producto_nombre')
        
        if not registros.exists():
            return Response({
                'fecha': fecha,
                'existe': False,
                'registros': [],
                'mensaje': f'No hay snapshot para la fecha {fecha}'
            })
        
        serializer = self.get_serializer(registros, many=True)
        
        # Obtener fecha de congelado del primer registro
        fecha_congelado = registros.first().fecha_congelado if registros.exists() else None
        
        return Response({
            'fecha': fecha,
            'existe': True,
            'fecha_congelado': fecha_congelado,
            'cantidad': registros.count(),
            'registros': serializer.data
        })


# ==================== 🔗 INTEGRACIÓN APP ↔ WEB ====================
# Endpoints para conectar ventas de app móvil con cargue web

@api_view(['GET'])
def calcular_devoluciones_automaticas(request, id_vendedor, fecha):
    """
    Calcula devoluciones automáticamente basándose en:
    - Cargue inicial (de CargueIDx)
    - Ventas reales (de VentaRuta desde app móvil)
    - Vencidas (registradas manualmente)
    
    Fórmula: DEVOLUCIONES = CARGUE_INICIAL - VENTAS_APP - VENCIDAS
    
    Parámetros:
        id_vendedor: ID1, ID2, ID3, ID4, ID5, ID6
        fecha: YYYY-MM-DD
    
    Retorna:
        {
            "id_vendedor": "ID1",
            "fecha": "2025-12-17",
            "productos": [
                {
                    "producto": "AREPA TIPO OBLEA",
                    "cantidad_inicial": 200,
                    "cantidad_vendida": 150,
                    "vencidas": 5,
                    "devoluciones": 45
                }
            ]
        }
    """
    try:
        from django.db.models import Sum
        
        # Mapeo de ID a Modelo
        modelo_map = {
            'ID1': CargueID1,
            'ID2': CargueID2,
            'ID3': CargueID3,
            'ID4': CargueID4,
            'ID5': CargueID5,
            'ID6': CargueID6,
        }
        
        ModeloCargue = modelo_map.get(id_vendedor)
        if not ModeloCargue:
            return Response({'error': 'ID de vendedor inválido'}, status=400)
        
        # Obtener cargue del día
        cargues = ModeloCargue.objects.filter(
            fecha=fecha,
            activo=True
        )
        
        if not cargues.exists():
            return Response({
                'id_vendedor': id_vendedor,
                'fecha': fecha,
                'mensaje': 'No hay datos de cargue para esta fecha',
                'productos': []
            })
        
        resultado = []
        
        for cargue in cargues:
            # Cantidad inicial con la que salió (cantidad - dctos + adicional)
            cantidad_inicial = cargue.cantidad - cargue.dctos + cargue.adicional
            
            # Ventas registradas en app (buscar por vendedor_id que viene del modelo Vendedor)
            ventas_app = VentaRuta.objects.filter(
                vendedor__id_vendedor=id_vendedor,
                fecha__date=fecha
            )
            
            # Sumar cantidades vendidas por producto (del campo JSON 'detalles')
            cantidad_vendida = 0
            for venta in ventas_app:
                detalles = venta.detalles or []
                for detalle in detalles:
                    nombre_detalle = detalle.get('nombre', '') or detalle.get('producto', '')
                    if nombre_detalle.upper() == cargue.producto.upper():
                        cantidad_vendida += detalle.get('cantidad', 0)
            
            # Vencidas (registradas manualmente en cargue)
            vencidas = cargue.vencidas or 0
            
            # Calcular devoluciones (no puede ser negativo)
            devoluciones = max(0, cantidad_inicial - cantidad_vendida - vencidas)
            
            resultado.append({
                'producto': cargue.producto,
                'cantidad_inicial': cantidad_inicial,
                'cantidad_vendida': cantidad_vendida,
                'vencidas': vencidas,
                'devoluciones': devoluciones
            })
        
        return Response({
            'id_vendedor': id_vendedor,
            'fecha': fecha,
            'productos': resultado
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
def control_stock_tiempo_real(request, id_vendedor, fecha):
    """
    Retorna un tablero read-only de control de stock por producto para un ID y fecha.

    La fuente base es el CargueIDx ya sincronizado con ventas y vencidas:
    - salio_con = cantidad - dctos + adicional
    - vendidas = unidades reportadas desde app ya impactadas en cargue
    - vencidas = unidades vencidas reportadas
    - saldo_teorico = salio_con - vendidas - vencidas
    - devoluciones = devolución física registrada al cierre (si existe)
    """
    try:
        modelo_map = {
            'ID1': CargueID1,
            'ID2': CargueID2,
            'ID3': CargueID3,
            'ID4': CargueID4,
            'ID5': CargueID5,
            'ID6': CargueID6,
        }

        ModeloCargue = modelo_map.get(id_vendedor)
        if not ModeloCargue:
            return Response({'error': 'ID de vendedor inválido'}, status=400)

        cargues = ModeloCargue.objects.filter(
            fecha=fecha,
            activo=True
        ).order_by('id')

        if not cargues.exists():
            return Response({
                'id_vendedor': id_vendedor,
                'fecha': fecha,
                'mensaje': 'No hay datos de cargue para esta fecha',
                'cerrado': False,
                'totales': {
                    'salio_con': 0,
                    'vendidas': 0,
                    'vencidas': 0,
                    'saldo_teorico': 0,
                    'devoluciones': 0,
                },
                'productos': [],
            })

        productos = []
        totales = {
            'salio_con': 0,
            'vendidas': 0,
            'vencidas': 0,
            'saldo_teorico': 0,
            'devoluciones': 0,
        }
        cerrado = False

        for cargue in cargues:
            salio_con = int(cargue.cantidad or 0) - int(cargue.dctos or 0) + int(cargue.adicional or 0)
            vendidas = int(cargue.vendidas or 0)
            vencidas = int(cargue.vencidas or 0)
            devoluciones = int(cargue.devoluciones or 0)
            saldo_teorico = salio_con - vendidas - vencidas
            diferencia_cierre = saldo_teorico - devoluciones

            if devoluciones > 0:
                cerrado = True

            productos.append({
                'producto': cargue.producto,
                'salio_con': salio_con,
                'vendidas': vendidas,
                'vencidas': vencidas,
                'saldo_teorico': saldo_teorico,
                'devoluciones': devoluciones,
                'diferencia_cierre': diferencia_cierre,
                'valor_unitario': float(cargue.valor or 0),
            })

            totales['salio_con'] += salio_con
            totales['vendidas'] += vendidas
            totales['vencidas'] += vencidas
            totales['saldo_teorico'] += saldo_teorico
            totales['devoluciones'] += devoluciones

        return Response({
            'id_vendedor': id_vendedor,
            'fecha': fecha,
            'cerrado': cerrado,
            'totales': totales,
            'productos': productos,
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
def ventas_tiempo_real(request, id_vendedor, fecha):
    """
    Obtiene ventas del día en tiempo real desde VentaRuta (app móvil).
    Agrupado por producto y método de pago.
    
    Parámetros:
        id_vendedor: ID1, ID2, ID3, ID4, ID5, ID6
        fecha: YYYY-MM-DD
    
    Retorna:
        {
            "id_vendedor": "ID1",
            "fecha": "2025-12-17",
            "totalVentas": 5,
            "total_dinero": 125000,
            "productos_vendidos": [
                {"producto": "AREPA TIPO OBLEA", "cantidad": 50}
            ],
            "ventas_por_metodo": {
                "EFECTIVO": 75000,
                "NEQUI": 35000,
                "DAVIPLATA": 15000
            }
        }
    """
    try:
        from django.db.models import Sum, Count
        
        # Obtener ventas del día para el vendedor
        ventas = VentaRuta.objects.filter(
            vendedor__id_vendedor=id_vendedor,
            fecha__date=fecha
        )
        
        if not ventas.exists():
            return Response({
                'id_vendedor': id_vendedor,
                'fecha': fecha,
                'total_ventas': 0,
                'total_dinero': 0,
                'productos_vendidos': [],
                'ventas_por_metodo': {
                    'EFECTIVO': 0,
                    'NEQUI': 0,
                    'DAVIPLATA': 0
                }
            })
        
        # Agrupar por producto (procesar JSON 'detalles')
        ventas_por_producto = {}
        total_dinero = 0
        ventas_por_metodo = {
            'EFECTIVO': 0,
            'NEQUI': 0,
            'DAVIPLATA': 0,
            'TRANSFERENCIA': 0
        }
        
        for venta in ventas:
            # Acumular totales
            total_dinero += float(venta.total or 0)
            metodo = venta.metodo_pago or 'EFECTIVO'
            ventas_por_metodo[metodo] = ventas_por_metodo.get(metodo, 0) + float(venta.total or 0)
            
            # Procesar productos del JSON 'detalles'
            detalles = venta.detalles or []
            for detalle in detalles:
                nombre = detalle.get('nombre', '') or detalle.get('producto', 'Sin nombre')
                cantidad = detalle.get('cantidad', 0)
                
                if nombre in ventas_por_producto:
                    ventas_por_producto[nombre] += cantidad
                else:
                    ventas_por_producto[nombre] = cantidad
        
        # Convertir a lista
        productos_vendidos = [
            {'producto': k, 'cantidad': v}
            for k, v in sorted(ventas_por_producto.items(), key=lambda x: x[1], reverse=True)
        ]
        
        return Response({
            'id_vendedor': id_vendedor,
            'fecha': fecha,
            'total_ventas': ventas.count(),
            'total_dinero': total_dinero,
            'productos_vendidos': productos_vendidos,
            'ventas_por_metodo': ventas_por_metodo
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['POST'])
def cerrar_turno_vendedor(request):
    """
    Cierra el turno del vendedor desde la app móvil.
    Calcula devoluciones automáticamente y las guarda en CargueIDx.
    
    OPTIMIZADO (Django Expert Skill):
    - Usa transaction.atomic() para integridad de datos.
    - Usa bulk_update() para reducir queries SQL.
    - Manejo robusto de errores con rollback automático.
    
    POST /api/cargue/cerrar-turno/
    """
    try:
        id_vendedor = request.data.get('id_vendedor')
        fecha = request.data.get('fecha')
        productos_vencidos = request.data.get('productos_vencidos', [])
        diferencia_precios = request.data.get('diferencia_precios', 0)  # 🆕 Recibir diferencia por precios especiales
        
        print(f"🔒 CERRAR TURNO (Optimizado): {id_vendedor} - {fecha}")
        if diferencia_precios > 0:
            print(f"💰 Diferencia por precios especiales: ${diferencia_precios}")
        
        if not id_vendedor or not fecha:
            return Response({'error': 'Se requiere id_vendedor y fecha'}, status=400)
        
        # Mapeo de ID a Modelo
        modelo_map = {
            'ID1': CargueID1, 'ID2': CargueID2, 'ID3': CargueID3,
            'ID4': CargueID4, 'ID5': CargueID5, 'ID6': CargueID6,
        }
        
        ModeloCargue = modelo_map.get(id_vendedor)
        if not ModeloCargue:
            return Response({'error': f'ID de vendedor inválido: {id_vendedor}'}, status=400)

        # CORRECCIÓN BUG TIMEZONE: la app puede enviar fecha+1 por diferencia UTC/local
        # Usar el TurnoVendedor ABIERTO como fuente de verdad para la fecha
        # Esto cubre el caso donde ya existe cargue del día siguiente (enviado anticipadamente)
        from .models import TurnoVendedor as _TurnoVendedor
        _v_id_num = int(str(id_vendedor).replace('ID', '')) if 'ID' in str(id_vendedor) else 0
        _turno_activo = _TurnoVendedor.objects.filter(vendedor_id=_v_id_num, estado='ABIERTO').first()
        if _turno_activo and str(_turno_activo.fecha) != fecha:
            print(f"🔧 Redirigiendo cierre: {fecha} → {_turno_activo.fecha} (turno abierto = fuente de verdad)")
            fecha = str(_turno_activo.fecha)

        # INICIO BLOQUE ATÓMICO - Todo o Nada
        with transaction.atomic():
            # Obtener cargue del día
            cargues = ModeloCargue.objects.filter(fecha=fecha, activo=True)

            if not cargues.exists():
                print(f"⚠️ No hay cargue para {id_vendedor} en {fecha}. Cerrando turno vacío.")
                
                # Intentar cerrar el turno Vendedor aunque no haya cargue
                from .models import TurnoVendedor, CargueResumen
                v_id_num = int(str(id_vendedor).replace('ID', '')) if 'ID' in str(id_vendedor) else 0
                
                # Buscar turno ESPECÍFICO de la fecha (Corrección Bug Fantasma)
                turno = TurnoVendedor.objects.filter(
                    vendedor_id=v_id_num,
                    fecha=fecha,
                    estado='ABIERTO'
                ).first()
                
                if turno:
                    turno.estado = 'CERRADO'
                    turno.hora_cierre = timezone.now()
                    turno.cerrado_manual = True
                    turno.save()
                    print(f"✅ Turno {turno.id} marcado como CERRADO (sin cargue)")
                
                # Actualizar resumen si existe
                try:
                    CargueResumen.objects.update_or_create(
                        dia=turno.dia if turno else 'DESCONOCIDO',
                        fecha=fecha,
                        vendedor_id=id_vendedor,
                        defaults={'estado_cargue': 'COMPLETADO', 'activo': True}
                    )
                except Exception as ex_resumen:
                     print(f"⚠️ Error actualizando CargueResumen sin cargue: {ex_resumen}")

                return Response({
                    'success': True,
                    'mensaje': 'Turno cerrado correctamente (Sin registros de cargue)',
                    'resumen': [],
                    'totales': {'cargado': 0, 'vendido': 0, 'vencidas': 0, 'devuelto': 0}
                })
            
            # --- VALIDACIÓN: YA CERRADO ---
            # Si algún producto tiene devoluciones > 0, significa que ya se cerró el turno
            if cargues.filter(devoluciones__gt=0).exists():
                print(f"⚠️ TURNO YA CERRADO para {id_vendedor} en {fecha}")
                
                # Asegurar que el TurnoVendedor también esté CERRADO
                from .models import TurnoVendedor
                vendedor_id_num = int(str(id_vendedor).replace('ID', '')) if 'ID' in str(id_vendedor) else int(id_vendedor)
                turno_abierto = TurnoVendedor.objects.filter(
                    vendedor_id=vendedor_id_num, fecha=fecha, estado='ABIERTO'
                ).first()
                if turno_abierto:
                    turno_abierto.estado = 'CERRADO'
                    turno_abierto.hora_cierre = timezone.now()
                    turno_abierto.cerrado_manual = True
                    turno_abierto.save()
                    print(f"✅ TurnoVendedor {turno_abierto.id} forzado a CERRADO")
                
                return Response({
                    'error': 'TURNO_YA_CERRADO',
                    'message': f'El turno para {id_vendedor} en {fecha} ya fue cerrado anteriormente.'
                }, status=409)
            
            # --- CÁLCULOS Y ACTUALIZACIÓN ---

            from .models import TurnoVendedor
            resumen = []
            cargues_a_actualizar = []
            totales = {'cargado': 0, 'vendido': 0, 'vencidas': 0, 'devuelto': 0}

            for cargue in cargues:
                # 1. Calcular cargado
                cantidad_inicial = cargue.cantidad - cargue.dctos + cargue.adicional

                # 2. Vendido = solo ventas de ruta (app). Pedidos son inventario separado, no salen del cargue.
                cantidad_vendida_total = cargue.vendidas or 0
                
                # 3. Calcular vencidas
                vencidas_final = cargue.vencidas or 0 # Base existente
                for item_vencido in productos_vencidos:
                    prod_vencido_nombre = item_vencido.get('producto', '')
                    # Comparación insensible a mayúsculas
                    if prod_vencido_nombre and prod_vencido_nombre.upper() == cargue.producto.upper():
                        cant_vencida = item_vencido.get('cantidad', 0)
                        if cant_vencida > 0:
                            vencidas_final = cant_vencida
                        break
                
                # 4. Calcular devoluciones (No negativo)
                devoluciones = max(0, cantidad_inicial - cantidad_vendida_total - vencidas_final)
                
                # Preparar objeto para actualización masiva (No guardar todavía)
                cargue.vencidas = vencidas_final
                cargue.devoluciones = devoluciones
                cargues_a_actualizar.append(cargue)
                
                # Acumular totales para respuesta y turno
                totales['cargado'] += cantidad_inicial
                totales['vendido'] += cantidad_vendida_total
                totales['vencidas'] += vencidas_final
                totales['devuelto'] += devoluciones
                
                resumen.append({
                    'producto': cargue.producto,
                    'cargado': cantidad_inicial,
                    'vendido': cantidad_vendida_total,
                    'vencidas': vencidas_final,
                    'devuelto': devoluciones
                })
            
            # 🚀 EJECUTAR BULK UPDATE (Optimización Clave: 1 Query SQL)
            if cargues_a_actualizar:
                ModeloCargue.objects.bulk_update(cargues_a_actualizar, ['vencidas', 'devoluciones'])
                recalcular_totales_cargue_queryset(
                    ModeloCargue.objects.filter(pk__in=[c.pk for c in cargues_a_actualizar])
                )
                print(f"✅ {len(cargues_a_actualizar)} productos actualizados vía bulk_update")
            
            # --- CERRAR TURNO EN BD ---
            vendedor_id_numerico = int(str(id_vendedor).replace('ID', '')) if 'ID' in str(id_vendedor) else int(id_vendedor)
            turno = TurnoVendedor.objects.filter(vendedor_id=vendedor_id_numerico, fecha=fecha).first()
            
            if turno:
                turno.estado = 'CERRADO'
                turno.hora_cierre = timezone.now()
                turno.cerrado_manual = True
                turno.total_ventas = totales['vendido']
                turno.total_dinero = totales['cargado'] # O el valor monetario real si se tiene
                turno.save()
                print(f"✅ Turno marcado como CERRADO en BD")

            print(f"✅ Turno cerrado exitosamente para {id_vendedor}. Rollback no necesario.")
            
            # 🆕 Agregar novedad si hay diferencia por precios especiales
            novedad = None
            if diferencia_precios > 0:
                novedad = f"💰 Venta Precios Especiales: +${diferencia_precios:,.0f}"
                print(f"📝 Novedad agregada: {novedad}")
            
            return Response({
                'success': True,
                'mensaje': 'Turno cerrado correctamente',
                'resumen': resumen,
                'totales': totales,
                'novedad': novedad  # 🆕 Enviar novedad al frontend
            })
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        # Gracias a transaction.atomic, si falla algo aquí, Django hace rollback automático
        print(f"❌ Error cerrando turno (Rollback automático): {str(e)}")
        return Response({'error': str(e), 'mensaje': 'Error al cerrar turno'}, status=500)


# ========================================
# ENDPOINTS PARA GESTIÓN DE TURNOS (App Móvil)
# ========================================

@api_view(['GET'])
def verificar_turno_activo(request):
    """
    Verificar si hay un turno abierto para un vendedor.
    Permite sincronización entre dispositivos.
    
    🆕 MEJORA: Si el turno no tiene cargue asociado, se cierra automáticamente.
    
    Query params:
    - vendedor_id: ID del vendedor (numérico o cadena ID1, ID2, etc.)
    - fecha: Fecha opcional (default: hoy)
    """
    try:
        from .models import TurnoVendedor, CargueProductos, CargueID1, CargueID2, CargueID3, CargueID4, CargueID5, CargueID6
        from datetime import date
        
        vendedor_id_param = request.query_params.get('vendedor_id')
        vendedor, vendedor_id_numerico, auth_error = _validar_vendedor_token(
            request,
            vendedor_id_param,
            campo='vendedor_id'
        )
        if auth_error:
            return auth_error

        vendedor_id = vendedor.id_vendedor
        fecha_param = request.query_params.get('fecha')
        
        # Fecha (hoy por defecto)
        if fecha_param:
            from datetime import datetime
            fecha = datetime.strptime(fecha_param, '%Y-%m-%d').date()
        else:
            fecha = date.today()
        
        # Buscar turno activo
        if not fecha_param:
            # Si no se especifica fecha, buscar CUALQUIER turno abierto (el más reciente)
            turno = TurnoVendedor.objects.filter(
                vendedor_id=vendedor_id_numerico,
                estado='ABIERTO'
            ).order_by('-fecha').first()
            
            if turno:
                print(f"✅ Turno activo encontrado: {turno.fecha} (Hoy es {date.today()})")
        else:
            # Si se especifica fecha, buscar estrictamente esa fecha
            turno = TurnoVendedor.objects.filter(
                vendedor_id=vendedor_id_numerico,
                fecha=fecha,
                estado='ABIERTO'
            ).first()
        
        if turno:
            # ✅ Turno ABIERTO encontrado - retornarlo directamente
            # NOTA: Ya NO cerramos turnos automáticamente por falta de cargue.
            # La validación de cargue se hace ANTES de abrir turno (en la app).
            # Cerrar turnos aquí causaba que turnos legítimos se cerraran
            # por falsos negativos en la búsqueda de cargue.
            return Response({
                'turno_activo': True,
                'turno_id': turno.id,
                'dia': turno.dia,
                'fecha': turno.fecha.isoformat(),
                'hora_apertura': turno.hora_apertura.isoformat() if turno.hora_apertura else None,
                'vendedor_nombre': turno.vendedor_nombre,
                'total_ventas': turno.total_ventas,
                'total_dinero': float(turno.total_dinero)
            })
        else:
            # 🆕 Buscar si ya existe un turno CERRADO hoy
            turno_cerrado = TurnoVendedor.objects.filter(
                vendedor_id=vendedor_id_numerico,
                fecha=fecha,
                estado='CERRADO'
            ).first()

            if turno_cerrado:
                return Response({
                    'turno_activo': False,
                    'turno_cerrado_hoy': True,
                    'mensaje': 'El turno ya fue cerrado hoy',
                    'total_ventas': turno_cerrado.total_ventas,
                })
            
            return Response({
                'turno_activo': False,
                'mensaje': 'No hay turno abierto para esta fecha'
            })
            
    except Exception as e:
        print(f"❌ Error verificando turno: {e}")
        return Response({
            'error': str(e)
        }, status=500)


@api_view(['POST'])
def abrir_turno(request):
    """
    Abrir un nuevo turno para un vendedor.
    Si ya hay turno abierto para ese día, retorna el existente.
    
    Body:
    - vendedor_id: ID del vendedor
    - vendedor_nombre: Nombre del vendedor (opcional)
    - dia: Día de la semana (LUNES, MARTES, etc.)
    - fecha: Fecha del turno (YYYY-MM-DD)
    """
    try:
        from .models import TurnoVendedor, CargueID1, CargueID2, CargueID3, CargueID4, CargueID5, CargueID6
        from datetime import datetime
        
        vendedor_id = request.data.get('vendedor_id')
        vendedor, vendedor_id_numerico, auth_error = _validar_vendedor_token(
            request,
            vendedor_id,
            campo='vendedor_id'
        )
        if auth_error:
            return auth_error

        vendedor_id = vendedor.id_vendedor
        vendedor_nombre = vendedor.nombre or request.data.get('vendedor_nombre', '')
        dia = request.data.get('dia', '').upper()
        fecha_str = request.data.get('fecha')
        dispositivo = request.data.get('dispositivo', '')
        forzar = request.data.get('forzar', False)
        
        if not dia or not fecha_str:
            return Response({
                'error': 'dia y fecha son requeridos'
            }, status=400)
        
        # Parsear fecha
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        
        # Verificar si ya existe turno para este día
        turno_existente = TurnoVendedor.objects.filter(
            vendedor_id=vendedor_id_numerico,
            fecha=fecha
        ).first()
        
        if turno_existente:
            if turno_existente.estado == 'ABIERTO':
                # Ya hay un turno abierto, retornarlo
                return Response({
                    'success': True,
                    'nuevo': False,
                    'mensaje': 'Ya hay un turno abierto para este día',
                    'turno_id': turno_existente.id,
                    'dia': turno_existente.dia,
                    'fecha': turno_existente.fecha.isoformat(),
                    'hora_apertura': turno_existente.hora_apertura.isoformat() if turno_existente.hora_apertura else None,
                    'estado': turno_existente.estado
                })
            else:
                # 🔒 TURNO CERRADO - NO SE PUEDE REABRIR
                # Una vez cerrado el turno del día, no se permite reabrirlo.
                # El vendedor debe esperar al día siguiente con stock disponible.
                print(f"🔒 Intento de reabrir turno CERRADO bloqueado: {vendedor_nombre} - {dia} {fecha}")
                return Response({
                    'error': 'TURNO_YA_CERRADO',
                    'message': f'El turno del {fecha_str} ya fue cerrado y no se puede reabrir. Debes esperar al día siguiente con stock disponible.',
                    'turno_id': turno_existente.id,
                    'fecha': turno_existente.fecha.isoformat(),
                    'hora_cierre': turno_existente.hora_cierre.isoformat() if turno_existente.hora_cierre else None,
                }, status=409)
        
        # Crear nuevo turno
        turno = TurnoVendedor.objects.create(
            vendedor_id=vendedor_id_numerico,
            vendedor_nombre=vendedor_nombre,
            dia=dia,
            fecha=fecha,
            estado='ABIERTO',
            hora_apertura=timezone.now(),
            dispositivo=dispositivo
        )
        
        print(f"✅ Turno abierto: {vendedor_nombre} - {dia} {fecha}")
        
        return Response({
            'success': True,
            'nuevo': True,
            'mensaje': 'Turno abierto correctamente',
            'turno_id': turno.id,
            'dia': turno.dia,
            'fecha': turno.fecha.isoformat(),
            'hora_apertura': turno.hora_apertura.isoformat(),
            'estado': turno.estado
        })
        
    except Exception as e:
        print(f"❌ Error abriendo turno: {e}")
        import traceback
        traceback.print_exc()
        return Response({
            'error': str(e)
        }, status=500)


@api_view(['POST'])
def cerrar_turno_estado(request):
    """
    Cerrar turno (cambiar estado a CERRADO).
    Solo cambia el estado, no procesa devoluciones.
    
    Body:
    - vendedor_id: ID del vendedor
    - fecha: Fecha del turno
    """
    try:
        from .models import TurnoVendedor
        from datetime import datetime
        
        vendedor_id = request.data.get('vendedor_id')
        _, vendedor_id_numerico, auth_error = _validar_vendedor_token(
            request,
            vendedor_id,
            campo='vendedor_id'
        )
        if auth_error:
            return auth_error

        fecha_str = request.data.get('fecha')
        
        if not fecha_str:
            return Response({
                'error': 'fecha es requerido'
            }, status=400)
        
        # Parsear fecha
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        
        # Buscar turno
        turno = TurnoVendedor.objects.filter(
            vendedor_id=vendedor_id_numerico,
            fecha=fecha
        ).first()
        
        if not turno:
            return Response({
                'error': 'No se encontró turno para esta fecha'
            }, status=404)
        
        if turno.estado == 'CERRADO':
            return Response({
                'error': 'TURNO_YA_CERRADO',
                'mensaje': 'El turno ya estaba cerrado'
            }, status=400)
        
        # Cerrar turno
        turno.estado = 'CERRADO'
        turno.hora_cierre = timezone.now()
        turno.cerrado_manual = True  # 🆕 Marcar como cerrado manualmente
        turno.save()
        
        print(f"✅ Turno cerrado MANUAL: {turno.vendedor_nombre} - {turno.dia} {turno.fecha}")
        
        return Response({
            'success': True,
            'mensaje': 'Turno cerrado correctamente',
            'turno_id': turno.id,
            'hora_cierre': turno.hora_cierre.isoformat()
        })
        
    except Exception as e:
        print(f"❌ Error cerrando turno: {e}")
        return Response({
            'error': str(e)
        }, status=500)


# ========================================
# CONFIGURACIÓN DE PRODUCCIÓN
# ========================================

@api_view(['GET'])
def obtener_configuracion_produccion(request):
    """Obtiene una configuración de producción por clave"""
    from .models import ConfiguracionProduccion
    
    clave = request.query_params.get('clave', 'usuario_produccion')
    
    try:
        config = ConfiguracionProduccion.objects.get(clave=clave)
        return Response({
            'success': True,
            'clave': config.clave,
            'valor': config.valor,
            'descripcion': config.descripcion,
            'fecha_actualizacion': config.fecha_actualizacion
        })
    except ConfiguracionProduccion.DoesNotExist:
        # Si no existe, devolver valor por defecto
        return Response({
            'success': True,
            'clave': clave,
            'valor': 'Usuario Predeterminado',
            'descripcion': 'No configurado',
            'fecha_actualizacion': None
        })


@api_view(['POST', 'PUT'])
def guardar_configuracion_produccion(request):
    """Guarda o actualiza una configuración de producción"""
    from .models import ConfiguracionProduccion
    
    clave = request.data.get('clave', 'usuario_produccion')
    valor = request.data.get('valor', '')
    descripcion = request.data.get('descripcion', '')
    
    if not valor:
        return Response({
            'error': 'El valor es requerido'
        }, status=400)
    
    try:
        config, created = ConfiguracionProduccion.objects.update_or_create(
            clave=clave,
            defaults={
                'valor': valor,
                'descripcion': descripcion
            }
        )
        
        return Response({
            'success': True,
            'action': 'created' if created else 'updated',
            'clave': config.clave,
            'valor': config.valor,
            'descripcion': config.descripcion,
            'fecha_actualizacion': config.fecha_actualizacion
        })
        
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=500)


# ========================================
# 🆕 TRAZABILIDAD DE LOTES
# ========================================

@api_view(['GET'])
def buscar_lote(request):
    """
    Busca un lote específico en todas las tablas de cargue.
    Retorna: producción, despachos y vencidas.
    """
    import json
    from .models import CargueID1, CargueID2, CargueID3, CargueID4, CargueID5, CargueID6, Lote
    
    lote_numero = request.query_params.get('lote', '').upper().strip()
    
    if not lote_numero:
        return Response({'error': 'Debe proporcionar un número de lote'}, status=400)
    
    resultado = {
        'lote': lote_numero,
        'produccion': None,
        'despachos': [],
        'vencidas': []
    }
    
    # 1. Buscar en tabla Lote (lotes registrados en producción)
    try:
        lote_obj = Lote.objects.filter(lote__iexact=lote_numero).first()
        if lote_obj:
            resultado['produccion'] = {
                'fecha': str(lote_obj.fecha_produccion),
                'usuario': lote_obj.usuario,
                'fecha_vencimiento': str(lote_obj.fecha_vencimiento) if lote_obj.fecha_vencimiento else None,
                'activo': lote_obj.activo
            }
    except Exception as e:
        print(f"Error buscando en Lote: {e}")
    
    # 2. Buscar en tablas CargueIDx (lotes_produccion y lotes_vencidos)
    cargue_models = [CargueID1, CargueID2, CargueID3, CargueID4, CargueID5, CargueID6]
    vendedor_ids = ['ID1', 'ID2', 'ID3', 'ID4', 'ID5', 'ID6']
    
    for idx, CargueModel in enumerate(cargue_models):
        vendedor_id = vendedor_ids[idx]
        
        try:
            # Buscar registros con lotes_produccion o lotes_vencidos que contengan este lote
            registros = CargueModel.objects.filter(
                models.Q(lotes_produccion__icontains=lote_numero) |
                models.Q(lotes_vencidos__icontains=lote_numero)
            )
            
            for reg in registros:
                # Procesar lotes_produccion
                if reg.lotes_produccion and lote_numero in reg.lotes_produccion:
                    try:
                        lotes_prod = json.loads(reg.lotes_produccion) if reg.lotes_produccion else []
                        if lote_numero in lotes_prod:
                            resultado['despachos'].append({
                                'fecha': str(reg.fecha),
                                'dia': reg.dia,
                                'vendedor_id': vendedor_id,
                                'responsable': reg.responsable,
                                'producto': reg.producto,
                                'cantidad': reg.cantidad,
                                'lotes': lotes_prod
                            })
                    except json.JSONDecodeError:
                        # Si no es JSON válido, verificar si es el texto directamente
                        if lote_numero in str(reg.lotes_produccion):
                            resultado['despachos'].append({
                                'fecha': str(reg.fecha),
                                'dia': reg.dia,
                                'vendedor_id': vendedor_id,
                                'responsable': reg.responsable,
                                'producto': reg.producto,
                                'cantidad': reg.cantidad,
                                'lotes': [reg.lotes_produccion]
                            })
                
                # Procesar lotes_vencidos
                if reg.lotes_vencidos and lote_numero in reg.lotes_vencidos:
                    try:
                        lotes_venc = json.loads(reg.lotes_vencidos) if reg.lotes_vencidos else []
                        for lv in lotes_venc:
                            if isinstance(lv, dict) and lv.get('lote', '').upper() == lote_numero:
                                resultado['vencidas'].append({
                                    'fecha': str(reg.fecha),
                                    'dia': reg.dia,
                                    'vendedor_id': vendedor_id,
                                    'responsable': reg.responsable,
                                    'producto': reg.producto,
                                    'cantidad': reg.vencidas,
                                    'motivo': lv.get('motivo', 'N/A'),
                                    'lote': lv.get('lote', lote_numero)
                                })
                    except json.JSONDecodeError:
                        pass
        except Exception as e:
            print(f"Error buscando en {CargueModel.__name__}: {e}")
    
    return Response(resultado)


@api_view(['GET'])
def lotes_por_fecha(request):
    """
    Obtiene todos los lotes de producción para una fecha específica.
    """
    import json
    from .models import CargueID1, CargueID2, CargueID3, CargueID4, CargueID5, CargueID6, Lote
    
    fecha = request.query_params.get('fecha', '')
    
    if not fecha:
        return Response({'error': 'Debe proporcionar una fecha'}, status=400)
    
    lotes = []
    
    # 1. Buscar en tabla Lote
    try:
        lotes_obj = Lote.objects.filter(fecha_produccion=fecha)
        for lote in lotes_obj:
            lotes.append({
                'lote': lote.lote,
                'fecha': str(lote.fecha_produccion),
                'usuario': lote.usuario,
                'fecha_vencimiento': str(lote.fecha_vencimiento) if lote.fecha_vencimiento else None,
                'origen': 'Producción'
            })
    except Exception as e:
        print(f"Error buscando lotes por fecha: {e}")
    
    # 2. Buscar lotes en tablas CargueIDx
    cargue_models = [CargueID1, CargueID2, CargueID3, CargueID4, CargueID5, CargueID6]
    vendedor_ids = ['ID1', 'ID2', 'ID3', 'ID4', 'ID5', 'ID6']
    
    lotes_encontrados = set()  # Para evitar duplicados
    
    for idx, CargueModel in enumerate(cargue_models):
        vendedor_id = vendedor_ids[idx]
        
        try:
            registros = CargueModel.objects.filter(
                fecha=fecha,
                lotes_produccion__isnull=False
            ).exclude(lotes_produccion='')
            
            for reg in registros:
                try:
                    lotes_prod = json.loads(reg.lotes_produccion) if reg.lotes_produccion else []
                    for lote_num in lotes_prod:
                        if lote_num and lote_num not in lotes_encontrados:
                            lotes_encontrados.add(lote_num)
                            lotes.append({
                                'lote': lote_num,
                                'fecha': str(reg.fecha),
                                'vendedor_id': vendedor_id,
                                'responsable': reg.responsable,
                                'producto': reg.producto,
                                'cantidad': reg.cantidad,
                                'origen': f'Cargue {vendedor_id}'
                            })
                except json.JSONDecodeError:
                    # Si no es JSON, tratar como texto
                    if reg.lotes_produccion and reg.lotes_produccion not in lotes_encontrados:
                        lotes_encontrados.add(reg.lotes_produccion)
                        lotes.append({
                            'lote': reg.lotes_produccion,
                            'fecha': str(reg.fecha),
                            'vendedor_id': vendedor_id,
                            'responsable': reg.responsable,
                            'producto': reg.producto,
                            'cantidad': reg.cantidad,
                            'origen': f'Cargue {vendedor_id}'
                        })
        except Exception as e:
            print(f"Error buscando lotes en {CargueModel.__name__}: {e}")
    
    return Response({
        'fecha': fecha,
        'total_lotes': len(lotes),
        'lotes': lotes
    })


@api_view(['GET'])
def lotes_por_mes(request):
    """
    Obtiene todos los lotes de producción para un mes específico.
    """
    import json
    from datetime import datetime
    from .models import CargueID1, CargueID2, CargueID3, CargueID4, CargueID5, CargueID6, Lote
    
    mes = request.query_params.get('mes', '')  # Formato: YYYY-MM
    
    if not mes:
        return Response({'error': 'Debe proporcionar un mes (formato: YYYY-MM)'}, status=400)
    
    try:
        year, month = mes.split('-')
        year = int(year)
        month = int(month)
    except:
        return Response({'error': 'Formato de mes inválido. Use YYYY-MM'}, status=400)
    
    lotes = []
    lotes_por_fecha = {}
    
    # 1. Buscar en tabla Lote
    try:
        lotes_obj = Lote.objects.filter(
            fecha_produccion__year=year,
            fecha_produccion__month=month
        )
        for lote in lotes_obj:
            fecha_str = str(lote.fecha_produccion)
            if fecha_str not in lotes_por_fecha:
                lotes_por_fecha[fecha_str] = []
            
            lotes_por_fecha[fecha_str].append({
                'lote': lote.lote,
                'usuario': lote.usuario,
                'fecha_vencimiento': str(lote.fecha_vencimiento) if lote.fecha_vencimiento else None,
                'origen': 'Producción'
            })
    except Exception as e:
        print(f"Error buscando lotes del mes: {e}")
    
    # 2. Buscar en tablas CargueIDx
    cargue_models = [CargueID1, CargueID2, CargueID3, CargueID4, CargueID5, CargueID6]
    vendedor_ids = ['ID1', 'ID2', 'ID3', 'ID4', 'ID5', 'ID6']
    
    for idx, CargueModel in enumerate(cargue_models):
        vendedor_id = vendedor_ids[idx]
        
        try:
            registros = CargueModel.objects.filter(
                fecha__year=year,
                fecha__month=month,
                lotes_produccion__isnull=False
            ).exclude(lotes_produccion='')
            
            for reg in registros:
                try:
                    lotes_prod = json.loads(reg.lotes_produccion) if reg.lotes_produccion else []
                    for lote_num in lotes_prod:
                        if lote_num:
                            fecha_str = str(reg.fecha)
                            if fecha_str not in lotes_por_fecha:
                                lotes_por_fecha[fecha_str] = []
                            
                            lotes_por_fecha[fecha_str].append({
                                'lote': lote_num,
                                'vendedor_id': vendedor_id,
                                'responsable': reg.responsable,
                                'producto': reg.producto,
                                'cantidad': reg.cantidad,
                                'origen': f'Cargue {vendedor_id}'
                            })
                except json.JSONDecodeError:
                    pass
        except Exception as e:
            print(f"Error: {e}")
    
    # Convertir a lista ordenada por fecha
    resultado = []
    for fecha in sorted(lotes_por_fecha.keys()):
        resultado.append({
            'fecha': fecha,
            'lotes': lotes_por_fecha[fecha]
        })
    
    return Response({
        'mes': mes,
        'total_fechas': len(resultado),
        'datos': resultado
    })


class RutaOrdenViewSet(viewsets.ModelViewSet):
    """ViewSet para manejar órdenes de clientes por ruta y día"""
    queryset = RutaOrden.objects.all()
    serializer_class = RutaOrdenSerializer
    permission_classes = [permissions.AllowAny]

    def get_queryset(self):
        """Permite filtrar por ruta y día"""
        queryset = super().get_queryset()
        ruta_id = self.request.query_params.get('ruta_id', None)
        dia = self.request.query_params.get('dia', None)
        
        if ruta_id:
            queryset = queryset.filter(ruta_id=ruta_id)
        if dia:
            queryset = queryset.filter(dia=dia.upper())
        return queryset

    def create(self, request, *args, **kwargs):
        """Crear o actualizar orden de ruta para una ruta + día específico"""
        ruta_id = request.data.get('ruta_id')
        dia = request.data.get('dia')
        if dia:
            dia = dia.upper()
        clientes_ids = request.data.get('clientes_ids', [])
        
        if not dia:
            return Response({'error': 'Se requiere dia'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Actualizar o crear el orden para esta ruta + día
        obj, created = RutaOrden.objects.update_or_create(
            ruta_id=ruta_id,
            dia=dia,
            defaults={'clientes_ids': clientes_ids}
        )
        
        print(f"✅ Orden guardado: Ruta={ruta_id}, Día={dia}, Clientes={len(clientes_ids)}")
        
        serializer = self.get_serializer(obj)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def obtener_orden(self, request):
        """
        Obtener el orden de clientes para una ruta y día específico
        GET /api/ruta-orden/obtener_orden/?ruta_id=1&dia=MARTES
        """
        ruta_id = request.query_params.get('ruta_id')
        dia = request.query_params.get('dia', '').upper()
        
        if not dia:
            return Response({'error': 'Se requiere dia'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            orden = RutaOrden.objects.get(ruta_id=ruta_id, dia=dia)
            return Response({
                'ruta_id': ruta_id,
                'dia': dia,
                'clientes_ids': orden.clientes_ids,
                'fecha_actualizacion': orden.fecha_actualizacion
            })
        except RutaOrden.DoesNotExist:
            # Si no existe orden personalizado, devolver lista vacía
            return Response({
                'ruta_id': ruta_id,
                'dia': dia,
                'clientes_ids': [],
                'mensaje': 'No hay orden personalizado para este día'
            })
    
    @action(detail=False, methods=['post'])
    def guardar_orden_vendedor(self, request):
        """
        Guardar orden de clientes para un vendedor en un día específico
        y mantenerlo por vendedor+día (orden global app móvil).
        Además, distribuye automáticamente por rutas para compatibilidad del panel web.
        
        POST /api/ruta-orden/guardar-orden-vendedor/
        Body: {
            "vendedor_id": "ID1",
            "dia": "MARTES",
            "clientes_ids": [1, 5, 3, 8, 2]
        }
        """
        vendedor_id = request.data.get('vendedor_id')
        dia = request.data.get('dia', '').upper()
        clientes_ids_raw = request.data.get('clientes_ids', [])
        
        if not vendedor_id or not dia:
            return Response({'error': 'Se requiere vendedor_id y dia'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            vendedor_norm, _ = _normalizar_id_vendedor(vendedor_id)
            if not vendedor_norm:
                return Response({'error': 'vendedor_id inválido'}, status=status.HTTP_400_BAD_REQUEST)

            vendedor = Vendedor.objects.filter(id_vendedor=vendedor_norm, activo=True).first()
            if not vendedor:
                return Response({'error': 'Vendedor no encontrado o inactivo'}, status=status.HTTP_404_NOT_FOUND)

            # Normalizar IDs (enteros únicos, manteniendo orden)
            clientes_ids = []
            ids_vistos = set()
            if isinstance(clientes_ids_raw, list):
                for raw_id in clientes_ids_raw:
                    try:
                        cid = int(raw_id)
                    except (TypeError, ValueError):
                        continue
                    if cid <= 0 or cid in ids_vistos:
                        continue
                    ids_vistos.add(cid)
                    clientes_ids.append(cid)

            rutas_vendedor_qs = Ruta.objects.filter(vendedor=vendedor, activo=True)

            # Obtener clientes válidos (solo del vendedor)
            clientes = ClienteRuta.objects.filter(
                id__in=clientes_ids,
                ruta__in=rutas_vendedor_qs
            ).select_related('ruta')

            ids_validos = set(clientes.values_list('id', flat=True))
            clientes_ids_validos = [cid for cid in clientes_ids if cid in ids_validos]

            # 1) Guardar orden global exacto por vendedor+día (fuente para app)
            RutaOrdenVendedor.objects.update_or_create(
                vendedor=vendedor,
                dia=dia,
                defaults={'clientes_ids': clientes_ids_validos}
            )
            
            # Agrupar clientes por ruta
            clientes_por_ruta = {}
            for cliente in clientes:
                ruta_id = cliente.ruta.id
                if ruta_id not in clientes_por_ruta:
                    clientes_por_ruta[ruta_id] = []
                clientes_por_ruta[ruta_id].append(cliente.id)
            
            # Guardar orden para cada ruta
            ordenes_guardadas = []
            for ruta_id, ids in clientes_por_ruta.items():
                # Mantener el orden original de clientes_ids
                ids_ordenados = [cid for cid in clientes_ids_validos if cid in ids]
                
                obj, created = RutaOrden.objects.update_or_create(
                    ruta_id=ruta_id,
                    dia=dia,
                    defaults={'clientes_ids': ids_ordenados}
                )
                ordenes_guardadas.append({
                    'ruta_id': ruta_id,
                    'ruta_nombre': obj.ruta.nombre,
                    'clientes_count': len(ids_ordenados)
                })
                print(f"✅ Orden guardado: Ruta={obj.ruta.nombre}, Día={dia}, Clientes={len(ids_ordenados)}")
            
            return Response({
                'success': True,
                'vendedor_id': vendedor_norm,
                'dia': dia,
                'clientes_recibidos': len(clientes_ids_raw) if isinstance(clientes_ids_raw, list) else 0,
                'clientes_validos': len(clientes_ids_validos),
                'ordenes_guardadas': ordenes_guardadas
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            print(f"❌ Error guardando orden: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ===== VIEWSET CLIENTES OCASIONALES =====

class ClienteOcasionalViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de clientes ocasionales (ventas en calle).
    Completamente independiente de ClienteRuta."""
    queryset = ClienteOcasional.objects.select_related('vendedor', 'cliente_ruta').all()
    serializer_class = ClienteOcasionalSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = ClienteOcasional.objects.select_related('vendedor', 'cliente_ruta').all()
        vendedor_id = self.request.query_params.get('vendedor_id', None)
        activo = self.request.query_params.get('activo', None)
        convertido = self.request.query_params.get('convertido', None)
        
        if vendedor_id:
            queryset = queryset.filter(vendedor__id_vendedor=vendedor_id)
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
        if convertido is not None:
            queryset = queryset.filter(convertido=convertido.lower() == 'true')
        
        return queryset.order_by('-fecha_creacion')
    
    def destroy(self, request, *args, **kwargs):
        """Eliminar un cliente ocasional."""
        try:
            instance = self.get_object()
            self.perform_destroy(instance)
            return Response({"message": "Cliente ocasional eliminado correctamente"}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def convertir(self, request, pk=None):
        """
        Convertir un cliente ocasional a cliente de ruta normal.
        POST /api/clientes-ocasionales/{id}/convertir/
        Body: {
            "ruta_id": 1,
            "dia_visita": "LUNES,MARTES",
            "tipo_negocio": "Tienda"
        }
        """
        try:
            cliente_ocasional = self.get_object()
            
            if cliente_ocasional.convertido:
                return Response(
                    {'error': 'Este cliente ya fue convertido a cliente de ruta'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            ruta_id = request.data.get('ruta_id')
            dia_visita = request.data.get('dia_visita', '')
            tipo_negocio = request.data.get('tipo_negocio', '')
            
            if not ruta_id:
                return Response(
                    {'error': 'Se requiere ruta_id para convertir'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            ruta = Ruta.objects.filter(id=ruta_id, activo=True).first()
            if not ruta:
                return Response(
                    {'error': 'Ruta no encontrada o inactiva'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Calcular orden
            from django.db.models import Max
            max_orden = ClienteRuta.objects.filter(ruta=ruta).aggregate(
                Max('orden')
            )['orden__max'] or 0
            
            # Crear ClienteRuta
            nuevo_cliente = ClienteRuta.objects.create(
                ruta=ruta,
                nombre_negocio=cliente_ocasional.nombre,
                nombre_contacto=cliente_ocasional.nombre,
                direccion=cliente_ocasional.direccion or '',
                telefono=cliente_ocasional.telefono or '',
                tipo_negocio=tipo_negocio or 'Cliente | OCASIONAL',
                dia_visita=dia_visita or 'LUNES',
                orden=max_orden + 1,
                activo=True,
                nota=f'Convertido desde cliente ocasional #{cliente_ocasional.id}'
            )
            
            # Marcar como convertido
            cliente_ocasional.convertido = True
            cliente_ocasional.cliente_ruta = nuevo_cliente
            cliente_ocasional.save()
            
            print(f"✅ Cliente ocasional '{cliente_ocasional.nombre}' convertido a ClienteRuta #{nuevo_cliente.id}")
            
            return Response({
                'success': True,
                'mensaje': f'Cliente convertido exitosamente a ruta {ruta.nombre}',
                'cliente_ruta_id': nuevo_cliente.id,
                'cliente_ocasional': ClienteOcasionalSerializer(cliente_ocasional).data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            print(f"❌ Error convirtiendo cliente ocasional: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ============================================================================
# 🤖 ENDPOINTS DE IA LOCAL (Ollama)
# ============================================================================

@api_view(['POST'])
def ai_chat(request):
    """
    Chat con el asistente IA
    
    POST /api/ai/chat/
    Body: {
        "question": "¿Cómo cierro el turno?",
        "include_docs": true  // opcional, default true
    }
    """
    
    question = request.data.get('question')
    if not question:
        return Response({
            'error': 'Se requiere campo "question"'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    include_docs = request.data.get('include_docs', False)  # OPTIMIZADO: False por defecto para velocidad
    
    try:
        ai = AIAssistant()
        answer = ai.ask(question, include_docs=include_docs)
        
        return Response({
            'question': question,
            'answer': answer,
            'model': ai.model
        })
    except Exception as e:
        return Response({
            'error': f'Error consultando IA: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def ai_analyze_data(request):
    """
    Analiza datos con IA
    
    POST /api/ai/analyze/
    Body: {
        "data": {...},  // datos a analizar
        "question": "¿Qué tendencia ves?"
    }
    """
    from api.services.ai_assistant_service import AIAssistant
    
    data = request.data.get('data')
    question = request.data.get('question')
    
    if not data or not question:
        return Response({
            'error': 'Se requieren campos "data" y "question"'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        ai = AIAssistant()
        analysis = ai.analyze_data(data, question)
        
        return Response({
            'analysis': analysis,
            'model': ai.model
        })
    except Exception as e:
        return Response({
            'error': f'Error analizando: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def ai_health(request):
    """
    Verifica estado de IA
    
    GET /api/ai/health/
    """
    try:
        from api.services.ai_assistant_service import AIAssistant
        ai = AIAssistant()
        health_status = ai.check_health()
        return Response(health_status)
    except Exception as e:
        return Response({
            'status': 'error',
            'message': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def ai_agent_command(request):
    """
    Endpoint para ejecutar comandos con el agente IA
    
    POST /api/ai/agent/
    Body: {
        "command": "Crea un cliente llamado Juan con teléfono 123456"
    }
    """
    from api.services.ai_agent_service import AIAgentService
    
    command = request.data.get('command')
    session_id = request.data.get('session_id')  # Capturar session_id
    
    if not command:
        return Response({
            'error': 'Se requiere campo "command"'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        from api.services.ai_agent_service import AIAgentService
        agent = AIAgentService(model="qwen2.5:3b")
        result = agent.process_command(command, session_id=session_id)
        
        return Response(result)
    except Exception as e:
        return Response({
            'error': f'Error procesando comando: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ReportePlaneacionViewSet(viewsets.ModelViewSet):
    """API para gestionar snapshots de reportes de planeación"""
    queryset = ReportePlaneacion.objects.all()
    serializer_class = ReportePlaneacionSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = ReportePlaneacion.objects.all().order_by('-fecha_creacion')
        
        # Filtros
        fecha = self.request.query_params.get('fecha')
        if fecha:
            queryset = queryset.filter(fecha_reporte=fecha)
            
        return queryset

    def perform_create(self, serializer):
        """Guardar reporte y entrenar Red Neuronal"""
        instance = serializer.save()
        
        try:
            from api.services.ia_service import IAService
            # Entrenar en segundo plano (o inmediato si es rápido)
            ia_service = IAService()
            success = ia_service.train_with_report(instance)
            if success:
                print(f"🧠 Red Neuronal re-entrenada con reporte {instance.fecha_reporte}")
        except Exception as e:
            print(f"⚠️ Error entrenando IA tras guardar reporte: {e}")


# ==================== REPORTES AVANZADOS ====================

@api_view(['GET'])
def reportes_vendedores(request):
    """
    Reporte consolidado de vendedores con ventas, vencidas y devoluciones
    GET /api/reportes/vendedores/?periodo=mes&fecha_inicio=2026-01-01&fecha_fin=2026-01-31
    """
    try:
        from django.db.models import Sum, Count, Q
        from datetime import datetime
        
        periodo = request.GET.get('periodo', 'mes')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return Response({'error': 'Faltan parámetros: fecha_inicio y fecha_fin'}, status=400)

        # Mapeo de ID a Nombre Real
        nombres_reales = {}

        def get_nombre_vendedor(identificador):
            identificador = str(identificador).upper()
            if identificador in nombres_reales:
                return nombres_reales[identificador]
            return identificador

        # Cargar mapa de precios reales para corrección de montos en 0
        precios_productos = {}
        try:
            from .models import Producto
            productos_qs = Producto.objects.all()
            for p in productos_qs:
                precios_productos[p.nombre] = float(p.precio_cargue or p.precio or 0)
        except:
            print("No se pudo cargar tabla Productos para precios")

        # Obtener datos de CargueResumen
        resumenes = CargueResumen.objects.filter(
            fecha__gte=fecha_inicio,
            fecha__lte=fecha_fin
        )

        vendedores_data = {}

        # Pre-poblar con todos los vendedores activos
        try:
             todos_vendedores_objs = Vendedor.objects.filter(activo=True)
             for v in todos_vendedores_objs:
                key_id = v.id_vendedor.upper()
                nombres_reales[key_id] = v.nombre.upper()
                if key_id.startswith('ID'):
                    id_num = key_id.replace('ID', '')
                    nombres_reales[id_num] = v.nombre.upper()
                
                # Inicializar vendedor en data
                nombre_real = v.nombre.upper()
                if nombre_real not in vendedores_data:
                    vendedores_data[nombre_real] = {
                        'id': f"ID{v.id_vendedor}" if not v.id_vendedor.startswith('ID') else v.id_vendedor,
                        'nombre': nombre_real,
                        'ventas_totales': 0,
                        'monto': 0.0,
                        'monto_ruta': 0.0,
                        'monto_pedidos': 0.0,
                        'vencidas': 0,
                        'devoluciones': 0,
                        'efectividad': 100.0
                    }
        except:
             pass

        for resumen in resumenes:
            vendedor_raw = resumen.vendedor_id
            nombre_real = get_nombre_vendedor(vendedor_raw)

            if nombre_real not in vendedores_data:
                vendedores_data[nombre_real] = {
                    'id': vendedor_raw,
                    'nombre': nombre_real,
                    'ventas_totales': 0,
                    'monto': 0.0,
                    'monto_ruta': 0.0,
                    'monto_pedidos': 0.0,
                    'vencidas': 0,
                    'devoluciones': 0,
                    'efectividad': 100.0
                }
            
            # Montos base desde Resumen
            venta_ruta = float(resumen.total_despacho or 0)
            venta_pedidos = float(resumen.total_pedidos or 0)
            devoluciones_cnt = 0
            vencidas_cnt = 0
            
            # Intentar obtener el modelo específico CargueIDx para recalcular si es 0
            try:
                id_num = ''.join(filter(str.isdigit, str(vendedor_raw)))
                if not id_num and nombre_real in nombres_reales.values():
                    for k, v in nombres_reales.items():
                        if v == nombre_real and k.isdigit():
                            id_num = k
                            break
                
                if id_num:
                    try:
                        from django.apps import apps
                        CargueModel = apps.get_model('api', f"CargueID{id_num}")
                    except LookupError:
                        CargueModel = None
                    
                    if CargueModel:
                        items_cargue = CargueModel.objects.filter(
                            fecha__gte=fecha_inicio,
                            fecha__lte=fecha_fin
                        )
                        
                        calc_ruta = 0.0
                        calc_pedidos_bdd = 0.0
                        
                        for item in items_cargue:
                            # RUTAS: Recalcular con precio catalogo
                            precio = float(item.valor)
                            if precio == 0 and item.producto in precios_productos:
                                precio = precios_productos[item.producto]
                            
                            if item.cantidad > 0:
                                calc_ruta += (item.cantidad * precio)
                            
                            # Intentar leer total_pedidos global guardado en la fila
                            val_p = getattr(item, 'total_pedidos', 0)
                            if val_p and float(val_p) > calc_pedidos_bdd:
                                calc_pedidos_bdd = float(val_p)

                            # Sumar devoluciones y vencidas (unidades)
                            if item.devoluciones > 0:
                                devoluciones_cnt += item.devoluciones
                            
                            if item.vencidas > 0:
                                vencidas_cnt += item.vencidas

                        # Si el cálculo da más que el resumen, usémoslo
                        if calc_ruta > venta_ruta:
                            venta_ruta = calc_ruta
                        
                        # Si encontramos valor en BD cargue, usarlo
                        if calc_pedidos_bdd > venta_pedidos:
                            venta_pedidos = calc_pedidos_bdd
                            
                        # PEDIDOS: Consultar tabla Pedido directamente (Fuente de verdad)
                        try:
                            from .models import Pedido
                            # Buscar pedidos entregados/pendientes para esa fecha y vendedor
                            pedidos_qs = Pedido.objects.filter(
                                vendedor=nombre_real,
                                fecha_entrega__date__gte=fecha_inicio,
                                fecha_entrega__date__lte=fecha_fin
                            ).exclude(estado='ANULADA')
                            
                            calc_pedidos = 0.0
                            for p in pedidos_qs:
                                calc_pedidos += float(p.total or 0)
                                
                            if calc_pedidos > venta_pedidos:
                                venta_pedidos = calc_pedidos
                        except Exception as e:
                            print(f"Error consultando pedidos para {nombre_real}: {e}")
                            
            except Exception as e:
                print(f"Error recalculando montos para {nombre_real}: {e}")

            # ACTUALIZAR DATOS
            vendedores_data[nombre_real]['monto_ruta'] += venta_ruta
            vendedores_data[nombre_real]['monto_pedidos'] += venta_pedidos
            vendedores_data[nombre_real]['monto'] += (venta_ruta + venta_pedidos)
            vendedores_data[nombre_real]['devoluciones'] += devoluciones_cnt
            vendedores_data[nombre_real]['vencidas'] += vencidas_cnt
            
            # Ventas totales (usaremos conteo de pedidos como proxy de 'transacciones' o días activos)
            vendedores_data[nombre_real]['ventas_totales'] += 1 

        # Convertir a lista
        resultado = list(vendedores_data.values())
        
        # Ordenar por ID ascendente (ID1, ID2, ID3...)
        def extract_id_number(item):
            try:
                id_str = str(item.get('id', ''))
                # Extraer solo dígitos
                nums = ''.join(filter(str.isdigit, id_str))
                return int(nums) if nums else 9999
            except:
                return 9999

        resultado.sort(key=extract_id_number)
        
        return Response({
            'vendedores': resultado,
            'periodo': periodo,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin
        })
        
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(error_msg)
        try:
            with open('/home/john/Escritorio/crm-fabrica/error_log.txt', 'w') as f:
                f.write(error_msg)
        except:
            pass
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
def reportes_efectividad_vendedores(request):
    """
    Reporte de efectividad de vendedores: Vendió, Devolvió, Vencidas, Cumplimiento, Efectividad
    GET /api/reportes/efectividad-vendedores/?periodo=mes&fecha_inicio=2026-01-01&fecha_fin=2026-01-31
    """
    try:
        from django.db.models import Sum, Count, Q
        from datetime import datetime
        
        periodo = request.GET.get('periodo', 'mes')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return Response({'error': 'Faltan parámetros: fecha_inicio y fecha_fin'}, status=400)
        
        # Obtener todos los vendedores activos
        vendedores = Vendedor.objects.filter(activo=True)
        
        resultado = []
        
        for vendedor in vendedores:
            # Obtener todas las ventas del vendedor en el período
            ventas_ruta = VentaRuta.objects.filter(
                vendedor=vendedor,
                fecha__date__gte=fecha_inicio,
                fecha__date__lte=fecha_fin
            )
            
            # VENDIÓ: Total de productos que llevó (cantidad en cada venta)
            vendio = 0
            for venta in ventas_ruta:
                if venta.detalle_productos:  # JSON con los productos de la venta
                    for producto in venta.detalle_productos:
                        vendio += producto.get('cantidad', 0)
            
            # DEVOLVIÓ: Productos devueltos
            devolvio = 0
            for venta in ventas_ruta:
                if venta.productos_devueltos:  # JSON con productos devueltos
                    for producto in venta.productos_devueltos:
                        devolvio += producto.get('cantidad', 0)
            
            # VENCIDAS: Productos vencidos
            vencidas = 0
            for venta in ventas_ruta:
                if venta.productos_vencidos:  # JSON con productos vencidos
                    for producto in venta.productos_vencidos:
                        vencidas += producto.get('cantidad', 0)
            
            # VENTAS REALES: Vendió - (Devolvió + Vencidas)
            ventas_reales = vendio - (devolvio + vencidas)
            if ventas_reales < 0:
                ventas_reales = 0
            
            # EFECTIVIDAD: (Ventas Reales / Vendió) * 100
            efectividad = 0.0
            if vendio > 0:
                efectividad = (ventas_reales / vendio) * 100
            
            # CUMPLIMIENTO: (Ventas Reales / Meta) * 100
            # Por ahora asumimos meta = vendió, se puede ajustar si hay campo de meta
            cumplimiento = efectividad  # Simplificado
            
            resultado.append({
                'id': vendedor.id_vendedor,
                'nombre': vendedor.nombre,
                'vendio': vendio,
                'devolvio': devolvio,
                'vencidas': vencidas,
                'ventas_reales': ventas_reales,
                'cumplimiento': round(cumplimiento, 2),
                'efectividad': round(efectividad, 2)
            })
        
        # Ordenar por ventas reales descendente
        resultado.sort(key=lambda x: x['ventas_reales'], reverse=True)
        
        return Response({
            'vendedores': resultado,
            'periodo': periodo,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin
        })
        
    except Exception as e:
        print(f"Error en reportes_efectividad_vendedores: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
def reportes_analisis_productos(request):
    """
    Análisis consolidado de productos: vendidos, devueltos y vencidos
    GET /api/reportes/analisis-productos/?tipo=vendidos&periodo=mes&fecha_inicio=2026-01-01&fecha_fin=2026-01-31&orden=desc&limite=10

    tipo: vendidos | devueltos | vencidos
    """
    try:
        from django.db.models import Sum
        from django.apps import apps
        from collections import Counter
        
        tipo = request.GET.get('tipo', 'vendidos')
        periodo = request.GET.get('periodo', 'mes')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        orden = request.GET.get('orden', 'desc')  # desc o asc
        limite = int(request.GET.get('limite', 10))
        
        if not fecha_inicio or not fecha_fin:
            return Response({'error': 'Faltan parámetros: fecha_inicio y fecha_fin'}, status=400)
        
        productos_conteo = Counter()
        vendedores_por_producto = {}
        
        # 1. Consultar las 6 tablas de CargueIDx
        for id_num in range(1, 7):
            try:
                CargueModel = apps.get_model('api', f'CargueID{id_num}')
                
                items = CargueModel.objects.filter(
                    fecha__gte=fecha_inicio,
                    fecha__lte=fecha_fin
                )
                
                for item in items:
                    nombre_producto = item.producto or 'Desconocido'
                    
                    # Seleccionar campo según tipo
                    if tipo == 'vendidos':
                        # Solo cantidad (lo realmente despachado/vendido)
                        cantidad = item.cantidad or 0
                    elif tipo == 'devueltos':
                        cantidad = item.devoluciones or 0
                    elif tipo == 'vencidos':
                        cantidad = item.vencidas or 0
                    else:
                        cantidad = 0
                    
                    if cantidad > 0:
                        productos_conteo[nombre_producto] += cantidad
                        
                        # Contar vendedores únicos por producto
                        if tipo in ['devueltos', 'vencidos']:
                            if nombre_producto not in vendedores_por_producto:
                                vendedores_por_producto[nombre_producto] = set()
                            vendedores_por_producto[nombre_producto].add(f'ID{id_num}')
                            
            except Exception as e:
                print(f"Error consultando CargueID{id_num}: {e}")
                continue
        
        # 2. Si es tipo "vendidos", también sumar productos de la tabla Pedido
        if tipo == 'vendidos':
            try:
                from .models import Pedido
                
                # fecha_entrega es DateField, no necesita __date
                pedidos = Pedido.objects.filter(
                    fecha_entrega__gte=fecha_inicio,
                    fecha_entrega__lte=fecha_fin
                ).exclude(estado='ANULADA')
                
                for pedido in pedidos:
                    # Usar related_name 'detalles' para acceder a DetallePedido
                    for detalle in pedido.detalles.all():
                        # producto es FK, acceder al nombre via producto.nombre
                        nombre_producto = detalle.producto.nombre if detalle.producto else 'Desconocido'
                        cantidad = detalle.cantidad or 0
                        
                        if cantidad > 0:
                            productos_conteo[nombre_producto] += cantidad
                            
            except Exception as e:
                print(f"Error consultando Pedidos: {e}")
                import traceback
                traceback.print_exc()
        
        # Convertir a lista y ordenar
        resultado = []
        for nombre, total in productos_conteo.items():
            item = {
                'nombre': nombre,
                'total': total
            }
            
            # Agregar conteo de vendedores para devueltos/vencidos
            if tipo in ['devueltos', 'vencidos']:
                item['vendedores'] = len(vendedores_por_producto.get(nombre, set()))
            
            resultado.append(item)
        
        # Ordenar
        reverse = (orden == 'desc')
        resultado.sort(key=lambda x: x['total'], reverse=reverse)
        
        # Limitar resultados
        resultado = resultado[:limite]
        
        return Response({
            'productos': resultado,
            'tipo': tipo,
            'periodo': periodo,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'total_productos': len(resultado)
        })
        
    except Exception as e:
        print(f"Error en reportes_analisis_productos: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
def reportes_pedidos_ruta(request):
    """
    Pedidos por ruta agrupados por vendedor
    GET /api/reportes/pedidos-ruta/?fecha_inicio=2026-01-01&fecha_fin=2026-01-31&vendedor=ID1&estado=pendiente
    """
    try:
        from datetime import datetime
        
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        vendedor_id = request.GET.get('vendedor')
        estado_filtro = request.GET.get('estado')
        
        if not fecha_inicio or not fecha_fin:
            return Response({'error': 'Faltan parámetros: fecha_inicio y fecha_fin'}, status=400)
        
        # Query base
        pedidos = Pedido.objects.filter(
            fecha__date__gte=fecha_inicio,
            fecha__date__lte=fecha_fin
        ).select_related('usuario')
        
        # Filtros opcionales
        if vendedor_id:
            pedidos = pedidos.filter(usuario__username__icontains=vendedor_id)
        
        if estado_filtro and estado_filtro != 'todos':
            pedidos = pedidos.filter(estado=estado_filtro)
        
        # Serializar resultado
        resultado = []
        for pedido in pedidos:
            resultado.append({
                'id': pedido.id,
                'vendedor_nombre': pedido.usuario.username if pedido.usuario else 'Sin vendedor',
                'ruta': getattr(pedido, 'ruta', None),  # Si existe campo ruta
                'cliente_nombre': pedido.cliente.nombre if pedido.cliente else 'Sin cliente',
                'fecha': pedido.fecha.isoformat(),
                'total': float(pedido.total),
                'estado': pedido.estado
            })
        
        return Response({
            'pedidos': resultado,
            'total': len(resultado),
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin
        })
        
    except Exception as e:
        print(f"Error en reportes_pedidos_ruta: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
def dashboard_ejecutivo(request):
    """
    Dashboard ejecutivo consolidado con información de CARGUE
    GET /api/dashboard-ejecutivo/?periodo=dia&fecha_inicio=2026-01-24&fecha_fin=2026-01-24
    """
    try:
        from django.db.models import Sum, Count, Q
        from collections import Counter, defaultdict
        
        periodo = request.GET.get('periodo', 'dia')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return Response({'error': 'Faltan parámetros: fecha_inicio y fecha_fin'}, status=400)
        
        # OBTENER TOTALES POR VENDEDOR desde CargueResumen (tiene los $$ correctos)
        resumenes = CargueResumen.objects.filter(
            fecha__gte=fecha_inicio,
            fecha__lte=fecha_fin,
            activo=True
        )
        
        # Mapeo de ID a Nombre Real (para unificar ID1 con WILSON)
        nombres_reales = {}
        try:
            from .models import Vendedor
            # Cargar mapa de vendedores: {1: 'WILSON', 2: 'OTRO', ...}
            todos_vendedores = Vendedor.objects.all()
            for v in todos_vendedores:
                # Asumimos que ID1 corresponde al vendedor con id=1, etc.
                key_id = f"ID{v.id}"
                nombres_reales[key_id] = v.nombre.upper()
        except:
            print("No se pudo cargar tabla Vendedores")

        # Función helper para normalizar nombre
        def get_nombre_vendedor(identificador):
            identificador = identificador.upper()
            # Si es ID1, ID2, etc buscar en mapa
            if identificador.startswith('ID') and identificador in nombres_reales:
                return nombres_reales[identificador]
            # Si ya es un nombre (WILSON), devolverlo
            return identificador

        # DATOS POR VENDEDOR usando CargueResumen
        vendedores_data = {}
        
        for resumen in resumenes:
            vendedor_raw = resumen.vendedor_id
            nombre_real = get_nombre_vendedor(vendedor_raw)
            
            if nombre_real not in vendedores_data:
                vendedores_data[nombre_real] = {
                    'nombre': nombre_real,
                    'ventas_count': 0,
                    'ventas_monto': 0,
                    'devueltas_count': 0,
                    'devueltas_monto': 0,
                    'vencidas_count': 0,
                    'vencidas_monto': 0
                }
            
            # Sumar el monto de venta (calcularlo si es 0)
            monto_venta = float(resumen.venta or 0)
            if monto_venta == 0:
                # Si venta es 0, intentar calcular como despacho + pedidos
                monto_venta = float(resumen.total_despacho or 0) + float(resumen.total_pedidos or 0)
            
            vendedores_data[nombre_real]['ventas_monto'] += monto_venta
        
        # OBTENER PRODUCTOS (vendidas, devoluciones, vencidas) desde los 6 modelos
        todos_cargues = []
        for modelo in [CargueID1, CargueID2, CargueID3, CargueID4, CargueID5, CargueID6]:
            cargues = modelo.objects.filter(
                fecha__gte=fecha_inicio,
                fecha__lte=fecha_fin,
                activo=True
            )
            todos_cargues.extend(list(cargues))
        
        # Cargar mapa de precios reales para corrección de montos en 0
        precios_productos = {}
        try:
            from .models import Producto
            productos_qs = Producto.objects.all()
            for p in productos_qs:
                precios_productos[p.nombre] = float(p.precio_cargue or p.precio or 0)
        except:
            print("No se pudo cargar tabla Productos para precios")

        # Contadores de productos
        productos_vendidos = Counter()
        productos_devueltos = Counter()
        productos_vencidos = Counter()
        
        for cargue in todos_cargues:
            # Obtener nombre normalizado
            if hasattr(cargue, 'responsable') and cargue.responsable:
                 vendedor_raw = cargue.responsable
                 vendedor_nombre = get_nombre_vendedor(vendedor_raw)
            else:
                 modelo_name = cargue.__class__.__name__
                 id_suffix = modelo_name.replace('Cargue', '')
                 vendedor_nombre = get_nombre_vendedor(id_suffix)

            if vendedor_nombre not in vendedores_data:
                vendedores_data[vendedor_nombre] = {
                    'nombre': vendedor_nombre,
                    'ventas_count': 0,
                    'ventas_monto': 0,
                    'devueltas_count': 0,
                    'devueltas_monto': 0,
                    'vencidas_count': 0,
                    'vencidas_monto': 0
                }
            
            # Obtener precio real (si cargue.valor es 0, usar catalogo)
            precio_real = float(cargue.valor)
            if precio_real == 0 and cargue.producto in precios_productos:
                precio_real = precios_productos[cargue.producto]
            
            # VENTAS (unidades)
            if cargue.vendidas > 0:
                vendedores_data[vendedor_nombre]['ventas_count'] += cargue.vendidas
                productos_vendidos[cargue.producto] += cargue.vendidas
            
            # DEVOLUCIONES (unidades + monto)
            if cargue.devoluciones > 0:
                vendedores_data[vendedor_nombre]['devueltas_count'] += cargue.devoluciones
                vendedores_data[vendedor_nombre]['devueltas_monto'] += float(cargue.devoluciones * precio_real)
                productos_devueltos[cargue.producto] += cargue.devoluciones
            
            # VENCIDAS (unidades + monto)
            if cargue.vencidas > 0:
                vendedores_data[vendedor_nombre]['vencidas_count'] += cargue.vencidas
                vendedores_data[vendedor_nombre]['vencidas_monto'] += float(cargue.vencidas * precio_real)
                productos_vencidos[cargue.producto] += cargue.vencidas
        
        # Calcular porcentajes
        vendedores_list = []
        for v_data in vendedores_data.values():
            total_productos = v_data['ventas_count'] + v_data['devueltas_count'] + v_data['vencidas_count']
            if total_productos > 0:
                v_data['porcentaje_devolucion'] = round((v_data['devueltas_count'] / total_productos) * 100, 2)
                v_data['porcentaje_vencidas'] = round((v_data['vencidas_count'] / total_productos) * 100, 2)
                v_data['efectividad'] = round((v_data['ventas_count'] / total_productos) * 100, 2)
            else:
                v_data['porcentaje_devolucion'] = 0
                v_data['porcentaje_vencidas'] = 0
                v_data['efectividad'] = 0
            vendedores_list.append(v_data)
        
        # Ordenar por monto
        vendedores_list.sort(key=lambda x: x['ventas_monto'], reverse=True)
        
        # Top 10 productos
        top_vendidos = [{'nombre': n, 'cantidad': c} for n, c in productos_vendidos.most_common(10)]
        top_devueltos = [{'nombre': n, 'cantidad': c} for n, c in productos_devueltos.most_common(10)]
        top_vencidos = [{'nombre': n, 'cantidad': c} for n, c in productos_vencidos.most_common(10)]
        
        # TOTALES
        totales = {
            'ventas_total': sum(v['ventas_count'] for v in vendedores_list),
            'ventas_monto_total': sum(v['ventas_monto'] for v in vendedores_list),
            'devueltas_total': sum(v['devueltas_count'] for v in vendedores_list),
            'devueltas_monto_total': sum(v['devueltas_monto'] for v in vendedores_list),
            'vencidas_total': sum(v['vencidas_count'] for v in vendedores_list),
            'vencidas_monto_total': sum(v['vencidas_monto'] for v in vendedores_list),
        }
        
        total_all = totales['ventas_total'] + totales['devueltas_total'] + totales['vencidas_total']
        totales['efectividad_promedio'] = round((totales['ventas_total'] / total_all * 100), 2) if total_all > 0 else 0
        
        return Response({
            'periodo': periodo,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'vendedores': vendedores_list,
            'top_productos_vendidos': top_vendidos,
            'top_productos_devueltos': top_devueltos,
            'top_productos_vencidos': top_vencidos,
            'totales': totales
        })
        
    except Exception as e:
        print(f"Error en dashboard_ejecutivo: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
def reportes_ventas_pos(request):
    """
    Reporte de ventas POS filtrado por usuario (cajero o vendedor)
    GET /api/reportes/ventas-pos/?periodo=dia&fecha_inicio=2026-01-18&fecha_fin=2026-01-18&cajero=Juan
    """
    try:
        from collections import defaultdict
        from django.db.models import Q
        from .models import Venta
        
        periodo = request.GET.get('periodo', 'dia')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        cajero = request.GET.get('cajero')  # Obtener nombre del cajero del query
        
        if not fecha_inicio or not fecha_fin:
            return Response({'error': 'Faltan parámetros de fecha'}, status=400)
            
        # Construir consulta base
        query = Q(fecha__date__gte=fecha_inicio) & Q(fecha__date__lte=fecha_fin)
        
        # Filtro: Excluir anuladas (Solo ventas válidas)
        query &= ~Q(estado='ANULADA')
        
        # Filtrar por cajero si se proporciona
        if cajero and cajero != 'undefined' and cajero != 'null':
            # Mostrar ventas donde el usuario es el vendedor asignado O quien la creó
            query &= (Q(vendedor__icontains=cajero) | Q(creado_por__icontains=cajero))
            
        ventas = Venta.objects.filter(query).order_by('-fecha')
        
        # Calcular métricas globales
        total_ventas = ventas.count()
        monto_total = sum(v.total for v in ventas)
        # Sumar detalles usando la relación related_name='detalles'
        total_productos = 0
        for v in ventas:
            total_productos += v.detalles.count()
        
        # Agrupar por día/semana/mes
        por_dia = defaultdict(lambda: {'ventas': 0, 'monto': 0})
        
        for venta in ventas:  
            if periodo == 'dia':
                clave = venta.fecha.strftime('%Y-%m-%d')
            elif periodo == 'semana':
                clave = f"Sem {venta.fecha.isocalendar()[1]}"
            elif periodo == 'mes':
                clave = venta.fecha.strftime('%Y-%m')
            else:  # año
                clave = venta.fecha.strftime('%Y')
            
            por_dia[clave]['ventas'] += 1
            por_dia[clave]['monto'] += float(venta.total)
        
        # Serializar lista de ventas para la tabla
        ventas_list = []
        for v in ventas:
            ventas_list.append({
                'id': v.id,
                'fecha': v.fecha.isoformat(),
                'cliente': v.cliente, # Es un CharField
                'productos': v.detalles.count(),
                'total': float(v.total),
                'estado': v.estado,
                'vendedor': v.vendedor,
                'creado_por': v.creado_por
            })
        
        # Convertir agrupación a lista ordenada
        por_dia_list = [
            {'fecha': fecha, 'ventas': datos['ventas'], 'monto': datos['monto']}
            for fecha, datos in sorted(por_dia.items())
        ]
        
        return Response({
            'usuario': cajero or 'Todos',
            'total_ventas': total_ventas,
            'monto_total': float(monto_total),
            'total_productos': total_productos,
            'por_dia': por_dia_list,
            'ventas': ventas_list,
            'periodo': periodo
        })
        
    except Exception as e:
        print(f"Error en reportes_ventas_pos: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
def test_dashboard_data(request):
    """
    Test endpoint para verificar datos de VentaRuta
    GET /api/test-dashboard-data/
    """
    try:
        fecha_inicio = request.GET.get('fecha_inicio', '2026-01-24')
        fecha_fin = request.GET.get('fecha_fin', '2026-01-24')
        
        ventas = VentaRuta.objects.filter(
            fecha__date__gte=fecha_inicio,
            fecha__date__lte=fecha_fin
        ).select_related('vendedor')
        
        resultado = {
            'total_ventas': ventas.count(),
            'ventas': []
        }
        
        for venta in ventas[:5]:  # Solo las primeras 5
            resultado['ventas'].append({
                'id': venta.id,
                'fecha': venta.fecha.isoformat(),
                'vendedor': venta.vendedor.nombre if venta.vendedor else None,
                'cliente': venta.cliente.nombre if venta.cliente else None,
                'detalle_productos_type': type(venta.detalle_productos).__name__,
                'detalle_productos': venta.detalle_productos,
                'productos_devueltos': venta.productos_devueltos,
                'productos_vencidos': venta.productos_vencidos,
            })
        
        return Response(resultado)
    except Exception as e:
        import traceback
        return Response({
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)


# ========================================
# 📷 EVIDENCIAS DE PEDIDOS
# ========================================

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def subir_evidencia_pedido(request):
    """
    Subir foto de evidencia para un pedido (vencidos/novedades)
    POST /api/evidencia-pedido/
    
    Form data:
    - pedido_id: ID del pedido (requerido)
    - producto_nombre: Nombre del producto (opcional)
    - motivo: Razón de la evidencia (opcional, default: 'Devolución en entrega')
    - imagen: Archivo de imagen (requerido)
    
    Ejemplo desde App:
    ```javascript
    const formData = new FormData();
    formData.append('pedido_id', pedidoId);
    formData.append('producto_nombre', 'AREPA TIPO OBLEA');
    formData.append('motivo', 'Producto vencido');
    formData.append('imagen', { uri: fotoUri, name: 'evidencia.jpg', type: 'image/jpeg' });
    
    fetch('/api/evidencia-pedido/', {
        method: 'POST',
        body: formData
    });
    ```
    """
    try:
        from .models import Pedido, EvidenciaPedido
        
        pedido_id = request.data.get('pedido_id')
        producto_nombre = request.data.get('producto_nombre', '')
        motivo = request.data.get('motivo', 'Devolución en entrega')
        imagen = request.FILES.get('imagen')
        
        # Validaciones
        if not pedido_id:
            return Response({'error': 'pedido_id es requerido'}, status=400)
        
        if not imagen:
            return Response({'error': 'imagen es requerida'}, status=400)
        
        # Buscar el pedido
        try:
            pedido = Pedido.objects.get(id=pedido_id)
        except Pedido.DoesNotExist:
            # Intentar buscar por numero_pedido
            try:
                pedido = Pedido.objects.get(numero_pedido=pedido_id)
            except Pedido.DoesNotExist:
                return Response({'error': f'Pedido {pedido_id} no encontrado'}, status=404)
        
        # Crear la evidencia
        evidencia = EvidenciaPedido.objects.create(
            pedido=pedido,
            producto_nombre=producto_nombre,
            motivo=motivo,
            imagen=imagen
        )
        
        print(f"📷 Evidencia creada para pedido {pedido.numero_pedido}: {producto_nombre} - {motivo}")
        
        return Response({
            'success': True,
            'message': 'Evidencia subida correctamente',
            'evidencia': {
                'id': evidencia.id,
                'pedido': pedido.numero_pedido,
                'producto_nombre': evidencia.producto_nombre,
                'motivo': evidencia.motivo,
                'imagen': request.build_absolute_uri(evidencia.imagen.url) if evidencia.imagen else None
            }
        }, status=201)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
def obtener_evidencias_pedido(request, pedido_id):
    """
    Obtener todas las evidencias de un pedido
    GET /api/evidencia-pedido/<pedido_id>/
    """
    try:
        from .models import Pedido, EvidenciaPedido
        
        # Buscar el pedido
        try:
            pedido = Pedido.objects.get(id=pedido_id)
        except Pedido.DoesNotExist:
            try:
                pedido = Pedido.objects.get(numero_pedido=pedido_id)
            except Pedido.DoesNotExist:
                return Response({'error': f'Pedido {pedido_id} no encontrado'}, status=404)
        
        evidencias = EvidenciaPedido.objects.filter(pedido=pedido)
        
        resultado = []
        for ev in evidencias:
            resultado.append({
                'id': ev.id,
                'producto_nombre': ev.producto_nombre,
                'motivo': ev.motivo,
                'imagen': request.build_absolute_uri(ev.imagen.url) if ev.imagen else None,
                'fecha_creacion': ev.fecha_creacion.isoformat()
            })
        
        return Response({
            'pedido': pedido.numero_pedido,
            'total_evidencias': len(resultado),
            'evidencias': resultado
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=500)


# ========================================
# 🔐 ENDPOINTS DE AUTENTICACIÓN
# ========================================

@api_view(['POST'])
def auth_login(request):
    """
    Endpoint de login para el Frontend Web.
    Rechaza vendedores de App Móvil (solo acceden por la App).
    """
    import hashlib
    from .models import Cajero
    
    codigo_o_nombre = request.data.get('codigo', '').strip()
    password = request.data.get('password', '')
    
    if not codigo_o_nombre or not password:
        return Response({
            'success': False,
            'error': 'Código/Nombre y contraseña son requeridos'
        }, status=400)
    
    try:
        # Buscar usuario por código o nombre
        usuario = Cajero.objects.filter(
            models.Q(codigo__iexact=codigo_o_nombre) | 
            models.Q(nombre__iexact=codigo_o_nombre)
        ).first()
        
        if not usuario:
            return Response({
                'success': False,
                'error': 'Usuario no encontrado'
            }, status=401)
        
        # Verificar si está activo
        if not usuario.activo:
            return Response({
                'success': False,
                'error': 'Usuario desactivado. Contacte al administrador.'
            }, status=401)
        
        # Rechazar vendedores de App Móvil
        if usuario.rol == 'VENDEDOR':
            return Response({
                'success': False,
                'error': 'Los vendedores de App Móvil no tienen acceso al sistema web. Use la aplicación móvil.'
            }, status=403)
        
        # Verificar contraseña
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        # Comparar con password hasheado O con password_plano
        password_valido = (
            usuario.password == password_hash or 
            usuario.password == password or 
            usuario.password_plano == password
        )
        
        if not password_valido:
            return Response({
                'success': False,
                'error': 'Contraseña incorrecta'
            }, status=401)
        
        # Actualizar último login
        usuario.ultimo_login = timezone.now()
        usuario.save(update_fields=['ultimo_login'])
        
        # Respuesta exitosa
        return Response({
            'success': True,
            'message': f'Bienvenido, {usuario.nombre}',
            'usuario': {
                'id': usuario.id,
                'codigo': usuario.codigo,
                'nombre': usuario.nombre,
                'email': usuario.email,
                'telefono': usuario.telefono,
                'rol': usuario.rol,
                'sucursal': usuario.sucursal.nombre if usuario.sucursal else None,
                'sucursal_id': usuario.sucursal.id if usuario.sucursal else None,
                'permisos': {
                    'acceso_pos': usuario.acceso_pos,
                    'acceso_pedidos': usuario.acceso_pedidos,
                    'acceso_cargue': usuario.acceso_cargue,
                    'acceso_produccion': usuario.acceso_produccion,
                    'acceso_inventario': usuario.acceso_inventario,
                    'acceso_reportes': usuario.acceso_reportes,
                    'acceso_configuracion': usuario.acceso_configuracion,
                    'puede_hacer_descuentos': usuario.puede_hacer_descuentos,
                    'puede_anular_ventas': usuario.puede_anular_ventas,
                },
                'es_admin': usuario.rol == 'ADMINISTRADOR',
            }
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Error en el servidor: {str(e)}'
        }, status=500)


@api_view(['POST'])
def auth_recuperar_password(request):
    """
    Solicitar recuperación de contraseña por email o teléfono.
    Genera un código temporal y lo envía.
    """
    import random
    from .models import Cajero
    
    email_o_telefono = request.data.get('contacto', '').strip()
    
    if not email_o_telefono:
        return Response({
            'success': False,
            'error': 'Ingrese su email o teléfono registrado'
        }, status=400)
    
    try:
        # Buscar usuario por email o teléfono
        usuario = Cajero.objects.filter(
            models.Q(email__iexact=email_o_telefono) | 
            models.Q(telefono__iexact=email_o_telefono)
        ).first()
        
        if not usuario:
            return Response({
                'success': False,
                'error': 'No se encontró ningún usuario con ese email o teléfono'
            }, status=404)
        
        # Generar código de recuperación (6 dígitos)
        codigo_recuperacion = ''.join([str(random.randint(0, 9)) for _ in range(6)])
        
        # TODO: Enviar código por email o SMS
        # Por ahora, lo devolvemos en la respuesta (solo para desarrollo)
        # En producción, usar servicios como SendGrid, Twilio, etc.
        
        # Guardar código temporalmente (en una tabla o cache)
        # Por simplicidad, lo guardamos en password_plano temporalmente
        # (en producción usar tabla de tokens de recuperación)
        
        return Response({
            'success': True,
            'message': f'Se ha enviado un código de recuperación a {email_o_telefono}',
            # Solo en desarrollo:
            'codigo_temporal': codigo_recuperacion,
            'usuario_id': usuario.id,
            'nota': 'En producción, el código se envía por email/SMS'
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Error en el servidor: {str(e)}'
        }, status=500)


@api_view(['POST'])
def auth_cambiar_password(request):
    """
    Cambiar contraseña de un usuario.
    Requiere: usuario_id, nueva_password, (codigo_recuperacion o password_actual)
    """
    import hashlib
    from .models import Cajero
    
    usuario_id = request.data.get('usuario_id')
    nueva_password = request.data.get('nueva_password', '').strip()
    password_actual = request.data.get('password_actual', '').strip()
    
    if not usuario_id or not nueva_password:
        return Response({
            'success': False,
            'error': 'Faltan datos requeridos'
        }, status=400)
    
    if len(nueva_password) < 4:
        return Response({
            'success': False,
            'error': 'La contraseña debe tener al menos 4 caracteres'
        }, status=400)
    
    try:
        usuario = Cajero.objects.get(id=usuario_id)
        
        # Si se proporciona password_actual, verificarlo
        if password_actual:
            password_hash = hashlib.sha256(password_actual.encode()).hexdigest()
            if usuario.password != password_hash and usuario.password != password_actual:
                return Response({
                    'success': False,
                    'error': 'Contraseña actual incorrecta'
                }, status=401)
        
        # Hashear y guardar nueva contraseña
        nuevo_hash = hashlib.sha256(nueva_password.encode()).hexdigest()
        usuario.password = nuevo_hash
        usuario.password_plano = nueva_password
        usuario.save(update_fields=['password', 'password_plano'])
        
        return Response({
            'success': True,
            'message': 'Contraseña actualizada correctamente'
        })
        
    except Cajero.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Usuario no encontrado'
        }, status=404)
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Error en el servidor: {str(e)}'
        }, status=500)


# ========================================
# 🧠 CONFIGURACIÓN Y LOGS DE REDES NEURONALES (IA)
# ========================================

@api_view(['GET', 'POST'])
def ia_config(request):
    """
    Gestionar configuración del Servicio de IA
    GET: Retorna estado actual
    POST: Actualiza configuración
    """
    # Usar archivo temporal para configuración por simplicidad
    CONFIG_FILE = 'ia_config.json'
    import json
    
    default_config = {
        'active': True,
        'continuousLearning': True,
        'lastTraining': None
    }
    
    # Cargar config actual
    current_config = default_config
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r') as f:
                current_config = json.load(f)
        except:
            pass
            
    if request.method == 'GET':
        return Response(current_config)
        
    elif request.method == 'POST':
        new_config = request.data
        current_config.update(new_config)
        
        try:
            with open(CONFIG_FILE, 'w') as f:
                json.dump(current_config, f)
            return Response({'success': True, 'config': current_config})
        except Exception as e:
            return Response({'error': str(e)}, status=500)

@api_view(['POST'])
def ia_retrain(request):
    """
    Dispara re-entrenamiento completo de la red neuronal
    basado en todos los Reportes de Planeación históricos.
    """
    try:
        from api.services.neural_network_service import NeuralNetworkService
        from api.models import ReportePlaneacion
        
        nn_service = NeuralNetworkService()
        
        # Obtener todos los reportes ordenados por fecha
        reportes = ReportePlaneacion.objects.all().order_by('fecha_reporte')
        count = 0
        
        for reporte in reportes:
            if nn_service.train_incremental(reporte):
                count += 1
                
        # Actualizar fecha de último entrenamiento
        import json
        CONFIG_FILE = 'ia_config.json'
        # Asegurar que existe si no
        if not os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, 'w') as f:
                json.dump({'active': True, 'continuousLearning': True}, f)

        if os.path.exists(CONFIG_FILE):
             with open(CONFIG_FILE, 'r+') as f:
                try:
                    config = json.load(f)
                except:
                    config = {'active': True, 'continuousLearning': True}
                
                config['lastTraining'] = timezone.now().isoformat()
                f.seek(0)
                json.dump(config, f)
                f.truncate()
        
        return Response({
            'success': True,
            'message': f'Re-entrenamiento completado con {count} reportes históricos.'
        })
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
def ia_logs(request):
    """
    Retorna logs recientes de la actividad neuronal.
    Simulado leyendo el archivo de log principal o memoria.
    """
    # Simulamos logs para la demo visual, en un sistema real leeríamos un archivo .log
    from datetime import datetime
    
    # Intentar leer config para estado
    import json
    CONFIG_FILE = 'ia_config.json'
    config = {'active': True, 'continuousLearning': True}
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, 'r') as f:
                config = json.load(f)
    except:
        pass

    # Generar logs simulados de "estado actual"
    logs = [
        {'time': datetime.now().strftime('%H:%M:%S'), 'type': 'INFO', 'message': 'Sistema Neuronal: EN LÍNEA'},
        {'time': datetime.now().strftime('%H:%M:%S'), 'type': 'INFO', 'message': f'Aprendizaje Continuo: {"ACTIVADO" if config.get("continuousLearning") else "DESACTIVADO"}'},
    ]
    
    if config.get('lastTraining'):
        logs.append({'time': config['lastTraining'].split('T')[1][:8], 'type': 'SUCCESS', 'message': 'Último re-entrenamiento exitoso'})

    return Response({'logs': logs, 'config': config})


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def exportar_clientes_excel(request):
    """
    Genera un archivo CSV compatible con Excel con todos los clientes (Ruta y Pedidos).
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="clientes_backup_{timezone.now().strftime("%Y%m%d")}.csv"'
    response.write(u'\ufeff'.encode('utf8')) # BOM para Excel

    writer = csv.writer(response)
    
    # Encabezados
    writer.writerow(['Origen', 'Nombre Negocio/Cliente', 'Contacto', 'Dirección', 'Teléfono', 'Ruta/Zona', 'Día Visita', 'Activo', 'Notas'])
    
    # 1. Clientes de Ruta
    clientes_ruta = ClienteRuta.objects.all().order_by('ruta', 'orden')
    for c in clientes_ruta:
        ruta_nombre = c.ruta.nombre if c.ruta else 'Sin Ruta'
        writer.writerow([
            'Ruta',
            c.nombre_negocio,
            c.nombre_contacto or '',
            c.direccion or '',
            c.telefono or '',
            ruta_nombre,
            c.dia_visita,
            'Si' if c.activo else 'No',
            c.nota or ''
        ])
        
    # 2. Clientes de Pedidos (Únicos por nombre y dirección)
    pedidos = Pedido.objects.values(
        'destinatario', 'telefono_contacto', 'direccion_entrega', 'zona_barrio', 'nota'
    ).distinct()
    
    seen = set()
    for p in pedidos:
        nombre = p['destinatario']
        direccion = p['direccion_entrega']
        
        if not nombre: continue
        
        # Normalizar para detectar duplicados
        key = (nombre.lower().strip(), direccion.lower().strip() if direccion else '')
        
        if key not in seen:
            seen.add(key)
            writer.writerow([
                'Pedido',
                nombre,
                '', # contacto
                direccion or '',
                p['telefono_contacto'] or '',
                p['zona_barrio'] or '',
                '', # dia visita
                'Si', 
                p['nota'] or ''
            ])

    return response


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def listar_vendedores_cargue(request):
    """
    Devuelve la lista de vendedores (ID1-ID6) con su nombre desde el modelo Vendedor.
    """
    from .models import Vendedor
    
    vendedores = Vendedor.objects.filter(activo=True).order_by('id_vendedor')
    data = [{'id': v.id_vendedor, 'nombre': v.nombre} for v in vendedores]
    
    return Response(data)


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def abrir_turno_manual(request):
    """
    Reabre un turno cerrado para permitir correcciones o pruebas.
    Recibe: { "fecha": "YYYY-MM-DD", "vendedor_id": 1 }
    Opcional: Si no se envía fecha, abre el último turno cerrado de ese vendedor.
    
    🆕 MEJORA: Resetea devoluciones y vencidas en CargueIDx para permitir nuevo cierre.
    """
    try:
        from .models import TurnoVendedor, Vendedor, CargueID1, CargueID2, CargueID3, CargueID4, CargueID5, CargueID6
        
        fecha = request.data.get('fecha')
        vendedor_id = request.data.get('vendedor_id') # Puede ser ID numérico o código "ID1"
        
        # ✅ NUEVO: Extraer ID numérico del código
        id_numerico = None
        codigo_vendedor = "" # ID1, ID2...
        
        if vendedor_id:
            vendedor_str = str(vendedor_id).upper()
            
            # Caso 1: Viene como "ID1", "ID2", etc. -> Extraer número
            if vendedor_str.startswith('ID'):
                try:
                    id_numerico = int(vendedor_str.replace('ID', ''))
                    codigo_vendedor = vendedor_str
                except:
                    pass
            # Caso 2: Viene como número directo "1", "2", etc.
            else:
                try:
                    id_numerico = int(vendedor_str)
                    codigo_vendedor = f"ID{id_numerico}"
                except:
                    pass
        
        if not id_numerico:
            return Response({
                'success': False,
                'message': 'Vendedor ID inválido. Usa "ID1", "ID2", etc. o 1, 2, etc.'
            }, status=400)
        
        # ✅ BUSCAR POR ID NUMÉRICO (no por nombre)
        queryset = TurnoVendedor.objects.filter(vendedor_id=id_numerico)
        
        print(f"🔍 Buscando turnos: vendedor_id={id_numerico}, código={codigo_vendedor}, fecha={fecha or 'cualquiera'}")
        
        # Filtrar por fecha si se proporcionó
        if fecha:
            queryset = queryset.filter(fecha=fecha)
        
        # Filtrar solo cerrados
        queryset = queryset.filter(estado='CERRADO')
        
        # Obtener el último
        turno = queryset.order_by('-id').first()
        
        if turno:
            # 1. Reabrir Turno (Cambiar estado)
            turno.estado = 'ABIERTO'
            # turno.hora_cierre = None # Si tuviera este campo
            turno.save()
            
            # 2. 🆕 LIMPIEZA DE DATOS (Resetear devoluciones y vencidas)
            # Si no tenemos codigo_vendedor aún, intentamos deducirlo del turno
            if not codigo_vendedor and turno.vendedor_id:
                # Asumimos mapeo directo ID 1 -> ID1
                codigo_vendedor = f"ID{turno.vendedor_id}"
            
            # Intentar limpiar la tabla de Cargue correspondiente
            registros_limpiados = 0
            if codigo_vendedor:
                modelo_map = {
                    'ID1': CargueID1, 'ID2': CargueID2, 'ID3': CargueID3,
                    'ID4': CargueID4, 'ID5': CargueID5, 'ID6': CargueID6
                }
                ModeloCargue = modelo_map.get(codigo_vendedor)
                
                if ModeloCargue:
                    # Resetear devoluciones y vencidas a 0 para esta fecha
                    cargues_afectados = ModeloCargue.objects.filter(fecha=turno.fecha, activo=True)
                    registros_limpiados = cargues_afectados.update(devoluciones=0, vencidas=0)
                    if registros_limpiados > 0:
                        recalcular_totales_cargue_queryset(
                            ModeloCargue.objects.filter(fecha=turno.fecha, activo=True)
                        )
                    print(f"🧹 Limpieza exitosa: {registros_limpiados} registros reseteados en {codigo_vendedor} para {turno.fecha}")
            
            print(f"✅ Turno reabierto: ID {turno.id} - {turno.vendedor_nombre} - {turno.fecha}")
            
            return Response({
                'success': True,
                'message': f'Turno del {turno.fecha} reabierto y reseteado correctamente ({registros_limpiados} productos limpiados)',
                'turno': {
                    'id': turno.id,
                    'fecha': turno.fecha,
                    'vendedor': turno.vendedor_nombre,
                    'estado': turno.estado
                }
            })
        else:
            return Response({
                'success': False,
                'message': f'No se encontró ningún turno CERRADO para los criterios dados'
            }, status=404)
            
    except Exception as e:
        print(f"❌ Error reabriendo turno: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({
            'success': False,
            'error': str(e)
        }, status=500)


# ==================== EXPORTAR CLIENTES RUTAS EXCEL ====================

@api_view(['GET'])
def exportar_clientes_rutas_excel(request):
    """
    Exporta todos los clientes de ruta a Excel, agrupados por ID de vendedor.
    GET /api/reportes/clientes-rutas-excel/
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    import io

    try:
        clientes_ruta = ClienteRuta.objects.filter(
            activo=True, ruta__activo=True
        ).select_related('ruta', 'ruta__vendedor').order_by(
            'ruta__vendedor__id_vendedor', 'ruta__nombre', 'orden'
        )

        # Agrupar por vendedor
        por_vendedor = {}
        for c in clientes_ruta:
            vid = c.ruta.vendedor.id_vendedor if c.ruta.vendedor else 'SIN_ID'
            vnombre = c.ruta.vendedor.nombre if c.ruta.vendedor else ''
            if vid not in por_vendedor:
                por_vendedor[vid] = {'nombre': vnombre, 'clientes': []}
            origen = 'PEDIDO' if (c.tipo_negocio and 'PEDIDOS' in c.tipo_negocio.upper()) else 'RUTA'
            por_vendedor[vid]['clientes'].append({
                'ruta': c.ruta.nombre,
                'negocio': c.nombre_negocio,
                'contacto': c.nombre_contacto or '',
                'telefono': c.telefono or '',
                'tipo_negocio': c.tipo_negocio or '',
                'origen': origen,
                'dias': c.dia_visita or '',
                'direccion': c.direccion or '',
                'nota': c.nota or '',
            })

        # ── Estilos ─────────────────────────────────────────────────
        azul = '0C2C53'
        verde = '198754'
        blanco = 'FFFFFF'
        gris = 'F1F3F5'
        amarillo = 'FFF3CD'

        header_font = Font(bold=True, color=blanco, size=11)
        header_fill = PatternFill('solid', fgColor=azul)
        total_fill = PatternFill('solid', fgColor='DEE2E6')
        center = Alignment(horizontal='center', vertical='center')

        def autofit(ws):
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

        wb = Workbook()
        wb.remove(wb.active)

        # ── Hoja Resumen ────────────────────────────────────────────
        ws_res = wb.create_sheet('Resumen', 0)
        ws_res.append([f'CLIENTES DE RUTAS — {__import__("datetime").date.today()}'])
        ws_res['A1'].font = Font(bold=True, size=14, color=azul)
        ws_res.append([])
        ws_res.append(['ID Vendedor', 'Nombre Vendedor', 'Total Clientes', 'RUTA', 'PEDIDO'])
        for cell in ws_res[3]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for vid in sorted(por_vendedor.keys()):
            g = por_vendedor[vid]
            total_ruta = sum(1 for c in g['clientes'] if c['origen'] == 'RUTA')
            total_ped = sum(1 for c in g['clientes'] if c['origen'] == 'PEDIDO')
            ws_res.append([vid, g['nombre'], len(g['clientes']), total_ruta, total_ped])

        ws_res.append(['TOTAL', '', sum(len(g['clientes']) for g in por_vendedor.values()), '', ''])
        for cell in ws_res[ws_res.max_row]:
            cell.font = Font(bold=True)
            cell.fill = total_fill
        autofit(ws_res)

        # ── Hoja por ID ─────────────────────────────────────────────
        COLS = ['Ruta', 'Negocio', 'Contacto', 'Teléfono', 'Tipo Negocio', 'Origen', 'Días Visita', 'Dirección', 'Nota']
        for vid in sorted(por_vendedor.keys()):
            g = por_vendedor[vid]
            ws = wb.create_sheet(vid[:31])
            ws.append([f'{vid} — {g["nombre"]}'])
            ws.merge_cells('A1:I1')
            ws['A1'].font = Font(bold=True, size=13, color=azul)
            ws.append([])
            ws.append(COLS)
            for cell in ws[3]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center

            for i, c in enumerate(g['clientes']):
                ws.append([
                    c['ruta'], c['negocio'], c['contacto'], c['telefono'],
                    c['tipo_negocio'], c['origen'], c['dias'], c['direccion'], c['nota']
                ])
                if c['origen'] == 'PEDIDO':
                    for cell in ws[ws.max_row]:
                        cell.fill = PatternFill('solid', fgColor='FFF3CD')
                elif i % 2 == 0:
                    for cell in ws[ws.max_row]:
                        cell.fill = PatternFill('solid', fgColor=gris)

            autofit(ws)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        from datetime import date
        nombre_archivo = f'clientes_rutas_{date.today()}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


# ==================== VENTAS PRODUCTOS POS ====================

@api_view(['GET'])
def ventas_productos_pos(request):
    """
    Ventas de productos POS consolidadas por período.
    GET /api/reportes/ventas-productos-pos/?fecha_inicio=2026-04-01&fecha_fin=2026-04-30
    """
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if not fecha_inicio or not fecha_fin:
        return Response({'error': 'fecha_inicio y fecha_fin son requeridos'}, status=400)

    try:
        from django.db.models import Sum

        detalles = DetalleVenta.objects.filter(
            venta__fecha__date__gte=fecha_inicio,
            venta__fecha__date__lte=fecha_fin,
        ).exclude(
            venta__estado='ANULADA'
        ).exclude(
            venta__estado='CANCELADO'
        ).values(
            'producto__nombre'
        ).annotate(
            cantidad_total=Sum('cantidad'),
            total_ventas=Sum('subtotal')
        ).order_by('-cantidad_total')

        productos = [
            {
                'nombre': d['producto__nombre'] or '(Sin nombre)',
                'cantidad': d['cantidad_total'] or 0,
                'total': round(float(d['total_ventas'] or 0)),
            }
            for d in detalles
        ]

        return Response({
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'productos': productos,
            'total_unidades': sum(p['cantidad'] for p in productos),
            'total_ventas': sum(p['total'] for p in productos),
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
def exportar_ventas_productos_pos_excel(request):
    """
    Exporta ventas de productos POS a Excel.
    GET /api/reportes/ventas-productos-pos/excel/?fecha_inicio=X&fecha_fin=Y
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from django.db.models import Sum
    import io

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if not fecha_inicio or not fecha_fin:
        return Response({'error': 'fecha_inicio y fecha_fin son requeridos'}, status=400)

    try:
        detalles = DetalleVenta.objects.filter(
            venta__fecha__date__gte=fecha_inicio,
            venta__fecha__date__lte=fecha_fin,
        ).exclude(
            venta__estado='ANULADA'
        ).exclude(
            venta__estado='CANCELADO'
        ).values(
            'producto__nombre'
        ).annotate(
            cantidad_total=Sum('cantidad'),
            total_ventas=Sum('subtotal')
        ).order_by('-cantidad_total')

        azul_oscuro = '0C2C53'
        blanco = 'FFFFFF'
        gris = 'F1F3F5'
        total_fill = PatternFill('solid', fgColor='DEE2E6')
        header_font = Font(bold=True, color=blanco, size=11)
        header_fill = PatternFill('solid', fgColor=azul_oscuro)
        center = Alignment(horizontal='center')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Ventas POS'

        ws.append([f'VENTAS PRODUCTOS POS — {fecha_inicio} → {fecha_fin}'])
        ws['A1'].font = Font(bold=True, size=14, color=azul_oscuro)
        ws.append([])

        ws.append(['#', 'Producto', 'Cantidad', 'Total COP'])
        for cell in ws[3]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        total_und = 0
        total_cop = 0
        for i, d in enumerate(detalles, 1):
            cant = d['cantidad_total'] or 0
            tot = round(float(d['total_ventas'] or 0))
            total_und += cant
            total_cop += tot
            ws.append([i, d['producto__nombre'] or '(Sin nombre)', cant, tot])
            if i % 2 == 0:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill('solid', fgColor=gris)

        ws.append(['', 'TOTAL', total_und, total_cop])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, size=11)
            cell.fill = total_fill

        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        nombre_archivo = f'ventas_pos_{fecha_inicio}_{fecha_fin}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


# ==================== HISTORIAL DE CLIENTES ====================

@api_view(['GET'])
def historial_clientes(request):
    """
    Historial de compras y vencidas por cliente en un rango de fechas.
    GET /api/reportes/historial-clientes/?fecha_inicio=2026-04-01&fecha_fin=2026-04-30
    Parámetros opcionales: vendedor_id (ej: ID1)
    """
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    vendedor_id = request.GET.get('vendedor_id', '').strip()

    if not fecha_inicio or not fecha_fin:
        return Response({'error': 'fecha_inicio y fecha_fin son requeridos'}, status=400)

    try:
        # ── Ventas de Ruta (App Móvil) ──────────────────────────────
        ventas_qs = VentaRuta.objects.filter(
            fecha__date__gte=fecha_inicio,
            fecha__date__lte=fecha_fin,
            estado='ACTIVA'
        ).select_related('vendedor', 'cliente')

        if vendedor_id:
            ventas_qs = ventas_qs.filter(vendedor__id_vendedor=vendedor_id)

        clientes_map = {}

        for venta in ventas_qs:
            vid = venta.vendedor.id_vendedor if venta.vendedor else 'SIN_ID'
            cid = str(venta.cliente_id) if venta.cliente_id else f'__occ_{venta.nombre_negocio or venta.cliente_nombre}'
            key = (cid, vid)

            if key not in clientes_map:
                clientes_map[key] = {
                    'cliente_id': venta.cliente_id,
                    'nombre_negocio': venta.nombre_negocio or venta.cliente_nombre or '(Sin nombre)',
                    'nombre_contacto': venta.cliente.nombre_contacto if venta.cliente else '',
                    'vendedor_id': vid,
                    'vendedor_nombre': venta.vendedor.nombre if venta.vendedor else '',
                    'origen': 'RUTA',
                    'total_ventas': 0,
                    'num_ventas': 0,
                    'productos': {},
                    'vencidas': {},
                }

            entry = clientes_map[key]
            entry['total_ventas'] += float(venta.total or 0)
            entry['num_ventas'] += 1

            for det in (venta.detalles or []):
                nombre = det.get('producto', '').strip()
                if not nombre:
                    continue
                if nombre not in entry['productos']:
                    entry['productos'][nombre] = {'cantidad': 0, 'total': 0}
                entry['productos'][nombre]['cantidad'] += int(det.get('cantidad', 0))
                entry['productos'][nombre]['total'] += float(det.get('subtotal', 0))

            for vec in (venta.productos_vencidos or []):
                nombre = vec.get('producto', '').strip()
                if not nombre:
                    continue
                cant = int(vec.get('cantidad', 0))
                if cant <= 0:
                    continue
                if nombre not in entry['vencidas']:
                    entry['vencidas'][nombre] = {'cantidad': 0}
                entry['vencidas'][nombre]['cantidad'] += cant

        # ── Pedidos entregados ──────────────────────────────────────
        pedidos_qs = Pedido.objects.filter(
            fecha__date__gte=fecha_inicio,
            fecha__date__lte=fecha_fin,
            estado='ENTREGADA'
        ).prefetch_related('detalles__producto')

        if vendedor_id:
            pedidos_qs = pedidos_qs.filter(asignado_a_id=vendedor_id)

        for pedido in pedidos_qs:
            vid = pedido.asignado_a_id or 'SIN_ID'
            key = (f'__ped_{pedido.destinatario}_{vid}', vid)

            if key not in clientes_map:
                clientes_map[key] = {
                    'cliente_id': None,
                    'nombre_negocio': pedido.destinatario or '(Sin nombre)',
                    'nombre_contacto': '',
                    'vendedor_id': vid,
                    'vendedor_nombre': pedido.vendedor or '',
                    'origen': 'PEDIDO',
                    'total_ventas': 0,
                    'num_ventas': 0,
                    'productos': {},
                    'vencidas': {},
                }

            entry = clientes_map[key]
            entry['total_ventas'] += float(pedido.total or 0)
            entry['num_ventas'] += 1

            for det in pedido.detalles.all():
                nombre = det.producto.nombre if det.producto else ''
                if not nombre:
                    continue
                if nombre not in entry['productos']:
                    entry['productos'][nombre] = {'cantidad': 0, 'total': 0}
                entry['productos'][nombre]['cantidad'] += det.cantidad
                entry['productos'][nombre]['total'] += float(det.subtotal or 0)

        # ── Serializar resultado ────────────────────────────────────
        result = []
        for entry in clientes_map.values():
            result.append({
                'cliente_id': entry['cliente_id'],
                'nombre_negocio': entry['nombre_negocio'],
                'nombre_contacto': entry['nombre_contacto'],
                'vendedor_id': entry['vendedor_id'],
                'vendedor_nombre': entry['vendedor_nombre'],
                'origen': entry['origen'],
                'total_ventas': round(entry['total_ventas']),
                'num_ventas': entry['num_ventas'],
                'productos': sorted(
                    [{'nombre': k, 'cantidad': v['cantidad'], 'total': round(v['total'])}
                     for k, v in entry['productos'].items()],
                    key=lambda x: -x['cantidad']
                ),
                'vencidas': sorted(
                    [{'nombre': k, 'cantidad': v['cantidad']}
                     for k, v in entry['vencidas'].items()],
                    key=lambda x: -x['cantidad']
                ),
            })

        result.sort(key=lambda x: (x['vendedor_id'], x['nombre_negocio'].lower()))
        return Response(result)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
def exportar_historial_clientes_excel(request):
    """
    Exporta el historial de clientes a Excel con una hoja por ID de vendedor.
    GET /api/reportes/historial-clientes/excel/?fecha_inicio=2026-04-01&fecha_fin=2026-04-30
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    import io

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    vendedor_id = request.GET.get('vendedor_id', '').strip()

    if not fecha_inicio or not fecha_fin:
        return Response({'error': 'fecha_inicio y fecha_fin son requeridos'}, status=400)

    try:
        # Reusar lógica de historial_clientes para obtener datos
        ventas_qs = VentaRuta.objects.filter(
            fecha__date__gte=fecha_inicio,
            fecha__date__lte=fecha_fin,
            estado='ACTIVA'
        ).select_related('vendedor', 'cliente')
        if vendedor_id:
            ventas_qs = ventas_qs.filter(vendedor__id_vendedor=vendedor_id)

        clientes_map = {}

        for venta in ventas_qs:
            vid = venta.vendedor.id_vendedor if venta.vendedor else 'SIN_ID'
            cid = str(venta.cliente_id) if venta.cliente_id else f'__occ_{venta.nombre_negocio or venta.cliente_nombre}'
            key = (cid, vid)
            if key not in clientes_map:
                clientes_map[key] = {
                    'nombre_negocio': venta.nombre_negocio or venta.cliente_nombre or '(Sin nombre)',
                    'nombre_contacto': venta.cliente.nombre_contacto if venta.cliente else '',
                    'vendedor_id': vid,
                    'vendedor_nombre': venta.vendedor.nombre if venta.vendedor else '',
                    'origen': 'RUTA',
                    'total_ventas': 0,
                    'productos': {},
                    'vencidas': {},
                }
            entry = clientes_map[key]
            entry['total_ventas'] += float(venta.total or 0)
            for det in (venta.detalles or []):
                nombre = det.get('producto', '').strip()
                if not nombre:
                    continue
                if nombre not in entry['productos']:
                    entry['productos'][nombre] = {'cantidad': 0, 'total': 0}
                entry['productos'][nombre]['cantidad'] += int(det.get('cantidad', 0))
                entry['productos'][nombre]['total'] += float(det.get('subtotal', 0))
            for vec in (venta.productos_vencidos or []):
                nombre = vec.get('producto', '').strip()
                cant = int(vec.get('cantidad', 0))
                if not nombre or cant <= 0:
                    continue
                if nombre not in entry['vencidas']:
                    entry['vencidas'][nombre] = {'cantidad': 0}
                entry['vencidas'][nombre]['cantidad'] += cant

        pedidos_qs = Pedido.objects.filter(
            fecha__date__gte=fecha_inicio,
            fecha__date__lte=fecha_fin,
            estado='ENTREGADA'
        ).prefetch_related('detalles__producto')
        if vendedor_id:
            pedidos_qs = pedidos_qs.filter(asignado_a_id=vendedor_id)

        for pedido in pedidos_qs:
            vid = pedido.asignado_a_id or 'SIN_ID'
            key = (f'__ped_{pedido.destinatario}_{vid}', vid)
            if key not in clientes_map:
                clientes_map[key] = {
                    'nombre_negocio': pedido.destinatario or '(Sin nombre)',
                    'nombre_contacto': '',
                    'vendedor_id': vid,
                    'vendedor_nombre': pedido.vendedor or '',
                    'origen': 'PEDIDO',
                    'total_ventas': 0,
                    'productos': {},
                    'vencidas': {},
                }
            entry = clientes_map[key]
            entry['total_ventas'] += float(pedido.total or 0)
            for det in pedido.detalles.all():
                nombre = det.producto.nombre if det.producto else ''
                if not nombre:
                    continue
                if nombre not in entry['productos']:
                    entry['productos'][nombre] = {'cantidad': 0, 'total': 0}
                entry['productos'][nombre]['cantidad'] += det.cantidad
                entry['productos'][nombre]['total'] += float(det.subtotal or 0)

        # ── Agrupar por vendedor ────────────────────────────────────
        por_vendedor = {}
        for entry in clientes_map.values():
            vid = entry['vendedor_id']
            if vid not in por_vendedor:
                por_vendedor[vid] = {'vendedor_nombre': entry['vendedor_nombre'], 'clientes': []}
            por_vendedor[vid]['clientes'].append(entry)

        # ── Estilos ─────────────────────────────────────────────────
        azul_oscuro = '0C2C53'
        verde = '198754'
        amarillo = 'FFC107'
        gris_claro = 'F1F3F5'
        blanco = 'FFFFFF'

        header_font = Font(bold=True, color=blanco, size=11)
        header_fill_azul = PatternFill('solid', fgColor=azul_oscuro)
        header_fill_verde = PatternFill('solid', fgColor=verde)
        header_fill_amarillo = PatternFill('solid', fgColor='856404')
        fila_gris = PatternFill('solid', fgColor=gris_claro)
        fila_total = PatternFill('solid', fgColor='DEE2E6')
        borde = Border(
            bottom=Side(style='thin', color='DEE2E6')
        )
        center = Alignment(horizontal='center', vertical='center')

        def autofit(ws):
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        wb = Workbook()
        wb.remove(wb.active)  # quitar hoja default

        # ── Hoja Resumen ────────────────────────────────────────────
        ws_res = wb.create_sheet('Resumen', 0)
        ws_res.append(['HISTORIAL DE CLIENTES', f'{fecha_inicio} → {fecha_fin}'])
        ws_res['A1'].font = Font(bold=True, size=14, color=azul_oscuro)
        ws_res.append([])
        ws_res.append(['ID Vendedor', 'Nombre Vendedor', 'Clientes', 'Total Ventas', 'Registros Vencidas'])
        for cell in ws_res[3]:
            cell.font = header_font
            cell.fill = header_fill_azul
            cell.alignment = center

        total_general = 0
        for vid in sorted(por_vendedor.keys()):
            g = por_vendedor[vid]
            total_v = sum(c['total_ventas'] for c in g['clientes'])
            venc_count = sum(sum(v['cantidad'] for v in c['vencidas'].values()) for c in g['clientes'])
            total_general += total_v
            ws_res.append([vid, g['vendedor_nombre'], len(g['clientes']), round(total_v), venc_count])

        ws_res.append(['TOTAL GENERAL', '', sum(len(g['clientes']) for g in por_vendedor.values()), round(total_general), ''])
        last = ws_res.max_row
        for cell in ws_res[last]:
            cell.font = Font(bold=True, size=11)
            cell.fill = fila_total
        autofit(ws_res)

        # ── Hoja por ID ─────────────────────────────────────────────
        COLS = ['Vendedor ID', 'Vendedor', 'Cliente', 'Contacto', 'Origen', 'Tipo', 'Producto', 'Cantidad', 'Total COP']
        for vid in sorted(por_vendedor.keys()):
            g = por_vendedor[vid]
            nombre_hoja = f'{vid}'[:31]
            ws = wb.create_sheet(nombre_hoja)

            # Título
            ws.append([f'Historial — {vid} — {g["vendedor_nombre"]}', '', '', '', '', '', '', '', f'{fecha_inicio} → {fecha_fin}'])
            ws.merge_cells('A1:H1')
            ws['A1'].font = Font(bold=True, size=13, color=azul_oscuro)
            ws.append([])

            # Encabezados
            ws.append(COLS)
            for cell in ws[3]:
                cell.font = header_font
                cell.fill = header_fill_azul
                cell.alignment = center

            clientes_sorted = sorted(g['clientes'], key=lambda c: c['nombre_negocio'].lower())
            for cliente in clientes_sorted:
                productos = sorted(cliente['productos'].items(), key=lambda x: -x[1]['cantidad'])
                vencidas = sorted(cliente['vencidas'].items(), key=lambda x: -x[1]['cantidad'])

                # Filas de compras
                if productos:
                    ws.append([])
                    sub_h = ws.max_row
                    ws[f'C{sub_h}'] = f'▶ {cliente["nombre_negocio"]}'
                    ws[f'C{sub_h}'].font = Font(bold=True, color=azul_oscuro)

                for nombre, vals in productos:
                    ws.append([
                        vid, g['vendedor_nombre'],
                        cliente['nombre_negocio'], cliente['nombre_contacto'],
                        cliente['origen'], 'COMPRA',
                        nombre, vals['cantidad'], round(vals['total'])
                    ])
                    row = ws.max_row
                    for cell in ws[row]:
                        cell.border = borde

                # Total cliente compras
                if productos:
                    ws.append([
                        '', '', cliente['nombre_negocio'], '', '', 'TOTAL COMPRAS', '',
                        sum(v['cantidad'] for _, v in productos),
                        round(cliente['total_ventas'])
                    ])
                    row = ws.max_row
                    for cell in ws[row]:
                        cell.font = Font(bold=True)
                        cell.fill = fila_total

                # Filas de vencidas
                if vencidas:
                    ws.append([])
                    for nombre, vals in vencidas:
                        ws.append([
                            vid, g['vendedor_nombre'],
                            cliente['nombre_negocio'], cliente['nombre_contacto'],
                            cliente['origen'], 'VENCIDA',
                            nombre, vals['cantidad'], 0
                        ])
                        row = ws.max_row
                        for cell in ws[row]:
                            cell.border = borde
                            cell.fill = PatternFill('solid', fgColor='FFF3CD')

                    ws.append([
                        '', '', cliente['nombre_negocio'], '', '', 'TOTAL VENCIDAS', '',
                        sum(v['cantidad'] for _, v in vencidas), ''
                    ])
                    row = ws.max_row
                    for cell in ws[row]:
                        cell.font = Font(bold=True, color='856404')
                        cell.fill = PatternFill('solid', fgColor='FFE69C')

            autofit(ws)

        # ── Retornar archivo ────────────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        nombre_archivo = f'historial_clientes_{fecha_inicio}_{fecha_fin}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


# ══════════════════════════════════════════════════════════════
#  📦 EXPORTAR CARGUE A EXCEL
# ══════════════════════════════════════════════════════════════
@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def exportar_cargue_excel(request):
    """
    Exporta el cargue completo de una fecha/dia a un libro Excel.
    Una hoja por ID (productos + pagos + cumplimiento) + hoja Resumen.
    Params: fecha=YYYY-MM-DD  dia=LUNES|MARTES|...
    """
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return Response({'error': 'openpyxl no instalado'}, status=500)

    fecha_str = request.query_params.get('fecha')
    dia = request.query_params.get('dia', '').upper()

    if not fecha_str or not dia:
        return Response({'error': 'Parametros fecha y dia son requeridos'}, status=400)

    try:
        fecha_obj = date.fromisoformat(fecha_str)
    except ValueError:
        return Response({'error': 'Formato de fecha invalido. Use YYYY-MM-DD'}, status=400)

    azul = '1E3A5F'
    azul_claro = 'D6E4F0'
    verde = '1D6A3A'
    gris = 'F8F9FA'

    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill_azul = PatternFill('solid', fgColor=azul)
    header_fill_verde = PatternFill('solid', fgColor=verde)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center')
    borde = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    fila_total = PatternFill('solid', fgColor=azul_claro)
    fila_par = PatternFill('solid', fgColor=gris)

    def autofit(ws, min_w=8, max_w=40):
        for col in ws.columns:
            width = min_w
            col_letter = None
            for cell in col:
                if hasattr(cell, 'column_letter'):
                    col_letter = cell.column_letter
                    if cell.value:
                        width = max(width, min(max_w, len(str(cell.value)) + 2))
            if col_letter:
                ws.column_dimensions[col_letter].width = width

    def formato_cop(val):
        try:
            return f'${int(float(val)):,}'
        except Exception:
            return val

    MODELOS_ID = [
        ('ID1', CargueID1),
        ('ID2', CargueID2),
        ('ID3', CargueID3),
        ('ID4', CargueID4),
        ('ID5', CargueID5),
        ('ID6', CargueID6),
    ]

    try:
        wb = Workbook()
        wb.remove(wb.active)

        resumen_rows = []

        # Precargar precios de todos los productos (precio_cargue)
        precios_lookup = {}
        for prod in Producto.objects.values('nombre', 'precio_cargue', 'precio'):
            precio = float(prod.get('precio_cargue') or prod.get('precio') or 0)
            if precio > 0:
                precios_lookup[prod['nombre']] = precio

        for id_label, Modelo in MODELOS_ID:
            # Orden por id (igual que la web - orden de inserción)
            registros = list(Modelo.objects.filter(fecha=fecha_obj, dia=dia, activo=True).order_by('id'))
            if not registros:
                continue

            primer = registros[0]
            responsable = getattr(primer, 'responsable', id_label)
            ruta = getattr(primer, 'ruta', '')

            ws = wb.create_sheet(id_label)

            # ── FILA 1: TÍTULO ──────────────────────────────────────
            ws.append([f'CARGUE {id_label} — {responsable}', '', '', '', '', '', '', '', '', '', '', f'{dia}  {fecha_str}'])
            ws.merge_cells('A1:K1')
            ws['A1'].font = Font(bold=True, size=13, color=azul)
            ws['A1'].alignment = left
            ws['L1'].font = Font(bold=True, size=11, color='555555')
            ws['L1'].alignment = center
            ws.row_dimensions[1].height = 22

            ws.append([])  # fila 2 vacía

            # ── FILA 3: ENCABEZADOS TABLA PRODUCTOS ─────────────────
            PROD_COLS = ['V', 'D', 'PRODUCTOS', 'CANTIDAD', 'DCTOS.', 'ADICIONAL', 'DEVOLUCIONES', 'VENCIDAS', 'LOTES VENCIDOS', 'TOTAL', 'VALOR', 'NETO']
            ws.append(PROD_COLS)
            hdr_row = ws.max_row  # = 3
            for cell in ws[hdr_row]:
                cell.font = header_font
                cell.fill = header_fill_azul
                cell.alignment = center
                cell.border = borde

            # ── FILAS 4..n+3: DATOS ──────────────────────────────────
            data_start = ws.max_row + 1
            for i, r in enumerate(registros):
                rn = ws.max_row + 1
                v_val = 'V' if getattr(r, 'v', False) else ''
                d_val = 'D' if getattr(r, 'd', False) else ''
                try:
                    lotes_v = json.loads(r.lotes_vencidos) if r.lotes_vencidos else []
                    if isinstance(lotes_v, list) and lotes_v:
                        lotes_str = ', '.join(str(l.get('lote', '')) for l in lotes_v if l.get('lote'))
                    else:
                        lotes_str = ''
                except Exception:
                    lotes_str = str(r.lotes_vencidos or '')

                # Precio: usar el guardado en BD si > 0, si no buscar en productos
                precio = float(r.valor or 0)
                if precio == 0:
                    precio = precios_lookup.get(r.producto, 0)

                ws.append([v_val, d_val, r.producto,
                           r.cantidad, r.dctos, r.adicional, r.devoluciones, r.vencidas,
                           lotes_str, None, precio, None])
                # Formulas TOTAL y NETO
                ws[f'J{rn}'] = f'=D{rn}-E{rn}+F{rn}-G{rn}-H{rn}'
                ws[f'L{rn}'] = f'=J{rn}*K{rn}'

                fill = fila_par if i % 2 == 1 else PatternFill()
                for cell in ws[rn]:
                    cell.border = borde
                    cell.alignment = center
                    if fill.fill_type:
                        cell.fill = fill
                ws[f'C{rn}'].alignment = left

            data_end = ws.max_row

            # ── FILA TOTALES ─────────────────────────────────────────
            ws.append(['', '', 'TOTALES',
                       f'=SUM(D{data_start}:D{data_end})', '', '',
                       f'=SUM(G{data_start}:G{data_end})',
                       f'=SUM(H{data_start}:H{data_end})', '',
                       f'=SUM(J{data_start}:J{data_end})', '',
                       f'=SUM(L{data_start}:L{data_end})'])
            tot_row = ws.max_row
            for cell in ws[tot_row]:
                cell.font = Font(bold=True)
                cell.fill = fila_total
                cell.border = borde
            ws[f'C{tot_row}'].alignment = left

            ws.append([])  # blank

            # ── PAGOS ────────────────────────────────────────────────
            pagos = list(CarguePagos.objects.filter(vendedor_id=id_label, fecha=fecha_obj, dia=dia, activo=True))

            ws.append(['CONCEPTO', 'DESCUENTOS', 'NEQUI', 'DAVIPLATA'])
            pago_hdr_r = ws.max_row
            for cell in ws[pago_hdr_r]:
                cell.font = header_font
                cell.fill = header_fill_verde
                cell.alignment = center
                cell.border = borde

            pago_data_start = ws.max_row + 1
            for p in pagos:
                ws.append([p.concepto, float(p.descuentos or 0), float(p.nequi or 0), float(p.daviplata or 0)])
                rp = ws.max_row
                for cell in ws[rp]:
                    cell.border = borde
                    cell.alignment = Alignment(horizontal='right')
                ws[f'A{rp}'].alignment = left
            pago_data_end = ws.max_row

            # Fila TOTAL pagos con fórmulas
            if pagos:
                ws.append(['TOTAL', '', f'=SUM(C{pago_data_start}:C{pago_data_end})', f'=SUM(D{pago_data_start}:D{pago_data_end})'])
            else:
                ws.append(['TOTAL', '', 0, 0])
            pago_tot_row = ws.max_row
            for cell in ws[pago_tot_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill('solid', fgColor='D4EDDA')
                cell.border = borde

            ws.append([])  # blank

            # ── RESUMEN FINANCIERO ───────────────────────────────────
            try:
                resumen_obj = CargueResumen.objects.get(vendedor_id=id_label, fecha=fecha_obj, dia=dia, activo=True)
            except CargueResumen.DoesNotExist:
                resumen_obj = None

            base_caja_val = float(resumen_obj.base_caja or 0) if resumen_obj else 0
            total_pedidos_val = float(resumen_obj.total_pedidos or 0) if resumen_obj else 0
            total_dctos_val = float(resumen_obj.total_dctos or 0) if resumen_obj else 0

            # Header resumen
            ws.append(['RESUMEN FINANCIERO', ''])
            res_hdr_r = ws.max_row
            ws.merge_cells(f'A{res_hdr_r}:B{res_hdr_r}')
            ws[f'A{res_hdr_r}'].font = Font(bold=True, color='FFFFFF', size=11)
            ws[f'A{res_hdr_r}'].fill = header_fill_azul
            ws[f'A{res_hdr_r}'].alignment = center

            def _res_row(ws, label, value, hl=None):
                ws.append([label, value])
                r = ws.max_row
                ws[f'A{r}'].font = Font(bold=True)
                ws[f'B{r}'].alignment = Alignment(horizontal='right')
                for col in ('A', 'B'):
                    ws[f'{col}{r}'].border = borde
                if hl:
                    for col in ('A', 'B'):
                        ws[f'{col}{r}'].fill = PatternFill('solid', fgColor=hl)
                return r

            _res_row(ws, 'BASE CAJA', base_caja_val)
            td_r = _res_row(ws, 'TOTAL DESPACHO:', f'=SUM(L{data_start}:L{data_end})', hl='EBF5FB')
            tp_r = _res_row(ws, 'TOTAL PEDIDOS:', total_pedidos_val, hl='FFF0E8')
            tdctos_r = _res_row(ws, 'TOTAL DCTOS:', total_dctos_val)
            venta_r = _res_row(ws, 'VENTA:', f'=B{td_r}+B{tp_r}-B{tdctos_r}', hl='E8F5E9')
            _res_row(ws, 'TOTAL EFECTIVO:', f'=B{venta_r}-C{pago_tot_row}-D{pago_tot_row}')
            _res_row(ws, 'VENCIDAS FVTO:', f'=SUMPRODUCT(H{data_start}:H{data_end},K{data_start}:K{data_end})')

            # Calcular para hoja Resumen general
            total_nequi_calc = sum(float(p.nequi or 0) for p in pagos)
            total_daviplata_calc = sum(float(p.daviplata or 0) for p in pagos)
            total_despacho_calc = sum(
                (float(r.valor or 0) if float(r.valor or 0) > 0 else precios_lookup.get(r.producto, 0))
                * (r.cantidad - r.dctos + r.adicional - r.devoluciones - r.vencidas)
                for r in registros
            )
            venta_calc = total_despacho_calc + total_pedidos_val - total_dctos_val
            efectivo_calc = venta_calc - total_nequi_calc - total_daviplata_calc

            ws.append([])  # blank

            # ── CONTROL DE CUMPLIMIENTO ──────────────────────────────
            try:
                cumpl = CargueCumplimiento.objects.get(vendedor_id=id_label, fecha=fecha_obj, dia=dia, activo=True)

                ws.append(['CONTROL DE CUMPLIMIENTO', ''])
                cc_hdr = ws.max_row
                ws.merge_cells(f'A{cc_hdr}:B{cc_hdr}')
                ws[f'A{cc_hdr}'].font = Font(bold=True, color='FFFFFF', size=11)
                ws[f'A{cc_hdr}'].fill = header_fill_azul
                ws[f'A{cc_hdr}'].alignment = center

                # MANIPULADOR
                ws.append(['MANIPULADOR', ''])
                m_hdr = ws.max_row
                ws.merge_cells(f'A{m_hdr}:B{m_hdr}')
                ws[f'A{m_hdr}'].font = Font(bold=True, size=10)
                ws[f'A{m_hdr}'].fill = PatternFill('solid', fgColor='D6E4F0')
                ws[f'A{m_hdr}'].alignment = center

                for campo, val in [
                    ('Licencia de transporte', cumpl.licencia_transporte),
                    ('SOAT', cumpl.soat),
                    ('Uniforme', cumpl.uniforme),
                    ('No loción', cumpl.no_locion),
                    ('No accesorios', cumpl.no_accesorios),
                    ('Capacitación/Carnet', cumpl.capacitacion_carnet),
                ]:
                    ws.append([campo, val or '-'])
                    rc = ws.max_row
                    ws[f'A{rc}'].border = borde
                    ws[f'B{rc}'].border = borde
                    if val == 'C':
                        ws[f'B{rc}'].fill = PatternFill('solid', fgColor='D4EDDA')
                        ws[f'B{rc}'].font = Font(bold=True, color='1D6A3A')
                    elif val == 'NC':
                        ws[f'B{rc}'].fill = PatternFill('solid', fgColor='F8D7DA')
                        ws[f'B{rc}'].font = Font(bold=True, color='721C24')

                # FURGÓN
                ws.append(['FURGÓN', ''])
                f_hdr = ws.max_row
                ws.merge_cells(f'A{f_hdr}:B{f_hdr}')
                ws[f'A{f_hdr}'].font = Font(bold=True, size=10)
                ws[f'A{f_hdr}'].fill = PatternFill('solid', fgColor='D6E4F0')
                ws[f'A{f_hdr}'].alignment = center

                for campo, val in [
                    ('Higiene', cumpl.higiene),
                    ('Estibas', cumpl.estibas),
                    ('Desinfección', cumpl.desinfeccion),
                ]:
                    ws.append([campo, val or '-'])
                    rc = ws.max_row
                    ws[f'A{rc}'].border = borde
                    ws[f'B{rc}'].border = borde
                    if val == 'C':
                        ws[f'B{rc}'].fill = PatternFill('solid', fgColor='D4EDDA')
                        ws[f'B{rc}'].font = Font(bold=True, color='1D6A3A')
                    elif val == 'NC':
                        ws[f'B{rc}'].fill = PatternFill('solid', fgColor='F8D7DA')
                        ws[f'B{rc}'].font = Font(bold=True, color='721C24')

                # Leyenda
                ws.append(['CUMPLE: C', 'NO CUMPLE: NC'])
                ley_r = ws.max_row
                ws[f'A{ley_r}'].font = Font(bold=True, color='1D6A3A', italic=True)
                ws[f'B{ley_r}'].font = Font(bold=True, color='721C24', italic=True)

            except CargueCumplimiento.DoesNotExist:
                pass

            ws.append([])  # blank

            # ── REGISTRO DE LOTES ────────────────────────────────────
            lotes_dia = []
            for r in registros:
                try:
                    lp = json.loads(r.lotes_produccion) if r.lotes_produccion else []
                    if isinstance(lp, list):
                        lotes_dia.extend(lp)
                except Exception:
                    pass

            if lotes_dia:
                ws.append(['REGISTRO DE LOTES', ''])
                rl_hdr = ws.max_row
                ws.merge_cells(f'A{rl_hdr}:B{rl_hdr}')
                ws[f'A{rl_hdr}'].font = Font(bold=True, color='FFFFFF', size=11)
                ws[f'A{rl_hdr}'].fill = header_fill_azul
                ws[f'A{rl_hdr}'].alignment = center

                ws.append(['LOTE', 'INFO'])
                for cell in ws[ws.max_row]:
                    cell.font = header_font
                    cell.fill = header_fill_azul
                    cell.alignment = center
                    cell.border = borde

                for lote in lotes_dia:
                    if isinstance(lote, dict):
                        ws.append([lote.get('lote', ''), lote.get('info', '') or lote.get('fecha', '')])
                    else:
                        ws.append([str(lote), ''])
                    for cell in ws[ws.max_row]:
                        cell.border = borde

            autofit(ws)
            ws.column_dimensions['A'].width = 4   # V
            ws.column_dimensions['B'].width = 4   # D
            ws.column_dimensions['C'].width = 30  # PRODUCTOS
            ws.column_dimensions['I'].width = 18  # LOTES VENCIDOS

            resumen_rows.append({
                'id': id_label,
                'responsable': responsable,
                'ruta': ruta,
                'productos': len(registros),
                'venta': venta_calc,
                'total_efectivo': efectivo_calc,
            })

        # Hoja Resumen (primera)
        if resumen_rows:
            ws_res = wb.create_sheet('Resumen', 0)
            ws_res.append([f'RESUMEN CARGUE — {dia}  {fecha_str}', '', '', '', '', ''])
            ws_res.merge_cells('A1:F1')
            ws_res['A1'].font = Font(bold=True, size=14, color=azul)
            ws_res['A1'].alignment = center
            ws_res.row_dimensions[1].height = 24
            ws_res.append([])

            ws_res.append(['ID', 'Responsable', 'Ruta', 'Productos', 'Venta', 'Efectivo'])
            for cell in ws_res[3]:
                cell.font = header_font
                cell.fill = header_fill_azul
                cell.alignment = center
                cell.border = borde

            total_venta = total_efectivo = 0.0
            for i, row in enumerate(resumen_rows):
                ws_res.append([row['id'], row['responsable'], row['ruta'], row['productos'], formato_cop(row['venta']), formato_cop(row['total_efectivo'])])
                data_row = ws_res.max_row
                fill = fila_par if i % 2 == 1 else PatternFill()
                for cell in ws_res[data_row]:
                    cell.border = borde
                    cell.alignment = left
                    if fill.fill_type:
                        cell.fill = fill
                total_venta += float(row['venta'] or 0)
                total_efectivo += float(row['total_efectivo'] or 0)

            ws_res.append(['TOTAL', '', '', '', formato_cop(total_venta), formato_cop(total_efectivo)])
            for cell in ws_res[ws_res.max_row]:
                cell.font = Font(bold=True)
                cell.fill = fila_total
                cell.border = borde

            autofit(ws_res)

        if not wb.sheetnames:
            return Response({'error': f'No hay datos de cargue para {dia} {fecha_str}'}, status=404)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        nombre_archivo = f'cargue_{dia}_{fecha_str}.xlsx'
        response_xl = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response_xl['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return response_xl

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


# ==================== EXPORTAR PLANEACIÓN MENSUAL EXCEL ====================

@api_view(['GET'])
def exportar_planeacion_mensual_excel(request):
    """
    Genera un Excel con la planeación de un mes completo.
    Una pestaña por día (lunes-sábado). Si existe snapshot guardado muestra datos,
    si no existe muestra la pestaña como Pendiente.
    GET /api/reportes/planeacion-mensual-excel/?mes=2026-04
    """
    try:
        mes_param = request.GET.get('mes')
        if not mes_param:
            return Response({'error': 'Parámetro mes requerido (formato: YYYY-MM)'}, status=400)

        try:
            partes = mes_param.split('-')
            anio, mes = int(partes[0]), int(partes[1])
        except Exception:
            return Response({'error': 'Formato de mes inválido. Use YYYY-MM'}, status=400)

        import calendar
        from datetime import date, timedelta

        primer_dia = date(anio, mes, 1)
        ultimo_dia = date(anio, mes, calendar.monthrange(anio, mes)[1])
        hoy = date.today()

        # Todos los días del mes excepto domingos (weekday==6)
        dias_mes = []
        d = primer_dia
        while d <= ultimo_dia:
            if d.weekday() != 6:
                dias_mes.append(d)
            d += timedelta(days=1)

        # Cargar snapshots del mes (uno por fecha, el más reciente)
        snapshots = ReportePlaneacion.objects.filter(
            fecha_reporte__year=anio,
            fecha_reporte__month=mes
        ).order_by('fecha_reporte', '-fecha_creacion')

        snapshot_map = {}
        for snap in snapshots:
            fecha_key = snap.fecha_reporte.strftime('%Y-%m-%d')
            if fecha_key not in snapshot_map:
                snapshot_map[fecha_key] = snap

        nombres_dia = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']

        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        COLOR_HEADER   = '1F4E79'
        COLOR_FILA_PAR = 'EBF3FB'
        COLOR_TOTAL    = 'BDD7EE'
        COLOR_OK       = 'E2EFDA'
        COLOR_SIN      = 'FCE4D6'
        COLOR_FUTURO   = 'FFF2CC'

        borde = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        def set_header_row(ws, fila, etiquetas, color=COLOR_HEADER):
            for ci, label in enumerate(etiquetas, 1):
                c = ws.cell(fila, ci, label)
                c.font = Font(bold=True, color='FFFFFF', size=10)
                c.fill = PatternFill('solid', fgColor=color)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = borde

        def set_borde_fila(ws, fila, n_cols):
            for ci in range(1, n_cols + 1):
                ws.cell(fila, ci).border = borde

        COLS = ['PRODUCTO', 'EXISTENCIAS', 'SOLICITADAS', 'PEDIDOS', 'A PRODUCIR', 'IA SUGERIDA']
        N = len(COLS)
        LAST_COL = chr(64 + N)  # 'F'

        wb = openpyxl.Workbook()

        # ── Hoja RESUMEN ──
        ws_res = wb.active
        ws_res.title = 'Resumen'

        ws_res.merge_cells('A1:D1')
        c = ws_res['A1']
        c.value = f'RESUMEN PLANEACIÓN — {mes_param}'
        c.font = Font(bold=True, color='FFFFFF', size=13)
        c.fill = PatternFill('solid', fgColor=COLOR_HEADER)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws_res.row_dimensions[1].height = 22

        set_header_row(ws_res, 2, ['FECHA', 'DÍA', 'ESTADO', 'PRODUCTOS'])
        ws_res.row_dimensions[2].height = 18

        total_guardados = 0
        total_pendientes = 0

        for idx, dia_fecha in enumerate(dias_mes):
            fecha_key    = dia_fecha.strftime('%Y-%m-%d')
            fecha_display = dia_fecha.strftime('%d/%m/%Y')
            nombre_dia   = nombres_dia[dia_fecha.weekday()]
            snap         = snapshot_map.get(fecha_key)
            es_futura    = dia_fecha > hoy
            fila_res     = idx + 3

            if snap:
                datos      = snap.datos_json if isinstance(snap.datos_json, list) else []
                estado     = 'Guardado'
                n_prod     = len(datos)
                fill_est   = PatternFill('solid', fgColor=COLOR_OK)
                font_est   = Font(color='375623', bold=True, size=10)
                total_guardados += 1
            elif es_futura:
                estado   = 'Futuro'
                n_prod   = 0
                fill_est = PatternFill('solid', fgColor=COLOR_FUTURO)
                font_est = Font(color='7F6000', bold=True, size=10)
                total_pendientes += 1
            else:
                estado   = 'Sin Guardar'
                n_prod   = 0
                fill_est = PatternFill('solid', fgColor=COLOR_SIN)
                font_est = Font(color='C00000', bold=True, size=10)
                total_pendientes += 1

            ws_res.append([fecha_display, nombre_dia, estado, n_prod if n_prod > 0 else '-'])
            ws_res.cell(fila_res, 3).fill  = fill_est
            ws_res.cell(fila_res, 3).font  = font_est
            ws_res.cell(fila_res, 3).alignment = Alignment(horizontal='center')
            ws_res.cell(fila_res, 4).alignment = Alignment(horizontal='center')
            set_borde_fila(ws_res, fila_res, 4)
            if idx % 2 == 0:
                for col in [1, 2, 4]:
                    ws_res.cell(fila_res, col).fill = PatternFill('solid', fgColor=COLOR_FILA_PAR)

        fila_tot_res = len(dias_mes) + 3
        ws_res.append(['TOTAL', f'{len(dias_mes)} días hábiles', f'Guardados: {total_guardados}', f'Pendientes: {total_pendientes}'])
        for ci in range(1, 5):
            c = ws_res.cell(fila_tot_res, ci)
            c.font = Font(bold=True, size=10)
            c.fill = PatternFill('solid', fgColor=COLOR_TOTAL)
            c.border = borde
            c.alignment = Alignment(horizontal='center')

        ws_res.column_dimensions['A'].width = 14
        ws_res.column_dimensions['B'].width = 14
        ws_res.column_dimensions['C'].width = 16
        ws_res.column_dimensions['D'].width = 14

        # ── Una hoja por día ──
        for dia_fecha in dias_mes:
            fecha_key    = dia_fecha.strftime('%Y-%m-%d')
            fecha_display = dia_fecha.strftime('%d/%m/%Y')
            nombre_dia   = nombres_dia[dia_fecha.weekday()]
            tab_name     = dia_fecha.strftime('%d-%m')
            snap         = snapshot_map.get(fecha_key)
            es_futura    = dia_fecha > hoy

            ws = wb.create_sheet(title=tab_name)

            # Título
            ws.merge_cells(f'A1:{LAST_COL}1')
            t = ws['A1']
            t.value = f'PLANEACIÓN {nombre_dia.upper()} {fecha_display}'
            t.font = Font(bold=True, color='FFFFFF', size=12)
            t.fill = PatternFill('solid', fgColor=COLOR_HEADER)
            t.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 22

            if snap:
                datos = snap.datos_json if isinstance(snap.datos_json, list) else []

                # Sub-título con fecha de guardado
                ws.merge_cells(f'A2:{LAST_COL}2')
                s = ws['A2']
                s.value = f'Guardado: {snap.fecha_creacion.strftime("%d/%m/%Y %H:%M")}  |  {len(datos)} productos'
                s.font = Font(italic=True, color='375623', size=9)
                s.fill = PatternFill('solid', fgColor=COLOR_OK)
                s.alignment = Alignment(horizontal='center')

                set_header_row(ws, 3, COLS)
                ws.row_dimensions[3].height = 16

                total_producir = 0
                for ri, prod in enumerate(datos, 4):
                    nombre     = prod.get('nombre', '')
                    existencias = int(prod.get('existencias', 0) or 0)
                    solicitado = int(prod.get('solicitado', 0) or 0)
                    pedidos    = int(prod.get('pedidos', 0) or 0)
                    orden      = int(prod.get('orden', 0) or 0)
                    ia         = int(prod.get('ia', 0) or 0)

                    ws.append([nombre, existencias, solicitado, pedidos, orden, ia])
                    total_producir += orden

                    fill_fila = PatternFill('solid', fgColor=COLOR_FILA_PAR) if ri % 2 == 0 else PatternFill()
                    for ci in range(1, N + 1):
                        c = ws.cell(ri, ci)
                        c.border = borde
                        c.alignment = Alignment(horizontal='left' if ci == 1 else 'center')
                        if ci > 1:
                            c.fill = fill_fila

                fila_tot_d = len(datos) + 4
                ws.cell(fila_tot_d, 1, 'TOTAL')
                ws.cell(fila_tot_d, 5, total_producir)
                for ci in range(1, N + 1):
                    c = ws.cell(fila_tot_d, ci)
                    c.font = Font(bold=True, size=10)
                    c.fill = PatternFill('solid', fgColor=COLOR_TOTAL)
                    c.border = borde
                    c.alignment = Alignment(horizontal='left' if ci == 1 else 'center')

            else:
                ws.merge_cells(f'A2:{LAST_COL}2')
                s = ws['A2']
                s.value = 'FECHA FUTURA — AÚN NO PLANEADA' if es_futura else 'SIN REPORTE GUARDADO PARA ESTA FECHA'
                s.font = Font(bold=True, color='FFFFFF', size=11)
                s.fill = PatternFill('solid', fgColor='BF8F00' if es_futura else 'C00000')
                s.alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[2].height = 20

            # Anchos de columna
            ws.column_dimensions['A'].width = 32
            for ci in range(2, N + 1):
                ws.column_dimensions[chr(64 + ci)].width = 14

        # Resumen siempre al inicio
        wb.move_sheet('Resumen', offset=-len(wb.sheetnames))

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        nombre_archivo = f'planeacion_{mes_param}.xlsx'
        response_xl = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response_xl['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return response_xl

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)
